"""Aplicativo web: Informe Financiero Mensual - Administracion Directa."""
import base64
import gzip
import hashlib
import hmac
import json
import logging
import shutil
import struct
import os
import secrets
import sqlite3
import threading
import time
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, session, abort, Response, send_file)
import io
from sqlalchemy import event, text
from sqlalchemy.engine import Engine


@event.listens_for(Engine, "connect")
def _activar_foreign_keys(dbapi_conn, _record):
    """SQLite no activa las claves foráneas por conexión: hay que hacerlo
    explícitamente para que el borrado en cascada respete la integridad."""
    cur = dbapi_conn.cursor()
    cur.execute("PRAGMA foreign_keys=ON")
    cur.close()
from werkzeug.security import generate_password_hash, check_password_hash

from models import (db, Proyecto, Presupuesto, Gasto, GastoDetalle,
                    AlmacenMovimiento, ActividadEjecutada, Trabajador, Usuario,
                    CLASIFICADORES, COMPONENTES, Suscripcion,
                    SuscripcionHistorial, LicenciaUtilizada, PLANES_SUSCRIPCION,
                    sumar_meses)
import version as _app_version
from helpers import (MESES, COMPONENTES_FE06, get_proyecto, get_suscripcion,
                     suscripcion_vigente, suscripcion_usuario, gastos_mes,
                     total_gastos_mes, kpis,
                     ejecucion_por_mes, ejecucion_por_componente, fe06_rows,
                     fe06_resumen, fe06_sintesis, f05_datos, almacen_items, almacen_diario,
                     saldo_insumo, almacen_valorizado, oc_para_material,
                     meses_con_ejecucion, meses_visibles, ampliacion_presupuestal,
                     mes_inicio_manifiesto, clasificadores_proyecto,
                     calendario_mes, resumen_tareo, panel_cuadro1, panel_datos,
                     actividades_mes,
                     presupuesto_filas, PRESUPUESTO_DETALLE, detalle_clasificador)
from planilla import (tabla_civil, calcular_obrero, calcular_tecnico,
                      TABLA_CIVIL_POR_ANIO)
from seed import seed
import databases as _bd
from databases import (tenant_path, tenant_engine, ensure_tenant,
                       tenant_session, dispose_tenant, init_databases, bind_session,
                       master_url, master_path, tablas_negocio, tablas_tenant,
                       tablas_maestras)

CARGOS_OBRERO = ["OPERARIO", "OFICIAL", "PEON", "GUARDIAN"]
CARGOS_TECNICO = ["RESIDENTE", "SUPERVISOR", "ADMINISTRADOR",
                  "ASISTENTE TECNICO", "ALMACENERO"]


def migrar_schema():
    """Agrega columnas nuevas a tablas existentes y migra los detalles de gasto."""
    with db.engine.connect() as conn:
        # Modo WAL: permite lecturas concurrentes con escritura (mejor para red local).
        conn.execute(text("PRAGMA journal_mode=WAL"))
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info(proyecto)"))]
        if not cols:
            # Maestra nueva o base sin tablas de negocio: no hay nada que migrar.
            return
        if "incluir_anios_anteriores" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN incluir_anios_anteriores BOOLEAN DEFAULT 1"))
            conn.commit()
        conn.execute(text(
            "UPDATE proyecto SET incluir_anios_anteriores = 1 "
            "WHERE incluir_anios_anteriores IS NULL"))
        conn.commit()

        for col in ["almacenero", "responsable_almacen", "administrador_obra"]:
            if col not in cols:
                conn.execute(text(
                    f"ALTER TABLE proyecto ADD COLUMN {col} VARCHAR(300) DEFAULT ''"))

        nuevos = {
            "fecha_inicio": "DATE",
            "fecha_fin": "DATE",
            "nuevo_final_obra": "DATE",
            "fecha_aprobacion": "DATE",
            "dias_ejecucion": "INTEGER DEFAULT 0",
            "dias_ampliacion": "INTEGER DEFAULT 0",
            "n_resolucion_adicional": "VARCHAR(200) DEFAULT ''",
            "monto_ampliacion": "FLOAT DEFAULT 0",
            "adicional_obra": "BOOLEAN DEFAULT 0",
            "adicionales": "TEXT DEFAULT ''",
            "ampliacion_presupuestal": "BOOLEAN DEFAULT 0",
        }
        for col, tipo in nuevos.items():
            if col not in cols:
                conn.execute(text(
                    f"ALTER TABLE proyecto ADD COLUMN {col} {tipo}"))
        conn.commit()

        for col in ["clasificador_personal", "clasificador_bienes",
                    "clasificador_servicios", "clasificador_expediente",
                    "clasificador_liquidacion"]:
            if col not in cols:
                conn.execute(text(
                    f"ALTER TABLE proyecto ADD COLUMN {col} VARCHAR(20) DEFAULT ''"))
        if "logo_path" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN logo_path VARCHAR(500) DEFAULT ''"))
        for cip_col in ["cip_supervisor", "cip_residente"]:
            if cip_col not in cols:
                conn.execute(text(
                    f"ALTER TABLE proyecto ADD COLUMN {cip_col} VARCHAR(50) DEFAULT ''"))
        for nuevo_col, tipo in {
            "colegiatura_admin": "VARCHAR(50) DEFAULT ''",
            "dni_responsable_almacen": "VARCHAR(50) DEFAULT ''",
            "asistente_tecnico": "VARCHAR(300) DEFAULT ''",
            "dni_cip_asistente": "VARCHAR(50) DEFAULT ''",
        }.items():
            if nuevo_col not in cols:
                conn.execute(text(
                    f"ALTER TABLE proyecto ADD COLUMN {nuevo_col} {tipo}"))
        conn.execute(text(
            "UPDATE proyecto SET clasificador_personal = '2.6.2.3.99.3' "
            "WHERE clasificador_personal IS NULL OR clasificador_personal = ''"))
        conn.execute(text(
            "UPDATE proyecto SET clasificador_bienes = '2.6.2.3.99.4' "
            "WHERE clasificador_bienes IS NULL OR clasificador_bienes = ''"))
        conn.execute(text(
            "UPDATE proyecto SET clasificador_servicios = '2.6.2.3.99.5' "
            "WHERE clasificador_servicios IS NULL OR clasificador_servicios = ''"))
        conn.execute(text(
            "UPDATE proyecto SET clasificador_expediente = '2.6.8.1.3.1' "
            "WHERE clasificador_expediente IS NULL OR clasificador_expediente = ''"))
        conn.execute(text(
            "UPDATE proyecto SET clasificador_liquidacion = 'LIQUIDACION' "
            "WHERE clasificador_liquidacion IS NULL OR clasificador_liquidacion = ''"))
        conn.commit()

        # Migracion: num_anios_anteriores y metas por año
        if "num_anios_anteriores" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN num_anios_anteriores INTEGER DEFAULT 3"))
        if "meta_ejec2023" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN meta_ejec2023 FLOAT DEFAULT 0"))
        if "meta_ejec2024" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN meta_ejec2024 FLOAT DEFAULT 0"))
        if "meta_ejec2025" not in cols:
            conn.execute(text(
                "ALTER TABLE proyecto ADD COLUMN meta_ejec2025 FLOAT DEFAULT 0"))
        conn.commit()

        # Migración de datos: renombra componente anterior
        conn.execute(text(
            "UPDATE presupuesto SET componente = 'Gastos de Supervisión' "
            "WHERE componente = 'Gestion de Supervisión'"))
        conn.execute(text(
            "UPDATE gasto SET componente = 'Gastos de Supervisión' "
            "WHERE componente = 'Gestion de Supervisión'"))
        conn.commit()

        # Migracion: devengado en gastos. Los registros existentes se consideran
        # ya devengados para no perder el contenido del informe actual.
        gcols = [r[1] for r in conn.execute(text("PRAGMA table_info(gasto)"))]
        if "devengado" not in gcols:
            conn.execute(text(
                "ALTER TABLE gasto ADD COLUMN devengado BOOLEAN DEFAULT 0"))
            conn.execute(text("UPDATE gasto SET devengado = 1"))
            conn.commit()
        if "nota_pago" not in gcols:
            conn.execute(text(
                "ALTER TABLE gasto ADD COLUMN nota_pago VARCHAR(100) DEFAULT ''"))
            conn.commit()
        if "fecha_devengado" not in gcols:
            conn.execute(text(
                "ALTER TABLE gasto ADD COLUMN fecha_devengado DATE"))
            conn.commit()

        # Migracion: devengado por trabajador. El estado de devengado de la
        # planilla lo marca el gasto PLLA del panel (ver mas abajo).
        tcols = [r[1] for r in conn.execute(text("PRAGMA table_info(trabajador)"))]
        if "devengado" not in tcols:
            conn.execute(text(
                "ALTER TABLE trabajador ADD COLUMN devengado BOOLEAN DEFAULT 0"))
            conn.commit()
        # Campo "Sueldo mensual" se reincorpora para la planilla de pagos del
        # personal tecnico/administrativo (D.L. 728).
        if "sueldo_mensual" not in tcols:
            conn.execute(text(
                "ALTER TABLE trabajador ADD COLUMN sueldo_mensual REAL DEFAULT 0.0"))
            conn.commit()
            tcols = [r[1] for r in conn.execute(text("PRAGMA table_info(trabajador)"))]
        if "aporte" not in tcols:
            conn.execute(text(
                "ALTER TABLE trabajador ADD COLUMN aporte VARCHAR(10) DEFAULT 'AFP'"))
            conn.execute(text("UPDATE trabajador SET aporte = 'AFP' WHERE aporte IS NULL"))
            conn.commit()
        # El estado de devengado de la planilla lo marca el gasto PLLA del panel:
        # se limpian los flags heredados del flujo anterior (casilla DEVENGADO).
        if "devengado" in tcols:
            conn.execute(text("UPDATE trabajador SET devengado = 0"))
            conn.commit()
        conn.execute(text(
            "UPDATE proyecto SET almacenero = asistente "
            "WHERE (almacenero IS NULL OR almacenero = '') AND asistente IS NOT NULL"))
        conn.execute(text(
            "UPDATE proyecto SET responsable_almacen = asistente "
            "WHERE (responsable_almacen IS NULL OR responsable_almacen = '') "
            "AND asistente IS NOT NULL"))
        conn.execute(text(
            "UPDATE proyecto SET administrador_obra = asistente "
            "WHERE (administrador_obra IS NULL OR administrador_obra = '') "
            "AND asistente IS NOT NULL"))
        conn.commit()

        # Migracion: cada gasto antiguo (con detalle en la misma fila) pasa a
        # tener un GastoDetalle para no perder la informacion registrada.
        gcols = [r[1] for r in conn.execute(text("PRAGMA table_info(gasto)"))]
        if "detalle" in gcols:
            filas = conn.execute(text(
                "SELECT id, detalle, und, cantidad, precio_unitario FROM gasto "
                "WHERE detalle IS NOT NULL AND detalle != ''")).fetchall()
            for gid, detalle, und, cantidad, pu in filas:
                ex = conn.execute(text(
                    "SELECT COUNT(*) FROM gasto_detalle WHERE gasto_id = :i"),
                    {"i": gid}).scalar()
                if ex == 0:
                    conn.execute(text(
                        "INSERT INTO gasto_detalle "
                        "(gasto_id, detalle, und, cantidad, precio_unitario, orden) "
                        "VALUES (:i, :d, :u, :c, :p, 1)"),
                        {"i": gid, "d": detalle or "", "u": und or "UND",
                         "c": cantidad or 1, "p": pu or 0})
                conn.commit()

        # Tabla de usuarios del aplicativo + usuario administrador inicial.
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS usuario ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "usuario VARCHAR(80) NOT NULL UNIQUE,"
            "clave VARCHAR(255) NOT NULL,"
            "nombres VARCHAR(200) DEFAULT '',"
            "rol VARCHAR(30) DEFAULT 'Usuario',"
            "activo BOOLEAN DEFAULT 1)"))
        conn.commit()
        count = conn.execute(text(
            "SELECT COUNT(*) FROM usuario")).scalar()
        if count == 0:
            conn.execute(text(
                "INSERT INTO usuario (usuario, clave, nombres, rol, activo) "
                "VALUES ('admin', :clave, 'Administrador', 'Administrador', 1)"),
                {"clave": generate_password_hash("admin")})
            conn.commit()

        # Permisos por usuario (claves de secciones autorizadas, formato JSON).
        ucols = [r[1] for r in conn.execute(text("PRAGMA table_info(usuario)"))]
        if "permisos" not in ucols:
            conn.execute(text(
                "ALTER TABLE usuario ADD COLUMN permisos VARCHAR(500) DEFAULT '[]'"))
            conn.commit()

        # Tabla de tareo o planilla (personal obrero y tecnico).
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS trabajador ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "tipo VARCHAR(10) DEFAULT 'OBRERO',"
            "nombre VARCHAR(300) DEFAULT '',"
            "dni VARCHAR(8) DEFAULT '',"
            "fecha_nacimiento DATE,"
            "cargo VARCHAR(200) DEFAULT '',"
            "sexo VARCHAR(1) DEFAULT 'M',"
            "fecha_inicio DATE,"
            "dias VARCHAR(100) DEFAULT '',"
            "mes INTEGER DEFAULT 6,"
            "anio INTEGER DEFAULT 2026)"))
        conn.commit()
        tcols = [r[1] for r in conn.execute(text("PRAGMA table_info(trabajador)"))]
        if "dias" not in tcols:
            conn.execute(text(
                "ALTER TABLE trabajador ADD COLUMN dias VARCHAR(100) DEFAULT ''"))
            conn.commit()
        # Migracion: dias trabajados numericos -> dias CSV. Se asume que el
        # registro fue de un mes en curso y los dias se marcan por secuencia.
        if "dias_trabajados" in tcols and "dias" in tcols:
            filas = conn.execute(text(
                "SELECT id, dias_trabajados FROM trabajador "
                "WHERE dias_trabajados > 0 AND (dias IS NULL OR dias = '')")).fetchall()
            for tid, n in filas:
                if n and n > 0:
                    conn.execute(text(
                        "UPDATE trabajador SET dias = :dias WHERE id = :id"),
                        {"dias": ",".join(str(d) for d in range(1, int(n) + 1)),
                         "id": tid})
            conn.commit()

        # Numero de pecosa o guia de remision en los movimientos de almacen.
        acols = [r[1] for r in conn.execute(text("PRAGMA table_info(almacen_movimiento)"))]
        if "pecosa_guia" not in acols:
            conn.execute(text(
                "ALTER TABLE almacen_movimiento ADD COLUMN pecosa_guia VARCHAR(50) DEFAULT ''"))
            conn.commit()
        if "numero_siaf" not in acols:
            conn.execute(text(
                "ALTER TABLE almacen_movimiento ADD COLUMN numero_siaf VARCHAR(50) DEFAULT ''"))
            conn.commit()

        # Configuracion presupuestal: filas PERSONAL para Gastos Generales y
        # Gastos de Supervisión (se agregan si aun no existen).
        pcols = [r[1] for r in conn.execute(text("PRAGMA table_info(presupuesto)"))]
        if pcols:
            personal = conn.execute(text(
                "SELECT clasificador_personal FROM proyecto LIMIT 1")).scalar()
            cod_personal = personal or "2.6.2.3.99.3"
            for comp in ["Gastos Generales", "Gastos de Supervisión"]:
                ex = conn.execute(text(
                    "SELECT COUNT(*) FROM presupuesto "
                    "WHERE componente = :c AND detalle = 'PERSONAL'"),
                    {"c": comp}).scalar()
                if ex == 0:
                    conn.execute(text(
                        "INSERT INTO presupuesto "
                        "(componente, clasificador, detalle, et, ejec2023, "
                        "ejec2024, ejec2025, pim2026) "
                        "VALUES (:c, :cl, 'PERSONAL', 0, 0, 0, 0, 0)"),
                        {"c": comp, "cl": cod_personal})
            conn.commit()


def migrar_suscripcion():
    """Crea las tablas de suscripción y la cuenta inicial del Super Usuario."""
    with db.engine.connect() as conn:
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS suscripcion ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "plan VARCHAR(20) DEFAULT 'Mensual',"
            "fecha_inicio DATE,"
            "fecha_fin DATE,"
            "activa BOOLEAN DEFAULT 1)"))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS suscripcion_historial ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "plan VARCHAR(20) DEFAULT '',"
            "fecha_inicio DATE,"
            "fecha_fin DATE,"
            "usuario VARCHAR(120) DEFAULT '',"
            "nota VARCHAR(500) DEFAULT '',"
            "accion VARCHAR(30) DEFAULT 'Renovación',"
            "fecha_registro DATETIME)"))
        conn.commit()
        # Suscripción inicial activa (1 mes) para que el aplicativo siga
        # funcionando desde el primer arranque.
        n = conn.execute(text("SELECT COUNT(*) FROM suscripcion")).scalar()
        if n == 0:
            hoy = date.today()
            conn.execute(text(
                "INSERT INTO suscripcion (plan, fecha_inicio, fecha_fin, activa) "
                "VALUES ('Mensual', :i, :f, 1)"),
                {"i": hoy, "f": sumar_meses(hoy, 1)})
            conn.commit()
        # Cuenta inicial del Super Usuario (controla la suscripción).
        scount = conn.execute(text(
            "SELECT COUNT(*) FROM usuario WHERE rol = 'Super Usuario'")).scalar()
        existe_nombre = conn.execute(text(
            "SELECT COUNT(*) FROM usuario WHERE usuario = 'super'")).scalar()
        if scount == 0 and existe_nombre == 0:
            conn.execute(text(
                "INSERT INTO usuario (usuario, clave, nombres, rol, activo, permisos) "
                "VALUES ('super', :clave, 'Cuenta Principal', 'Super Usuario', 1, :permisos)"),
                {"clave": generate_password_hash("1989@#John"),
                 "permisos": json.dumps([c for c, _ in PERMISOS_SECCIONES])})
            conn.commit()
        # El nombre mostrado de la cuenta principal no debe revelar el rol.
        conn.execute(text(
            "UPDATE usuario SET nombres = 'Cuenta Principal' "
            "WHERE rol = 'Super Usuario' AND usuario = 'super' "
            "AND nombres IN ('Super Usuario', 'Cuenta Principal')"))
        conn.commit()
        # Licencia propia por Administrador: si susc_activa es NULL el
        # Administrador hereda la licencia global; False = sin licencia.
        ucols = [r[1] for r in conn.execute(text("PRAGMA table_info(usuario)"))]
        if "susc_activa" not in ucols:
            conn.execute(text("ALTER TABLE usuario ADD COLUMN susc_plan VARCHAR(20)"))
            conn.execute(text("ALTER TABLE usuario ADD COLUMN susc_inicio DATE"))
            conn.execute(text("ALTER TABLE usuario ADD COLUMN susc_fin DATE"))
            conn.execute(text("ALTER TABLE usuario ADD COLUMN susc_activa BOOLEAN"))
            conn.commit()


def _migrar_a_multidb():
    """Convierte una base única antigua al nuevo esquema (una sola vez).

    Si la base maestra todavía contiene datos de negocio (proyecto), los
    traslada a la base del primer Administrador y mueve allí los operadores,
    dejando la maestra solo con cuentas y suscripción. Es idempotente: al no
    quedar datos de negocio en la maestra no vuelve a ejecutarse.
    """
    origen = master_path()
    if not os.path.exists(origen):
        return
    src = sqlite3.connect(origen)
    try:
        try:
            n = src.execute("SELECT COUNT(*) FROM proyecto").fetchone()[0]
        except sqlite3.OperationalError:
            return
        if not n:
            return
        fila = src.execute(
            "SELECT id FROM usuario WHERE rol = 'Administrador' "
            "ORDER BY id LIMIT 1").fetchone()
        if not fila:
            return
        admin_id = fila[0]
        destino = tenant_path(admin_id)
        ensure_tenant(admin_id)
        existentes = {r[0] for r in src.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table'")}
        src.execute("ATTACH DATABASE ? AS tenant", (destino,))
        for tabla in tablas_negocio():
            if tabla not in existentes:
                continue
            # Si el tenant ya tiene datos (migración previa), no duplicar.
            if src.execute(
                    f"SELECT COUNT(*) FROM tenant.{tabla}").fetchone()[0]:
                continue
            cols = ", ".join(c.name for c in db.metadata.tables[tabla].columns)
            src.execute(
                f"INSERT INTO tenant.{tabla} ({cols}) SELECT {cols} "
                f"FROM main.{tabla}")
        ucols = ", ".join(c.name for c in db.metadata.tables["usuario"].columns)
        src.execute(
            f"INSERT INTO tenant.usuario ({ucols}) SELECT {ucols} "
            f"FROM main.usuario WHERE rol = 'Usuario' "
            f"AND usuario NOT IN (SELECT usuario FROM tenant.usuario)")
        src.execute("DELETE FROM main.usuario WHERE rol = 'Usuario'")
        for tabla in tablas_negocio():
            src.execute(f"DROP TABLE IF EXISTS main.{tabla}")
        src.commit()
    finally:
        src.close()


def _asegurar_tenants():
    """Crea las bases (vacías) de todos los Administradores que falten."""
    for adm in _bd.master_session.query(Usuario).filter(
            Usuario.rol == "Administrador").all():
        ensure_tenant(adm.id)


def _seed_inicial():
    """Semilla de datos de ejemplo (deshabilitada para instaladores).

    La función original creaba datos de demostración (Proyecto, Presupuesto,
    Gastos, Almacén) para el primer Administrador.  Se desactiva para que
    las instalaciones nuevas arranquen con la BD vacía y solo el Super
    Usuario; el usuario crea sus datos desde la interfaz.
    """
    return


def create_app():
    base = os.path.abspath(os.path.dirname(__file__))
    instance = os.path.join(base, "instance")
    os.makedirs(instance, exist_ok=True)

    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.environ.get(
        "SECRET_KEY", "informe-financiero-toraya-2026")
    # Ruta de la base de datos (INFORME_DB permite usar otra DB, p.ej. en pruebas).
    app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
        "INFORME_DB") or "sqlite:///" + os.path.join(instance, "informe.db")
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    # Mitigacion CSRF a nivel de cookie: las peticiones cross-site no incluyen la cookie.
    app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
    app.config["SESSION_COOKIE_HTTPONLY"] = True
    # Flask 3.1 usa por defecto SEND_FILE_MAX_AGE_DEFAULT=None, lo que sirve los
    # estaticos con "Cache-Control: no-cache" y fuerza re-descarga en cada pagina.
    # Cache de 30 dias: css/js usan versionado (?v=N) y el logo tambien, asi que
    # el navegador solo vuelve a bajar lo que realmente cambia.
    app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 2592000
    app.jinja_env.finalize = lambda v: "" if v is None else v

    # Version del aplicativo (para el check de actualizaciones).
    app.config["INFORME_VERSION"] = os.environ.get(
        "INFORME_VERSION") or _app_version.__version__
    # Repositorio GitHub para el chequeo de actualizaciones.
    app.config["INFORME_REPO"] = "jacj01/informe-mensual-obra"
    # Token GitHub (PAT con permisos minimos de lectura de releases).
    # Se usa solo como fallback si gh CLI no esta autenticado.
    # Se puede overridear con la variable de entorno INFORME_GH_TOKEN.
    app.config["INFORME_GH_TOKEN"] = os.environ.get("INFORME_GH_TOKEN", "")

    @app.context_processor
    def _csrf_contexto():
        """Genera y expone el token CSRF por sesión (defensa en profundidad)."""
        token = session.get("_csrf")
        if not token:
            token = secrets.token_hex(16)
            session["_csrf"] = token
        return {"csrf_token": token,
                "version_actual": app.config.get("INFORME_VERSION", "1.0.0"),
                "es_admin_actual": es_admin_actual,
                "es_super_usuario": es_super_usuario}

    @app.after_request
    def cabeceras_seguridad(resp):
        """Headers de seguridad en todas las respuestas."""
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        resp.headers.setdefault("Referrer-Policy", "same-origin")
        resp.headers.setdefault("Content-Security-Policy",
            "default-src 'self'; script-src 'self' 'unsafe-inline'; "
            "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
            "img-src 'self' data:; "
            "font-src 'self' data: https://fonts.gstatic.com; connect-src 'self'")
        return resp

    _GZIP_TIPOS = {"text/html", "text/css", "text/plain", "application/javascript",
                   "text/javascript", "application/json", "application/xml",
                   "application/x-javascript", "image/svg+xml"}

    @app.after_request
    def _comprimir_respuesta(resp):
        """Comprime con gzip las respuestas de texto cuando el navegador lo
        acepta, para reducir los bytes transferidos en cada navegación.
        Los binarios (png, woff2...) no se tocan."""
        if resp.status_code in (204, 304):
            return resp
        if resp.mimetype not in _GZIP_TIPOS:
            return resp
        if "gzip" not in request.headers.get("Accept-Encoding", ""):
            return resp
        try:
            if request.path.startswith("/static/"):
                base = os.path.realpath(app.static_folder)
                ruta = os.path.realpath(
                    os.path.join(base, request.path[len("/static/"):].lstrip("/")))
                if not ruta.startswith(base) or not os.path.isfile(ruta):
                    return resp
                with open(ruta, "rb") as fh:
                    data = fh.read()
            else:
                data = resp.get_data()
        except Exception:
            return resp
        if not data or len(data) < 600:
            return resp
        try:
            comp = gzip.compress(data, 6)
        except Exception:
            return resp
        if len(comp) >= len(data):
            return resp
        resp.set_data(comp)
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Vary"] = "Accept-Encoding"
        resp.headers["Content-Length"] = str(len(comp))
        return resp

    @app.before_request
    def validar_origen():
        """Defensa adicional contra CSRF: rechaza POST de origen externo
        cuando el navegador envía Origin/Referer que no coincide con el host."""
        if request.method == "POST":
            host = request.host
            origen = request.headers.get("Origin") or request.headers.get("Referer")
            if origen:
                from urllib.parse import urlsplit
                try:
                    ohost = urlsplit(origen).netloc
                except ValueError:
                    ohost = ""
                if ohost and ohost != host:
                    return ("Origen no autorizado.", 403)

    @app.before_request
    def validar_csrf():
        """Valida el token CSRF por sesión en toda petición POST."""
        if request.method != "POST":
            return
        token = session.get("_csrf")
        if not token:
            return ("Token CSRF no generado. Recargue la página.", 403)
        enviado = request.form.get("_csrf", "") or \
            request.headers.get("X-CSRF-Token", "")
        if not enviado or not secrets.compare_digest(enviado, token):
            return ("Token CSRF no válido. Recargue la página e intente de nuevo.", 403)

    db.init_app(app)
    with app.app_context():
        init_databases()
        # La maestra solo guarda cuentas y suscripción; las tablas de negocio
        # viven en la base de cada Administrador (se crean vía ensure_tenant).
        db.metadata.create_all(bind=db.engine, tables=tablas_maestras())
        migrar_schema()
        migrar_suscripcion()
        _migrar_a_multidb()
        _asegurar_tenants()
        _seed_inicial()
        bind_session(_bd.master_engine)
        # El respaldo diario corre en un hilo de fondo: si las bases son grandes
        # o el disco es lento (p.ej. carpetas sincronizadas), copiarlas en el
        # arranque retrasa la primera carga. El servidor queda listo de inmediato.
        threading.Thread(target=_respaldo_diario_automatico,
                         daemon=True, name="respaldo-diario").start()

    @app.before_request
    def _enlazar_base_por_peticion():
        """Apunta la sesión ORM a la base del administrador de la sesión
        (o a la base maestra cuando no hay base asignada).

        Los archivos estáticos no necesitan base de datos: se saltean para que
        Flask no acceda a la cookie de sesión (evita Vary: Cookie y permite que
        el navegador los cachee por 30 días)."""
        if request.endpoint == "static":
            return None
        tid = session.get("tenant")
        bind_session(tenant_engine(tid) if tid else _bd.master_engine)

    registrar_rutas(app)
    return app


def parse_fecha(val):
    if not val:
        return None
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError:
        return None


def fecha_defecto(p):
    """Fecha por defecto segun el mes/anio configurado en Datos del Proyecto."""
    try:
        return date(p.anio, p.mes_actual, 1)
    except (TypeError, ValueError):
        return date.today()


def param_int(name, default, lo=None, hi=None):
    """Lee un parametro de la query string como entero con fallback seguro."""
    raw = request.args.get(name)
    if raw in (None, ""):
        return default
    try:
        val = int(raw)
    except (TypeError, ValueError):
        return default
    if lo is not None and val < lo:
        return default
    if hi is not None and val > hi:
        return default
    return val


def usuario_actual():
    """Devuelve el Usuario de la sesión activa o None.

    La cuenta de un Administrador vive en la base maestra, mientras que la de
    un operador vive en la base de su Administrador (la que está enlazada en
    esta petición).
    """
    uid = session.get("usuario_id")
    if not uid:
        return None
    if session.get("usuario_rol") == "Administrador":
        return _bd.master_session.get(Usuario, uid)
    return db.session.get(Usuario, uid)


def _dueno_suscripcion(u):
    """Cuenta cuya licencia rige al usuario actual.

    Un operador (rol Usuario) hereda la licencia de su Administrador; el
    Administrador y la cuenta principal usan la suya.
    """
    if u is None or getattr(u, "rol", None) != "Usuario":
        return u
    tid = session.get("tenant")
    if tid:
        return _bd.master_session.get(Usuario, tid)
    return u


def _buscar_operador(admin_id, nombre):
    """Busca un operador (rol Usuario) en la base del Administrador indicado."""
    ensure_tenant(admin_id)
    S = tenant_session(admin_id)
    try:
        return S.query(Usuario).filter(Usuario.usuario == nombre).first()
    finally:
        S.close()


def _es_cuenta_propia(u):
    """True si u es la cuenta con la que el usuario ingresó.

    Compara id y rol porque las bases de los administradores y la maestra
    pueden coincidir en números de id (cada base tiene su propio contador).
    """
    return bool(u and u.id == session.get("usuario_id")
                and u.rol == session.get("usuario_rol"))


PERMISOS_SECCIONES = [
    ("cabecera", "Cabecera del Proyecto"),
    ("ordenes", "Ingreso de O/C o O/S"),
    ("gastos", "Manifiesto de Gastos"),
    ("formatos", "Formatos Financiero"),
    ("almacen", "Almacén de Obra"),
    ("tareo", "Tareo y Planilla"),
    ("configuracion", "Configuración Presupuestal"),
    ("respaldo", "Respaldo de Datos"),
]

# Rol de máxima prioridad: controla la suscripción y puede modificar todo,
# incluidos los administradores, sin las restricciones habituales.
ROL_SUPER = "Super Usuario"
ROLES_TOTALES = ("Administrador", ROL_SUPER)


def es_super_actual():
    """True si el usuario de la sesión activa es el Super Usuario."""
    u = usuario_actual()
    return bool(u and u.rol == ROL_SUPER)


def es_admin_actual():
    """True si el usuario activo puede administrar (Administrador o Super)."""
    u = usuario_actual()
    if not u:
        return False
    return u.rol in ROLES_TOTALES


def es_super_usuario():
    """True SOLO para el Super Usuario (rol exclusivo). Usado para operaciones
    de publicacion/release, que requieren el maximo de privilegios."""
    u = usuario_actual()
    if not u:
        return False
    return u.rol == ROL_SUPER


# ----------------------------------------------------------------------------
# Licencia de suscripción: archivo cifrado (SHA-256) que el Super Usuario
# genera y entrega al cliente. El cliente lo sube para renovar automáticamente.
# El contenido no es legible en un editor de texto y cualquier modificación
# invalida la integridad (HMAC-SHA256), por lo que la licencia no puede
# alterarse sin conocer la clave del emisor.
# ----------------------------------------------------------------------------
LICENCIA_SECRETO = "informe-mensual-obra::licencia::v1::2026"
LICENCIA_EMISOR = "Informe Mensual de Obra"
_MAGICO_LICENCIA = b"IML1:"


def _clave_licencia():
    """Clave de cifrado derivada de la semilla mediante SHA-256 (32 bytes)."""
    return hashlib.sha256(LICENCIA_SECRETO.encode("utf-8")).digest()


def _keystream(clave, nonce, largo):
    """Keystream pseudoaleatorio SHA-256 en modo CTR (bloques de 32 bytes)."""
    out = bytearray()
    contador = 0
    while len(out) < largo:
        out.extend(hashlib.sha256(
            clave + nonce + struct.pack(">I", contador)).digest())
        contador += 1
    return bytes(out[:largo])


def cifrar_licencia(contenido):
    """Cifra bytes con XOR(keystream SHA-256) y sella con HMAC-SHA256.
    Devuelve bytes con prefijo 'IML1:' (contenido no legible)."""
    clave = _clave_licencia()
    nonce = secrets.token_bytes(16)
    ks = _keystream(clave, nonce, len(contenido))
    cifrado = bytes(a ^ b for a, b in zip(contenido, ks))
    tag = hmac.new(clave, nonce + cifrado, hashlib.sha256).digest()
    return _MAGICO_LICENCIA + base64.urlsafe_b64encode(nonce + cifrado + tag)


def descifrar_licencia(blob):
    """Devuelve el dict de la licencia o None si el archivo es inválido,
    está alterado o no fue generado por el emisor (integridad HMAC)."""
    try:
        if not isinstance(blob, (bytes, bytearray)) or not blob.startswith(_MAGICO_LICENCIA):
            return None
        datos = base64.urlsafe_b64decode(blob[len(_MAGICO_LICENCIA):])
        nonce = datos[:16]
        tag = datos[-32:]
        cifrado = datos[16:-32]
        clave = _clave_licencia()
        esperado = hmac.new(clave, nonce + cifrado, hashlib.sha256).digest()
        if not hmac.compare_digest(esperado, tag):
            return None
        ks = _keystream(clave, nonce, len(cifrado))
        plano = bytes(a ^ b for a, b in zip(cifrado, ks))
        d = json.loads(plano.decode("utf-8"))
        return d if isinstance(d, dict) else None
    except Exception:
        return None


def _licencia_valida(d):
    """True si el dict descifrado tiene datos de plan/meses correctos."""
    if not isinstance(d, dict):
        return False
    plan = d.get("plan")
    if plan not in PLANES_SUSCRIPCION:
        return False
    # Toda licencia debe indicar a qué usuario está vinculada y su serie.
    if not d.get("usuario") or not d.get("serie"):
        return False
    try:
        meses = int(d.get("meses", PLANES_SUSCRIPCION[plan]))
    except (TypeError, ValueError):
        return False
    return 1 <= meses <= 60


def generar_licencia(plan, usuario="", usuario_id=None, nombres=""):
    """Devuelve el contenido cifrado (str) de un archivo de licencia.

    La licencia queda vinculada al usuario indicado: solo la cuenta de ese
    usuario podrá activarla y solo una vez.
    """
    if plan not in PLANES_SUSCRIPCION:
        plan = "Mensual"
    d = {
        "version": 1,
        "emisor": LICENCIA_EMISOR,
        "plan": plan,
        "meses": PLANES_SUSCRIPCION[plan],
        "emitida": date.today().isoformat(),
        "serie": secrets.token_hex(4).upper(),
        "usuario_id": int(usuario_id) if usuario_id else None,
        "usuario": (usuario or "").strip(),
        "nombres": (nombres or "").strip(),
    }
    return cifrar_licencia(
        json.dumps(d, ensure_ascii=False).encode("utf-8")).decode("ascii")


def aplicar_licencia(d):
    """Aplica una licencia validada a la suscripción del Administrador titular.

    La licencia activa la suscripción propia del Administrador que la aplica
    (encadenando su vigencia); la cuenta principal renovaría la suscripción
    global. Devuelve un dict con plan/fechas o None si los datos son inválidos,
    si la licencia ya fue utilizada o si la aplica un usuario distinto al
    titular.
    """
    if not _licencia_valida(d):
        return None
    plan = d.get("plan", "Mensual")
    if plan not in PLANES_SUSCRIPCION:
        return None
    meses = int(d.get("meses", PLANES_SUSCRIPCION[plan]))
    serie = (d.get("serie") or "").strip()
    actor = usuario_actual()
    # La licencia solo puede activarla la cuenta del usuario titular.
    if (not actor or not actor.usuario
            or actor.usuario.strip() != (d.get("usuario") or "").strip()):
        return None
    ms = _bd.master_session
    # Uso único: la serie de la licencia no puede haberse activado antes.
    if serie and ms.query(LicenciaUtilizada).filter_by(serie=serie).first():
        return None
    hoy = date.today()
    es_admin = actor.rol == "Administrador"
    # Si la vigencia actual sigue activa se encadena desde su fin; si venció o
    # no hay licencia, se inicia desde hoy.
    if es_admin:
        base = actor.susc_fin if (actor.susc_activa and actor.susc_fin
                                  and actor.susc_fin >= hoy) else hoy
    else:
        s = get_suscripcion()
        base = s.fecha_fin if (s and s.fecha_fin and s.fecha_fin >= hoy) else hoy
    fin = sumar_meses(base, meses)
    if es_admin:
        actor.susc_plan = plan
        actor.susc_inicio = base
        actor.susc_fin = fin
        actor.susc_activa = True
    else:
        s = get_suscripcion()
        if s is None:
            s = Suscripcion(plan="Mensual")
            ms.add(s)
        s.plan = plan
        s.fecha_inicio = base
        s.fecha_fin = fin
        s.activa = True
    ms.add(SuscripcionHistorial(
        plan=plan, fecha_inicio=base, fecha_fin=fin,
        usuario=(actor.nombres or actor.usuario) if actor else "",
        nota=f"Renovación por licencia (serie {serie})",
        accion="Renovación"))
    ms.add(LicenciaUtilizada(
        serie=serie or "SIN-SERIE",
        usuario=(actor.usuario or "").strip(),
        plan=plan))
    ms.commit()
    return {"plan": plan, "fecha_inicio": base, "fecha_fin": fin}

# Endpoints a los que solo accede el Administrador.
_ENDPOINTS_ADMIN = ("usuarios", "usuario_nuevo", "usuario_editar", "usuario_eliminar")

# Endpoints de control de la suscripción: solo el Super Usuario.
_ENDPOINTS_SUPER = ("suscripcion", "suscripcion_renovar", "suscripcion_pausar")


def permiso_requerido(ep):
    """Clave de permiso necesaria para un endpoint (None = acceso libre)."""
    if ep in ("dashboard", "api_resumen", "login", "logout", "static",
              "suscripcion_vencida", "licencia_subir", "licencia_form"):
        return None
    if ep in _ENDPOINTS_SUPER:
        return "__super__"
    if ep in _ENDPOINTS_ADMIN:
        return "__admin__"
    if ep.startswith("orden"):
        return "ordenes"
    if ep.startswith("gasto"):
        return "gastos"
    if ep.startswith("almacen"):
        return "almacen"
    if ep.startswith("tareo"):
        return "tareo"
    if ep in ("planilla_opciones", "planilla_imprimir"):
        return "tareo"
    if ep == "cabecera":
        return "cabecera"
    if ep == "formatos":
        return "formatos"
    if ep in ("imprimir_manifiesto", "manifiesto_excel"):
        return "formatos"
    if ep in ("imprimir_fe05", "imprimir_fe06", "imprimir_panel"):
        return "formatos"
    if ep in ("imprimir_fe07", "imprimir_fe08"):
        return "almacen"
    if ep == "configuracion":
        return "configuracion"
    if ep == "respaldo":
        return "respaldo"
    return None


def home_usuario(u):
    """Página inicial tras ingresar: admin -> Dashboard, super -> usuarios,
    usuario -> su primera sección autorizada."""
    if not u:
        return url_for("dashboard")
    if u.rol == ROL_SUPER:
        return url_for("usuarios")
    if u.rol == "Administrador":
        return url_for("dashboard")
    for clave, _ in PERMISOS_SECCIONES:
        if clave in u.permiso_lista:
            return url_for(clave)
    return url_for("dashboard")


def _permisos_guardados(rol, form):
    """Claves de permiso a guardar: el Administrador y el Super Usuario siempre tienen todas."""
    validas = {clave for clave, _ in PERMISOS_SECCIONES}
    if rol in ROLES_TOTALES:
        return list(validas)
    return [c for c in form.getlist("permisos") if c in validas]


def leer_lista(archivo):
    """Lee un archivo de texto y devuelve sus líneas no vacías."""
    if not os.path.exists(archivo):
        return []
    with open(archivo, encoding="utf-8") as f:
        return [linea.strip() for linea in f if linea.strip()]


# ----------------------------------------------------------------------------
# Respaldo y recuperacion de la base de datos
# ----------------------------------------------------------------------------
BASE_PROYECTO = os.path.dirname(os.path.abspath(__file__))
# INFORME_RESPALDO_DIR permite aislar los respaldos (p.ej. en pruebas).
RESPALDO_DIR = (os.environ.get("INFORME_RESPALDO_DIR")
                or os.path.join(BASE_PROYECTO, "Respaldo BD"))


def ruta_db():
    """Ruta SQLite en uso durante esta petición.

    Si el usuario autenticado tiene base propia (Administrador u operador)
    devuelve la ruta de esa base; en caso contrario la base maestra. Fuera de
    una petición devuelve siempre la base maestra.
    """
    try:
        tid = session.get("tenant")
    except (RuntimeError, AttributeError):
        tid = None
    if tid:
        return tenant_path(tid)
    uri = os.environ.get("INFORME_DB")
    if uri and uri.startswith("sqlite:///"):
        return uri[len("sqlite:///"):]
    return os.path.join(BASE_PROYECTO, "instance", "informe.db")


# ------------------- CIFRADO DE RESPALDOS -------------------
_MAGIC = b"INFRES"  # cabecera mágica para identificar archivos cifrados
_SALT = b"InformeMensual2026!@#"


def _derivar_clave():
    """Derive una clave de 32 bytes a partir del hostname de la máquina."""
    import socket
    host = socket.gethostname().encode("utf-8")
    return hashlib.sha256(_SALT + host).digest()


def _cifrar_datos(data):
    """Cifra datos con XOR usando la clave derivada del hostname."""
    clave = _derivar_clave()
    # Stream cipher XOR con clave repetida
    out = bytearray(len(data))
    for i in range(len(data)):
        out[i] = data[i] ^ clave[i % len(clave)]
    return _MAGIC + bytes(out)


def _descifrar_datos(data):
    """Descifra datos XOR si tienen cabecera mágica; si no, retorna tal cual (compat)."""
    if data[:len(_MAGIC)] == _MAGIC:
        payload = data[len(_MAGIC):]
        clave = _derivar_clave()
        out = bytearray(len(payload))
        for i in range(len(payload)):
            out[i] = payload[i] ^ clave[i % len(clave)]
        return bytes(out)
    return data  # compatibilidad con respaldos antiguos sin cifrar


def _es_cifrado(path):
    """True si el archivo tiene cabecera mágica de cifrado."""
    try:
        with open(path, "rb") as f:
            return f.read(len(_MAGIC)) == _MAGIC
    except Exception:
        return False


def crear_respaldo(nombre=None, db_path=None):
    """Copia la base indicada (o la base en uso) a la carpeta de respaldos y la cifra.

    db_path permite respaldar explícitamente la base de un Administrador
    (p.ej. al restablecer o eliminar su proyecto).
    """
    os.makedirs(RESPALDO_DIR, exist_ok=True)
    if nombre is None:
        nombre = f"informe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    dest = os.path.join(RESPALDO_DIR, nombre)
    src = sqlite3.connect(db_path or ruta_db())
    dst = sqlite3.connect(dest)
    try:
        with dst:
            src.backup(dst)
    finally:
        dst.close()
        src.close()
    # Cifrar el archivo de respaldo
    try:
        with open(dest, "rb") as f:
            raw = f.read()
        with open(dest, "wb") as f:
            f.write(_cifrar_datos(raw))
    except Exception as e:
        logging.getLogger("respaldo").warning("No se pudo cifrar respaldo: %s", e)
    problemas = verificar_integridad(dest)
    if problemas:
        logging.getLogger("respaldo").warning(
            "Respaldo con problemas de integridad (%s): %s", dest, problemas)
    else:
        logging.getLogger("respaldo").info("Respaldo verificado e íntegro: %s", dest)
    return dest


def verificar_integridad(db_path):
    """Ejecuta PRAGMA integrity_check sobre una base SQLite (cifrada o no)."""
    try:
        with open(db_path, "rb") as f:
            raw = f.read()
        data = _descifrar_datos(raw)
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.write(data)
        tmp.close()
        try:
            con = sqlite3.connect(tmp.name)
            try:
                filas = con.execute("PRAGMA integrity_check").fetchall()
                return [f[0] for f in filas if f[0] != "ok"]
            finally:
                con.close()
        finally:
            try:
                os.remove(tmp.name)
            except Exception:
                pass
    except sqlite3.Error as e:
        return [str(e)]


def _respaldo_diario_automatico():
    """Crea (una vez al día) un respaldo automático de la base maestra y de
    cada base de Administrador, verificando la integridad de cada copia.

    Se ejecuta al arrancar el servidor; si ya existe el respaldo del día se
    omite para no duplicar archivos en cada reinicio.
    """
    os.makedirs(RESPALDO_DIR, exist_ok=True)
    hoy = datetime.now().strftime("%Y%m%d")
    bases = [("maestra", master_path())]
    try:
        admins = _bd.master_session.query(Usuario).filter(
            Usuario.rol == "Administrador",
            Usuario.activo == True).all()
    except Exception:
        admins = []
    for adm in admins:
        ruta = tenant_path(adm.id)
        if os.path.exists(ruta):
            bases.append((f"admin_{adm.id}", ruta))
    log = logging.getLogger("respaldo")
    for etiqueta, ruta in bases:
        nombre = f"auto_{hoy}_{etiqueta}.db"
        if os.path.exists(os.path.join(RESPALDO_DIR, nombre)):
            continue
        try:
            crear_respaldo(nombre, db_path=ruta)
            log.info("Respaldo automático diario creado: %s", nombre)
        except Exception as e:
            log.exception("Respaldo automático fallido (%s): %s", etiqueta, e)


def _hay_datos_proyecto(admin_id=None):
    """True si existe cualquier dato del proyecto (en la base en uso o en la
    base del Administrador indicado)."""
    modelos = (Proyecto, Presupuesto, Gasto, GastoDetalle,
               AlmacenMovimiento, Trabajador)
    if admin_id:
        S = tenant_session(admin_id)
        try:
            return any(S.query(m).first() for m in modelos)
        finally:
            S.close()
    return any(m.query.first() for m in modelos)


def _limpiar_datos_proyecto(admin_id=None):
    """Limpia todos los datos del proyecto para iniciar uno nuevo en cero.

    Si admin_id está presente, limpia la base de ese Administrador (lo usa el
    Super Usuario al restablecer el proyecto de un administrador); si no,
    limpia la base en uso durante la petición.

    Antes de borrar crea un respaldo automatico de la base afectada. Devuelve
    la ruta del respaldo generado. No toca usuarios, ni la suscripcion/licencia.
    """
    respaldo = crear_respaldo(
        f"inicio_nuevo_proyecto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
        db_path=tenant_path(admin_id) if admin_id else None)
    if admin_id:
        S = tenant_session(admin_id)
        try:
            S.query(GastoDetalle).delete(synchronize_session=False)
            S.query(Gasto).delete(synchronize_session=False)
            S.query(Presupuesto).delete(synchronize_session=False)
            S.query(AlmacenMovimiento).delete(synchronize_session=False)
            S.query(Trabajador).delete(synchronize_session=False)
            p = S.query(Proyecto).first()
            if p is None:
                p = Proyecto()
                S.add(p)
            _reiniciar_proyecto(p)
            S.commit()
        finally:
            S.close()
    else:
        GastoDetalle.query.delete(synchronize_session=False)
        Gasto.query.delete(synchronize_session=False)
        Presupuesto.query.delete(synchronize_session=False)
        AlmacenMovimiento.query.delete(synchronize_session=False)
        Trabajador.query.delete(synchronize_session=False)
        p = Proyecto.query.first()
        if p is None:
            p = Proyecto()
            db.session.add(p)
        _reiniciar_proyecto(p)
        db.session.commit()
    return respaldo


def _reiniciar_proyecto(p):
    """Deja el registro Proyecto en su estado por defecto (datos en cero)."""
    for col in Proyecto.__table__.columns:
        if col.name == "id":
            continue
        default = getattr(col.default, "arg", None) if col.default is not None else None
        if callable(default):
            default = None
        setattr(p, col.name, default)
    p.anio = date.today().year
    p.mes_actual = date.today().month


def listar_respaldos():
    """Devuelve la lista de archivos de respaldo ordenados del mas reciente al mas antiguo."""
    os.makedirs(RESPALDO_DIR, exist_ok=True)
    out = []
    for f in sorted(os.listdir(RESPALDO_DIR), reverse=True):
        ruta = os.path.join(RESPALDO_DIR, f)
        if os.path.isfile(ruta):
            out.append({"nombre": f, "ruta": ruta,
                        "tamano": os.path.getsize(ruta),
                        "mod": datetime.fromtimestamp(os.path.getmtime(ruta))})
    return out


def restaurar_respaldo(nombre):
    """Reemplaza la base de datos en uso (maestra o del administrador) por el
    contenido de un respaldo. Antes de restaurar se genera automaticamente una
    copia de seguridad del estado actual para no perder datos."""
    os.makedirs(RESPALDO_DIR, exist_ok=True)
    if not nombre or nombre != os.path.basename(nombre):
        raise ValueError("Nombre de respaldo no válido.")
    src = os.path.join(RESPALDO_DIR, nombre)
    if not os.path.isfile(src):
        raise FileNotFoundError("El respaldo no existe.")
    # Descifrar si es necesario
    try:
        with open(src, "rb") as f:
            raw = f.read()
        decrypted = _descifrar_datos(raw)
        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp.write(decrypted)
        tmp.close()
        restore_src = tmp.name
    except Exception:
        restore_src = src
    con = sqlite3.connect(restore_src)
    try:
        con.execute("SELECT COUNT(*) FROM sqlite_master").fetchone()
    finally:
        con.close()
    tid = session.get("tenant")
    target = ruta_db()
    pre = crear_respaldo(
        f"pre_restauracion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
    db.session.remove()
    if tid:
        dispose_tenant(tid)
    else:
        db.engine.dispose()
    src_con = sqlite3.connect(restore_src)
    dst_con = sqlite3.connect(target)
    try:
        with dst_con:
            src_con.backup(dst_con)
    finally:
        dst_con.close()
        src_con.close()
    # Limpiar archivo temporal si se creó
    if restore_src != src:
        try:
            os.remove(restore_src)
        except Exception:
            pass
    if tid:
        eng = tenant_engine(tid)
        db.metadata.create_all(bind=eng, tables=tablas_tenant())
        bind_session(eng)
    else:
        db.metadata.create_all(bind=db.engine, tables=tablas_maestras())
        migrar_schema()
        migrar_suscripcion()
        _migrar_a_multidb()
        # Limpiar estado de actualizacion previa si existe (no es un progreso activo)
        _estado = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                               "actualizar.estado")
        if os.path.exists(_estado):
            try:
                os.remove(_estado)
            except Exception:
                pass
        bind_session(_bd.master_engine)
    return pre


def _publicar_thread(root, ver, msj, estado):
    """Hilo de fondo que ejecuta bump + commit + push + workflow_dispatch de una
    nueva version, reportando el avance al archivo publicar.estado."""
    import re
    import subprocess

    def prog(fase, pct, mensaje):
        try:
            with open(estado, "w", encoding="utf-8") as fh:
                json.dump({"fase": fase, "porcentaje": pct, "mensaje": mensaje,
                           "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S")}, fh)
        except Exception:
            pass

    try:
        prog("publicando", 10, "Iniciando publicación de v%s..." % ver)
        # 1) bump version.py
        vpath = os.path.join(root, "informe_web", "version.py")
        txt = open(vpath, encoding="utf-8").read()
        txt2 = re.sub(r'__version__\s*=\s*"[^"]*"', '__version__  = "%s"' % ver, txt)
        if txt2 == txt:
            raise RuntimeError("no se encontró __version__ en version.py")
        prog("publicando", 25, "Actualizando versión a v%s..." % ver)
        open(vpath, "w", encoding="utf-8").write(txt2)
        # 2) commit + push
        msg = ("Release v%s" % ver) if not msj else \
            ("Release v%s (%s)" % (ver, msj.replace("\n", " ")[:140]))
        env = dict(os.environ)
        prog("publicando", 40, "Preparando commit...")
        subprocess.run(["git", "-C", root, "add", "-A"], env=env,
                       check=True, capture_output=True, text=True,
                       creationflags=0x08000000)
        subprocess.run(["git", "-C", root, "commit", "-m", msg], env=env,
                       check=True, capture_output=True, text=True,
                       creationflags=0x08000000)
        prog("publicando", 60, "Subiendo a GitHub (push)...")
        push = subprocess.run(["git", "-C", root, "push", "origin", "master"],
                              env=env, capture_output=True, text=True,
                              creationflags=0x08000000)
        if push.returncode != 0:
            raise RuntimeError("git push falló: %s" %
                               (push.stderr or push.stdout or "").strip()[:200])
        # 3) disparar workflow_dispatch (la release se construye en GitHub Actions)
        prog("publicando", 80, "Disparando build de release...")
        run = subprocess.run(["gh", "workflow", "run", "Build & Release"],
                             cwd=root, env=env, capture_output=True, text=True,
                             creationflags=0x08000000)
        if run.returncode != 0:
            raise RuntimeError("workflow no disparado: %s" %
                               (run.stderr or "").strip()[:200])
        prog("listo", 100, "Nueva versión v%s publicada." % ver)
    except Exception as e:
        prog("error", 100, str(e))


def registrar_rutas(app):
    """Registra todas las rutas del aplicativo."""
    @app.route("/robots.txt")
    def robots():
        base = request.url_root.rstrip("/")
        txt = ("User-agent: *\n"
               "Allow: /\n"
               "Disallow: /login\n"
               "Disallow: /usuarios\n"
               "Disallow: /suscripcion\n"
               "Disallow: /respaldo\n"
               "Disallow: /licencia\n"
               "Disallow: /api/\n"
               "Disallow: /*_imprimir\n"
               f"Sitemap: {base}/sitemap.xml\n")
        return txt, 200, {"Content-Type": "text/plain; charset=utf-8"}

    @app.route("/api/version")
    def api_version():
        """Version y commit del aplicativo, para el check de actualizaciones."""
        return jsonify({
            "version": app.config.get("INFORME_VERSION", "1.0.0"),
            "commit": os.environ.get("INFORME_COMMIT", ""),
        })

    @app.route("/api/actualizacion")
    def api_actualizacion():
        """Consulta la release mas reciente publicada en GitHub y decide si hay
        una version NUEVA disponible (remota > local).
        Solo Administradores y el Super Usuario pueden consultar.

        Estrategia: intenta gh CLI primero; si falla (no instalado / sin auth),
        usa la REST API de GitHub con un PAT embebido como fallback."""
        if not es_admin_actual():
            return jsonify({"error": "No autorizado"}), 403
        repo = app.config.get("INFORME_REPO", "jacj01/informe-mensual-obra")
        local_v = app.config.get("INFORME_VERSION", "1.0.0")

        def _vt(s):
            import re
            m = re.search(r"v?(\d+\.\d+\.\d+)", s or "")
            return tuple(int(x) for x in m.group(1).split(".")) if m else (0, 0, 0)

        data = None
        tag = ""

        # --- Estrategia 1: gh CLI ---
        try:
            import subprocess, re
            cmd = ["gh", "-R", repo, "release", "view", "--json",
                   "tagName,name,publishedAt,assets"]
            out = subprocess.run(cmd, capture_output=True, text=True, timeout=15,
                                 creationflags=0x08000000)
            if out.returncode == 0 and out.stdout.strip():
                raw = json.loads(out.stdout)
                tag = raw.get("tagName", "")
                asset = next((a for a in raw.get("assets", [])
                              if a.get("name", "").endswith(".zip")), None)
                data = {
                    "tag": tag,
                    "nombre": raw.get("name", ""),
                    "publicada": raw.get("publicatedAt", ""),
                    "asset": asset.get("name") if asset else None,
                    "url": asset.get("url") if asset else None,
                }
        except Exception:
            data = None  # fallback a REST API

        # --- Estrategia 2: REST API de GitHub (sin auth, repo publico) ---
        if data is None:
            try:
                import urllib.request, urllib.error, ssl
                api_url = f"https://api.github.com/repos/{repo}/releases/latest"
                req = urllib.request.Request(api_url)
                req.add_header("Accept", "application/vnd.github+json")
                req.add_header("User-Agent", "InformeObra/1.0")
                # Intentar con certificados del sistema; si falla (Python
                # embebido sin CA certs), usar contexto no verificado.
                try:
                    ctx = ssl.create_default_context()
                except Exception:
                    ctx = ssl._create_unverified_context()
                try:
                    with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
                        raw = json.loads(resp.read().decode("utf-8"))
                except (ssl.SSLError, OSError):
                    ctx2 = ssl._create_unverified_context()
                    with urllib.request.urlopen(req, timeout=15, context=ctx2) as resp:
                        raw = json.loads(resp.read().decode("utf-8"))
                tag = raw.get("tag_name", "")
                asset = next((a for a in raw.get("assets", [])
                              if a.get("name", "").endswith(".zip")), None)
                data = {
                    "tag": tag,
                    "nombre": raw.get("name", ""),
                    "publicada": raw.get("published_at", ""),
                    "asset": asset.get("name") if asset else None,
                    "url": asset.get("browser_download_url") if asset else None,
                }
            except Exception as e:
                Log('[UPDATE] REST API fallo: ' + str(e))
                return jsonify({"disponible": None, "error": str(e),
                                "version_actual": local_v}), 502

        hay_nueva = _vt(tag) > _vt(local_v)
        return jsonify({
            "disponible": bool(hay_nueva), "tag": tag,
            "nombre": data.get("nombre", ""),
            "publicada": data.get("publicada", ""),
            "asset": data.get("asset"),
            "url": data.get("url"),
            "version_actual": local_v,
        })

    @app.route("/actualizar", methods=["POST"])
    def aplicar_actualizacion():
        """Dispara la actualizacion en el propio equipo.
        Descarga el ZIP via gh release download (si gh esta disponible) o
        via REST API directa (repo publico). Lanza actualizar.ps1.
        Solo Admin / Super."""
        if not es_admin_actual():
            return jsonify({"error": "No autorizado"}), 403
        try:
            import subprocess, tempfile, urllib.request, ssl
            root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            ps1 = os.path.join(root, "actualizar.ps1")
            repo = app.config.get("INFORME_REPO", "jacj01/informe-mensual-obra")
            local_v = app.config.get("INFORME_VERSION", "1.0.0")

            # Obtener tag de la release remota
            tag = ""
            # Estrategia 1: gh CLI
            try:
                cmd = ["gh", "-R", repo, "release", "view", "--json", "tagName"]
                out = subprocess.run(cmd, capture_output=True, text=True, timeout=15,
                                     creationflags=0x08000000)
                if out.returncode == 0:
                    tag = json.loads(out.stdout).get("tagName", "")
            except Exception:
                pass
            # Estrategia 2: REST API publica
            if not tag:
                try:
                    api_url = f"https://api.github.com/repos/{repo}/releases/latest"
                    req = urllib.request.Request(api_url)
                    req.add_header("Accept", "application/vnd.github+json")
                    req.add_header("User-Agent", "InformeObra/1.0")
                    try:
                        ctx = ssl.create_default_context()
                    except Exception:
                        ctx = ssl._create_unverified_context()
                    try:
                        with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
                            tag = json.loads(resp.read().decode()).get("tag_name", "")
                    except (ssl.SSLError, OSError):
                        ctx2 = ssl._create_unverified_context()
                        with urllib.request.urlopen(req, timeout=15, context=ctx2) as resp:
                            tag = json.loads(resp.read().decode()).get("tag_name", "")
                except Exception:
                    pass

            if not tag:
                return jsonify({"ok": False,
                                "error": "No se pudo obtener la version remota. "
                                         "Verifique su conexion a internet."}), 400

            # Descargar ZIP: intentar gh primero, luego REST API directa
            tmp_dir = tempfile.mkdtemp(prefix="informe_upd_")
            zip_path = None
            # Intento 1: gh release download
            try:
                r = subprocess.run(["gh", "-R", repo, "release", "download", tag,
                                    "--pattern", "*.zip", "--dir", tmp_dir],
                                   capture_output=True, text=True, timeout=120,
                                   creationflags=0x08000000)
                if r.returncode == 0:
                    zips = [f for f in os.listdir(tmp_dir) if f.endswith(".zip")]
                    if zips:
                        zip_path = os.path.join(tmp_dir, zips[0])
            except Exception:
                zip_path = None
            # Intento 2: REST API directa (repo publico)
            if not zip_path:
                try:
                    api_dl = f"https://api.github.com/repos/{repo}/releases/tags/{tag}"
                    req2 = urllib.request.Request(api_dl)
                    req2.add_header("Accept", "application/vnd.github+json")
                    req2.add_header("User-Agent", "InformeObra/1.0")
                    try:
                        ctx2 = ssl.create_default_context()
                    except Exception:
                        ctx2 = ssl._create_unverified_context()
                    try:
                        with urllib.request.urlopen(req2, timeout=15, context=ctx2) as resp2:
                            rel = json.loads(resp2.read().decode())
                    except (ssl.SSLError, OSError):
                        ctx2 = ssl._create_unverified_context()
                        with urllib.request.urlopen(req2, timeout=15, context=ctx2) as resp2:
                            rel = json.loads(resp2.read().decode())
                    zip_asset = next((a for a in rel.get("assets", [])
                                      if a.get("name", "").endswith(".zip")), None)
                    if zip_asset:
                        dl_url = zip_asset.get("browser_download_url")
                        req3 = urllib.request.Request(dl_url)
                        req3.add_header("User-Agent", "InformeObra/1.0")
                        with urllib.request.urlopen(req3, timeout=120, context=ctx2) as dl_resp:
                            zip_path = os.path.join(tmp_dir, zip_asset["name"])
                            with open(zip_path, "wb") as f:
                                while True:
                                    chunk = dl_resp.read(65536)
                                    if not chunk:
                                        break
                                    f.write(chunk)
                except Exception:
                    zip_path = None

            if not zip_path:
                return jsonify({"ok": False,
                                "error": "No se pudo descargar el paquete. "
                                         "Ejecute 'gh auth login' para autenticar."}), 400

            # Lanzar el updater con -ZipFile
            ps_args = ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass",
                       "-File", ps1, "-Instalar",
                       "-VersionLocal", local_v, "-ZipFile", zip_path]

            _si = subprocess.STARTUPINFO()
            _si.dwFlags = subprocess.STARTF_USESHOWWINDOW
            _si.wShowWindow = 0  # SW_HIDE
            subprocess.Popen(ps_args, cwd=root,
                             startupinfo=_si,
                             creationflags=subprocess.CREATE_NO_WINDOW,
                             stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL,
                             stderr=subprocess.DEVNULL)
            return jsonify({"ok": True, "msg": "Actualizacion iniciada; el servidor se reiniciara en breve."})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500

    @app.route("/api/actualizacion/progreso")
    def progreso_actualizacion():
        """Lee el archivo de progreso escrito por actualizar.ps1.
        Solo Administradores y el Super Usuario."""
        if not es_admin_actual():
            return jsonify({"error": "No autorizado"}), 403
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pf = os.path.join(root, "actualizar.estado")
        if not os.path.exists(pf):
            return jsonify({"fase": "inactivo", "porcentaje": 0, "mensaje": ""})
        try:
            # actualizar.ps1 escribe con PowerShell 5.1 (BOM UTF-8): usar utf-8-sig
            with open(pf, encoding="utf-8-sig") as fh:
                data = json.load(fh)
            return jsonify(data)
        except Exception:
            return jsonify({"fase": "inactivo", "porcentaje": 0, "mensaje": ""})

    @app.route("/api/actualizar-publicar", methods=["POST"])
    def publicar_nueva_version():
        """Dispara la publicacion de una nueva release en GitHub: bump de
        version.py, commit, push y workflow_dispatch del release.
        Corre en un hilo de fondo que reporta avance a /api/actualizar-publicar/progreso.
        SOLO el Super Usuario."""
        if not es_super_usuario():
            return jsonify({"error": "No autorizado: se requiere Super Usuario"}), 403
        import re
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        payload = request.get_json(silent=True) or {}
        ver = (payload.get("version") or "").strip()
        if not re.match(r"^\d+\.\d+\.\d+$", ver):
            return jsonify({"ok": False, "error": "version invalida; use formato X.Y.Z"}), 400
        estado = os.path.join(root, "publicar.estado")
        # Si ya hay una publicacion en curso (no antigua), no duplicar.
        if os.path.exists(estado):
            try:
                with open(estado, encoding="utf-8-sig") as fh:
                    prev = json.load(fh)
                if prev.get("fase") == "publicando":
                    try:
                        ts = datetime.strptime(prev.get("ts", ""), "%Y-%m-%d %H:%M:%S")
                        fresco = (datetime.now() - ts).total_seconds() < 600
                    except Exception:
                        fresco = True
                    if fresco:
                        return jsonify({"ok": False,
                                        "error": "Ya hay una publicación en curso. Espere a que termine."}), 409
            except Exception:
                pass
        msj = (payload.get("mensaje") or "").strip()
        threading.Thread(target=_publicar_thread, args=(root, ver, msj, estado),
                         daemon=True).start()
        return jsonify({"ok": True, "msg": "Publicación iniciada; avance en la barra de progreso."})

    @app.route("/api/actualizar-publicar/progreso")
    def progreso_publicacion():
        """Lee el archivo de progreso escrito por el hilo de publicacion.
        SOLO el Super Usuario."""
        if not es_super_usuario():
            return jsonify({"error": "No autorizado"}), 403
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pf = os.path.join(root, "publicar.estado")
        if not os.path.exists(pf):
            return jsonify({"fase": "inactivo", "porcentaje": 0, "mensaje": ""})
        try:
            with open(pf, encoding="utf-8-sig") as fh:
                return jsonify(json.load(fh))
        except Exception:
            return jsonify({"fase": "inactivo", "porcentaje": 0, "mensaje": ""})


    @app.route("/sitemap.xml")
    def sitemap():
        base = request.url_root.rstrip("/")
        urls = [
            ("", "1.0", "daily"),
            ("/dashboard", "0.9", "daily"),
            ("/cabecera", "0.8", "weekly"),
            ("/ordenes", "0.8", "weekly"),
            ("/gastos", "0.8", "weekly"),
            ("/almacen", "0.8", "weekly"),
            ("/tareo", "0.8", "weekly"),
            ("/inventario", "0.7", "monthly"),
            ("/formatos", "0.7", "monthly"),
            ("/configuracion", "0.5", "monthly"),
        ]
        today = date.today().isoformat()
        items = "".join(
            f"  <url>\n"
            f"    <loc>{base}{u}</loc>\n"
            f"    <lastmod>{today}</lastmod>\n"
            f"    <changefreq>{cf}</changefreq>\n"
            f"    <priority>{p}</priority>\n"
            f"  </url>\n"
            for u, p, cf in urls
        )
        xml = ('<?xml version="1.0" encoding="UTF-8"?>\n'
               '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9"\n'
               '        xmlns:xhtml="http://www.w3.org/1999/xhtml">\n'
               f"{items}</urlset>\n")
        return Response(xml, mimetype="application/xml")

    @app.context_processor
    def inyectar_activo():
        ep = request.endpoint or ""
        if ep == "dashboard":
            active = "dashboard"
        elif ep == "cabecera":
            active = "cabecera"
        elif ep in ("ordenes", "orden_nuevo", "orden_editar", "orden_devengar",
                    "orden_material_nuevo", "orden_material_editar",
                    "orden_material_eliminar", "orden_eliminar"):
            active = "ordenes"
        elif ep in ("gastos", "gasto_nuevo", "gasto_editar",
                    "gasto_detalle_nuevo", "gasto_detalle_editar",
                    "gasto_detalle_eliminar", "gasto_eliminar"):
            active = "gastos"
        elif ep in ("formatos", "formatos_actividades"):
            active = "formatos"
        elif ep in ("almacen", "almacen_nuevo", "almacen_editar",
                    "almacen_eliminar", "almacen_oc", "almacen_agregar_oc"):
            active = "almacen"
        elif ep in ("tareo", "tareo_guardar", "tareo_imprimir",
                    "trabajador_nuevo",
                    "trabajador_editar", "trabajador_eliminar"):
            active = "tareo"
        elif ep == "configuracion":
            active = "configuracion"
        elif ep in ("usuarios", "usuario_nuevo", "usuario_editar", "usuario_eliminar"):
            active = "usuarios"
        elif ep in ("suscripcion", "suscripcion_renovar", "suscripcion_pausar"):
            active = "suscripcion"
        elif ep == "respaldo":
            active = "respaldo"
        else:
            active = ""
        u = usuario_actual()
        if not u:
            pset = set()
        elif u.rol == ROL_SUPER:
            # El Super Usuario solo ve gestión de usuarios, suscripción y respaldo.
            pset = {"respaldo"}
        elif u.rol in ROLES_TOTALES:
            pset = {clave for clave, _ in PERMISOS_SECCIONES}
        else:
            pset = set(u.permiso_lista)
        # Estado de la suscripción para el banner de aviso / vencimiento.
        susc = suscripcion_usuario(_dueno_suscripcion(u)) if u else None
        return {"active": active, "usuario": u, "permisos_usuario": pset,
                "susc": susc}

    @app.template_filter("moneda")
    def moneda(n):
        return f"S/ {n:,.2f}"

    @app.template_filter("fecha_larga")
    def fecha_larga(d):
        if not d:
            return ""
        meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                 "julio", "agosto", "septiembre", "octubre", "noviembre",
                 "diciembre"]
        try:
            return f"{d.day} de {meses[d.month - 1]} del {d.year}"
        except (AttributeError, IndexError, TypeError):
            return ""

    @app.template_filter("cant")
    def cant(n):
        if n is None:
            return ""
        n = float(n)
        if n == int(n):
            return f"{int(n):,}"
        return f"{n:,.2f}"

    def lista_clasificadores(p):
        """Opciones (codigo, nombre) para el select de clasificador."""
        return list(clasificadores_proyecto().items())

    def clasificadores_oc(p):
        """Opciones para O/C y O/S: Bienes, Servicios, Expediente Tecnico, Liquidacion y extras."""
        base = clasificadores_proyecto()
        codigos = [p.clasificador_bienes or "2.6.2.3.99.4",
                   p.clasificador_servicios or "2.6.2.3.99.5",
                   p.clasificador_expediente or "2.6.8.1.3.1",
                   p.clasificador_liquidacion or "LIQUIDACION"]
        excluidos = {p.clasificador_personal or "2.6.2.3.99.3"}
        codigos_set = set(codigos) | excluidos
        result = [(c, base.get(c, c)) for c in codigos]
        for c, n in base.items():
            if c not in codigos_set:
                result.append((c, n))
        return result

    def cls_nombre(p, code):
        """Nombre legible de un clasificador (usa los configurados en la cabecera)."""
        if not code:
            return ""
        base = clasificadores_proyecto()
        if code in base:
            return base[code]
        return CLASIFICADORES.get(code, code)

    @app.before_request
    def verificar_login():
        ep = request.endpoint or ""
        if ep in ("login", "logout", "static", "robots", "sitemap",
                  "api_version", "api_actualizacion"):
            return None
        # Cierre de sesión por inactividad (respaldo al temporizador del navegador).
        ahora = time.time()
        ultimo = session.get("_ultimo_acceso")
        if session.get("usuario_id"):
            if ultimo is not None and ahora - ultimo > 15 * 60:
                session.clear()
                flash("Su sesión se cerró por inactividad (15 minutos).", "info")
                if request.headers.get("X-Modal"):
                    return ("Debe iniciar sesión.<script>setTimeout(function(){location.reload()},800)</script>",
                            401)
                return redirect(url_for("login"))
            session["_ultimo_acceso"] = ahora
        if not session.get("usuario_id"):
            if request.headers.get("X-Modal"):
                return ("Debe iniciar sesión.<script>setTimeout(function(){location.reload()},800)</script>",
                        401)
            return redirect(url_for("login", next=request.path))
        u = usuario_actual()
        if not u:
            session.clear()
            if request.headers.get("X-Modal"):
                return ("Debe iniciar sesión.<script>setTimeout(function(){location.reload()},800)</script>",
                        401)
            return redirect(url_for("login", next=request.path))
        es_super = u.rol == ROL_SUPER
        if ep in ("suscripcion_vencida", "licencia_subir", "licencia_form",
                  "licencia_generar"):
            return None
        # Bloqueo por suscripción vencida o pausada: todos excepto el Super Usuario.
        if not es_super and not suscripcion_vigente(_dueno_suscripcion(u)):
            if request.headers.get("X-Modal"):
                return ("Suscripción vencida.<script>setTimeout(function(){location.href='" +
                        url_for("suscripcion_vencida") + "'},800)</script>", 401)
            return redirect(url_for("suscripcion_vencida"))
        # El Super Usuario gestiona cuentas y licencia; no tiene datos de
        # proyecto propios (esos viven en la base de cada Administrador).
        if es_super:
            if ep == "dashboard":
                return redirect(url_for("usuarios"))
            if ep not in ("usuarios", "usuario_nuevo", "usuario_editar",
                          "usuario_eliminar", "respaldo", "suscripcion",
                          "suscripcion_renovar", "suscripcion_pausar",
                          "aplicar_actualizacion", "api_actualizacion", "publicar_nueva_version",
                          "progreso_publicacion", "progreso_actualizacion"):
                flash("La cuenta principal gestiona cuentas y licencia; "
                      "esta sección corresponde a los datos de un proyecto.",
                      "error")
                return redirect(url_for("usuarios"))
            return None
        req = permiso_requerido(ep)
        if req is None:
            return None
        if u.rol == "Administrador":
            if req == "__super__":
                flash("No tiene permisos para gestionar la suscripción.", "error")
                return redirect(home_usuario(u))
            return None
        if req in ("__admin__", "__super__"):
            flash("No tiene permisos para acceder a esta sección.", "error")
            return redirect(home_usuario(u))
        if req not in u.permiso_lista:
            if request.headers.get("X-Modal"):
                return ("No tiene acceso a esta sección.<script>setTimeout(function(){location.reload()},800)</script>",
                        401)
            flash("No tiene acceso autorizado a esta sección.", "error")
            return redirect(home_usuario(u))
        return None

    def admin_requerido(f):
        @wraps(f)
        def envuelto(*args, **kwargs):
            if not session.get("usuario_id"):
                return redirect(url_for("login", next=request.path))
            u = usuario_actual()
            if not u or u.rol not in ROLES_TOTALES:
                flash("No tiene permisos para acceder a la gestión de usuarios.", "error")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return envuelto

    def super_requerido(f):
        @wraps(f)
        def envuelto(*args, **kwargs):
            if not session.get("usuario_id"):
                return redirect(url_for("login", next=request.path))
            u = usuario_actual()
            if not u or u.rol != ROL_SUPER:
                flash("No tiene permisos para acceder a esta opción.", "error")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return envuelto

    # ------------------------- ACCESO (LOGIN) --------------------------
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if session.get("usuario_id"):
            return redirect(home_usuario(usuario_actual()))
        error = None
        if request.method == "POST":
            nombre = request.form.get("usuario", "").strip()
            clave = request.form.get("clave", "")
            usr = None
            tid = None
            # Las cuentas de Super Usuario y Administrador viven en la maestra.
            maestro = (_bd.master_session.query(Usuario)
                       .filter(Usuario.usuario == nombre).first())
            if maestro and maestro.activo \
                    and check_password_hash(maestro.clave, clave):
                usr = maestro
                # El Administrador entra con su propia base de proyecto.
                if usr.rol == "Administrador":
                    tid = usr.id
            else:
                # Los operadores (rol Usuario) viven en la base de su
                # Administrador: se busca en cada una de ellas.
                for adm in (_bd.master_session.query(Usuario)
                            .filter(Usuario.rol == "Administrador").all()):
                    op = _buscar_operador(adm.id, nombre)
                    if op and op.activo \
                            and check_password_hash(op.clave, clave):
                        usr, tid = op, adm.id
                        break
            if usr:
                if tid:
                    ensure_tenant(tid)
                    session["tenant"] = tid
                session["usuario_id"] = usr.id
                session["usuario_nombre"] = usr.nombres or usr.usuario
                session["usuario_rol"] = usr.rol
                return redirect(home_usuario(usr))
            error = "Usuario o contraseña incorrectos."
        return render_template("login.html", error=error)

    @app.route("/logout")
    def logout():
        session.clear()
        flash("Sesión cerrada correctamente.", "success")
        return redirect(url_for("login"))

    # ------------------------- GESTION DE USUARIOS ----------------------
    @app.route("/usuarios")
    @admin_requerido
    def usuarios():
        orden = {ROL_SUPER: 0, "Administrador": 1}
        actor = usuario_actual()
        es_super = bool(actor and actor.rol == ROL_SUPER)
        lista = sorted(Usuario.query.all(),
                       key=lambda u: (orden.get(u.rol, 2), u.usuario.lower()))
        if not es_super:
            lista = [u for u in lista if u.rol != ROL_SUPER]
        return render_template("usuarios.html", usuarios=lista,
                               labels=dict(PERMISOS_SECCIONES),
                               total=len(lista),
                               admins=sum(1 for u in lista if u.rol == "Administrador"),
                               activos=sum(1 for u in lista if u.activo),
                               es_super=es_super)

    @app.route("/usuarios/nuevo", methods=["GET", "POST"])
    @admin_requerido
    def usuario_nuevo():
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        actor = usuario_actual()
        es_super = bool(actor and actor.rol == ROL_SUPER)
        error = None
        if request.method == "POST":
            nombre = request.form.get("usuario", "").strip()
            clave = request.form.get("clave", "")
            rol = request.form.get("rol", "Usuario")
            if not nombre:
                error = "Debe indicar el nombre de usuario."
            elif not clave:
                error = "Debe indicar una contraseña."
            elif len(clave) < 6:
                error = "La contraseña debe tener al menos 6 caracteres."
            elif rol == ROL_SUPER and not es_super:
                error = "No tiene permisos para crear cuentas con ese rol."
            elif rol == "Administrador" and not es_super:
                error = "No tiene permisos para crear usuarios Administradores."
            elif Usuario.query.filter(Usuario.usuario == nombre).first():
                error = f"El usuario '{nombre}' ya existe."
            else:
                opt_licencia = request.form.get("licencia", "nueva")
                nu = Usuario(
                    usuario=nombre,
                    clave=generate_password_hash(clave),
                    nombres=request.form.get("nombres", "").strip(),
                    rol=rol,
                    activo=bool(request.form.get("activo")),
                    permisos=json.dumps(_permisos_guardados(rol, request.form)))
                # Un Administrador nuevo inicia sin licencia (debe comprarla),
                # salvo que el super elija mantener la licencia actual.
                if es_super and rol == "Administrador" and opt_licencia != "mantener":
                    nu.susc_activa = False
                db.session.add(nu)
                db.session.flush()
                nuevo_id = nu.id
                db.session.commit()
                try:
                    crear_respaldo(
                        f".usuario_nuevo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
                except Exception:
                    logging.getLogger("respaldo").warning(
                        "No se pudo crear respaldo al crear usuario", exc_info=True)
                if es_super and rol == "Administrador":
                    # Cada Administrador tiene su propia base, que nace vacía.
                    ensure_tenant(nuevo_id)
                    if opt_licencia == "mantener":
                        flash(f"Usuario '{nombre}' creado como Administrador. "
                              f"Mantiene la licencia actual y su proyecto "
                              f"comienza en cero.", "success")
                    else:
                        flash(f"Usuario '{nombre}' creado como Administrador. "
                              f"Su proyecto comienza en cero y no afecta a los "
                              f"demás administradores. Debe adquirir su "
                              f"licencia para ingresar.", "success")
                else:
                    flash(f"Usuario '{nombre}' creado.", "success")
                return redirect(url_for("usuarios"))
        u = Usuario()
        status = 400 if (es_modal and error) else 200
        return render_template("_usuario_form.html", u=u, es_edicion=False,
                               permisos_opciones=PERMISOS_SECCIONES, error=error,
                               es_super=es_super), status

    @app.route("/usuarios/editar/<int:uid>", methods=["GET", "POST"])
    @admin_requerido
    def usuario_editar(uid):
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        actor = usuario_actual()
        es_super = bool(actor and actor.rol == ROL_SUPER)
        u = db.session.get(Usuario, uid)
        if not u:
            abort(404)
        # La cuenta principal solo la administra su propio responsable.
        if u.rol == ROL_SUPER and not es_super:
            flash("No tiene permisos para modificar esa cuenta.", "error")
            return redirect(url_for("usuarios"))
        # El Administrador no modifica cuentas de otro Administrador (excepto la
        # propia cuenta: el administrador puede corregir sus datos/contraseña).
        if u.rol == "Administrador" and not es_super \
                and not _es_cuenta_propia(u):
            flash("No tiene permisos para modificar cuentas de Administrador.",
                  "error")
            return redirect(url_for("usuarios"))
        error = None
        if request.method == "POST":
            nombre = request.form.get("usuario", "").strip()
            nuevo_rol = request.form.get("rol", "Usuario")
            # La cuenta principal conserva siempre su rol (no se ofrece en el
            # formulario de rol).
            if u.rol == ROL_SUPER and _es_cuenta_propia(u):
                nuevo_rol = ROL_SUPER
            if not nombre:
                error = "Debe indicar el nombre de usuario."
            elif nuevo_rol == ROL_SUPER and not es_super:
                error = "No tiene permisos para asignar ese rol."
            elif (nuevo_rol == "Administrador" and not es_super
                    and u.rol != "Administrador"):
                error = "No tiene permisos para asignar ese rol."
            elif (_es_cuenta_propia(u) and u.rol == "Administrador"
                    and nuevo_rol != "Administrador" and not es_super):
                error = "No puede quitarse el rol de Administrador a sí mismo."
            else:
                otro = Usuario.query.filter(Usuario.usuario == nombre,
                                            Usuario.id != uid).first()
                if otro:
                    error = f"El usuario '{nombre}' ya existe."
                elif (_es_cuenta_propia(u)
                        and not bool(request.form.get("activo"))):
                    error = "No puede desactivar su propio usuario."
                elif (request.form.get("clave", "")
                        and len(request.form.get("clave", "")) < 6):
                    error = "La contraseña debe tener al menos 6 caracteres."
                elif (_es_cuenta_propia(u) and u.rol == ROL_SUPER
                        and nuevo_rol != ROL_SUPER):
                    error = "No puede quitarse su propio rol."
                else:
                    u.usuario = nombre
                    u.nombres = request.form.get("nombres", "").strip()
                    u.rol = nuevo_rol
                    u.activo = bool(request.form.get("activo"))
                    u.permisos = json.dumps(
                        _permisos_guardados(u.rol, request.form))
                    clave = request.form.get("clave", "")
                    if clave:
                        u.clave = generate_password_hash(clave)
                    db.session.commit()
                    try:
                        crear_respaldo(
                            f".usuario_editar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
                    except Exception:
                        logging.getLogger("respaldo").warning(
                            "No se pudo crear respaldo al editar usuario", exc_info=True)
                    if es_super and nuevo_rol == "Administrador":
                        ensure_tenant(u.id)
                    if _es_cuenta_propia(u):
                        session["usuario_nombre"] = u.nombres or u.usuario
                        session["usuario_rol"] = u.rol
                    if (es_super and u.rol == "Administrador"
                            and request.form.get("limpiar_datos") == "1"):
                        try:
                            respaldo = _limpiar_datos_proyecto(u.id)
                            flash(f"Usuario actualizado. Proyecto del administrador "
                                  f"reiniciado en cero. Respaldo automático en "
                                  f"'Respaldo BD/{os.path.basename(respaldo)}'.",
                                  "success")
                        except Exception as e:
                            flash(f"Usuario actualizado, pero no se pudo reiniciar "
                                  f"el proyecto: {e}", "error")
                    else:
                        flash("Usuario actualizado.", "success")
                    return redirect(url_for("usuarios"))
        status = 400 if (es_modal and error) else 200
        return render_template("_usuario_form.html", u=u, es_edicion=True,
                               permisos_opciones=PERMISOS_SECCIONES, error=error,
                               es_super=es_super, es_cuenta_propia=_es_cuenta_propia(u),
                               hay_datos=bool(es_super and u.rol == "Administrador"
                                              and _hay_datos_proyecto(uid))), status

    @app.route("/usuarios/eliminar/<int:uid>", methods=["POST"])
    @admin_requerido
    def usuario_eliminar(uid):
        actor = usuario_actual()
        es_super = bool(actor and actor.rol == ROL_SUPER)
        u = db.session.get(Usuario, uid)
        if not u:
            flash("Usuario no encontrado.", "error")
        elif _es_cuenta_propia(u):
            flash("No puede eliminar su propio usuario.", "error")
        elif u.rol == ROL_SUPER and not es_super:
            flash("No tiene permisos para eliminar esa cuenta.", "error")
        elif u.rol == "Administrador" and not es_super:
            flash("No tiene permisos para eliminar esa cuenta.", "error")
        else:
            admins = Usuario.query.filter(Usuario.rol == "Administrador",
                                          Usuario.activo == True).count()
            # La protección del "último administrador" aplica solo al Administrador:
            # la cuenta principal puede eliminar administradores sin restricciones.
            if (not es_super and u.rol == "Administrador" and u.activo
                    and admins <= 1):
                flash("No puede eliminar el último administrador activo.", "error")
            else:
                borrado_bd = False
                if es_super and u.rol == "Administrador":
                    # Libera el motor y elimina la base propia del Administrador
                    # (con respaldo previo) para no dejar datos huerfanos.
                    dispose_tenant(u.id)
                    ruta_ten = tenant_path(u.id)
                    if os.path.exists(ruta_ten):
                        crear_respaldo(
                            f"admin_{u.usuario}_"
                            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
                            db_path=ruta_ten)
                        borrado_bd = True
                    shutil.rmtree(os.path.dirname(ruta_ten), ignore_errors=True)
                db.session.delete(u)
                db.session.commit()
                flash(f"Usuario '{u.usuario}' eliminado."
                      + (" Su base de datos fue respaldada y eliminada."
                         if borrado_bd else ""), "success")
        return redirect(url_for("usuarios"))

    # ------------------------- SUSCRIPCION -----------------------------
    @app.route("/suscripcion")
    @super_requerido
    def suscripcion():
        s = get_suscripcion()
        historial = (SuscripcionHistorial.query
                     .order_by(SuscripcionHistorial.fecha_registro.desc())
                     .limit(50).all())
        admins = (_bd.master_session.query(Usuario)
                  .filter(Usuario.rol == "Administrador", Usuario.activo.is_(True))
                  .order_by(Usuario.usuario).all())
        return render_template("suscripcion.html", s=s, historial=historial,
                               PLANES=PLANES_SUSCRIPCION, admins=admins)

    @app.route("/suscripcion/renovar", methods=["POST"])
    @super_requerido
    def suscripcion_renovar():
        s = get_suscripcion()
        if not s:
            s = Suscripcion(plan="Mensual")
            db.session.add(s)
        plan = request.form.get("plan", "Mensual")
        if plan not in PLANES_SUSCRIPCION:
            plan = "Mensual"
        nota = request.form.get("nota", "").strip()
        hoy = date.today()
        # Si la suscripción sigue vigente, la renovación se encadena desde su
        # fecha de fin actual; si venció, se inicia desde hoy.
        base = s.fecha_fin if (s.fecha_fin and s.fecha_fin >= hoy) else hoy
        fin = sumar_meses(base, PLANES_SUSCRIPCION[plan])
        s.plan = plan
        s.fecha_inicio = base
        s.fecha_fin = fin
        s.activa = True
        actor = usuario_actual()
        db.session.add(SuscripcionHistorial(
            plan=plan, fecha_inicio=base, fecha_fin=fin,
            usuario=(actor.nombres or actor.usuario) if actor else "",
            nota=nota, accion="Renovación"))
        db.session.commit()
        flash(f"Suscripción {plan} vigente hasta el {fin.strftime('%d/%m/%Y')}.",
              "success")
        return redirect(url_for("suscripcion"))

    @app.route("/suscripcion/pausar", methods=["POST"])
    @super_requerido
    def suscripcion_pausar():
        s = get_suscripcion()
        if s:
            s.activa = not s.activa
            actor = usuario_actual()
            db.session.add(SuscripcionHistorial(
                plan=s.plan, fecha_inicio=s.fecha_inicio, fecha_fin=s.fecha_fin,
                usuario=(actor.nombres or actor.usuario) if actor else "",
                nota="", accion="Pausada" if not s.activa else "Reactivada"))
            db.session.commit()
            flash("Suscripción pausada: el aplicativo quedó bloqueado." if not s.activa
                  else "Suscripción reactivada.", "success")
        return redirect(url_for("suscripcion"))

    @app.route("/suscripcion/vencida")
    def suscripcion_vencida():
        u = usuario_actual()
        dueno = _dueno_suscripcion(u)
        s = suscripcion_usuario(dueno)
        return render_template("suscripcion_vencida.html", s=s,
                               es_super=bool(u and u.rol == ROL_SUPER),
                               admin_usuario=(dueno.usuario if dueno else ""),
                               admin_nombres=(dueno.nombres if dueno else ""))

    @app.route("/licencia/generar", methods=["POST"])
    @super_requerido
    def licencia_generar():
        """Genera y descarga un archivo de licencia para el plan y el usuario
        (Administrador) indicado. La licencia queda vinculada a ese usuario."""
        plan = request.form.get("plan", "Mensual")
        if plan not in PLANES_SUSCRIPCION:
            plan = "Mensual"
        uid = request.form.get("usuario_id", type=int)
        titular = (_bd.master_session.get(Usuario, uid)
                   if uid else None)
        if not titular or titular.rol != "Administrador":
            flash("Seleccione el usuario Administrador para la licencia.", "error")
            return redirect(url_for("suscripcion"))
        contenido = generar_licencia(plan, usuario=titular.usuario,
                                     usuario_id=titular.id,
                                     nombres=titular.nombres)
        nombre = (f"licencia_{titular.usuario}_{plan.lower()}_"
                  f"{date.today().strftime('%Y%m%d')}.lic")
        return Response(
            contenido, mimetype="application/octet-stream",
            headers={"Content-Disposition":
                     f"attachment; filename={nombre}"})

    @app.route("/licencia/subir", methods=["POST"])
    def licencia_subir():
        """El cliente sube el archivo de licencia para renovar la suscripción."""
        f = request.files.get("licencia")
        if f is None or not f.filename:
            flash("Seleccione el archivo de licencia (.lic o .json).", "error")
            return redirect(url_for("suscripcion_vencida"))
        s = aplicar_licencia(descifrar_licencia(f.read()))
        if s is None:
            flash("El archivo de licencia no es válido, ya fue utilizado o no "
                  "corresponde a la cuenta de su usuario.", "error")
            return redirect(url_for("suscripcion_vencida"))
        plan = s["plan"]
        hoy = date.today()
        if s["fecha_inicio"] and s["fecha_inicio"] > hoy:
            dias_rest = (s["fecha_inicio"] - hoy).days
            flash(f"Suscripción ampliada por licencia ({plan}): se sumaron "
                  f"{dias_rest} días restantes más el período de la licencia. "
                  f"Nueva vigencia hasta el {s['fecha_fin'].strftime('%d/%m/%Y')}.",
                  "success")
        else:
            flash(f"Suscripción renovada por licencia ({plan}): nueva vigencia "
                  f"hasta el {s['fecha_fin'].strftime('%d/%m/%Y')}.", "success")
        return redirect(url_for("dashboard"))

    @app.route("/licencia/form")
    def licencia_form():
        """Formulario de subida de licencia (usado en el modal 'Ampliar licencia')."""
        return render_template("_licencia_form.html")

    @app.route("/")
    def dashboard():
        u = usuario_actual()
        if u and u.rol == ROL_SUPER:
            return redirect(url_for("usuarios"))
        p = get_proyecto()
        k = kpis()
        num_anios = p.num_anios_anteriores if p.num_anios_anteriores is not None else 3
        por_mes = ejecucion_por_mes(p.anio)
        por_comp = ejecucion_por_componente(p.anio)
        recientes = (Gasto.query.order_by(Gasto.id.desc()).limit(8).all())
        return render_template(
            "dashboard.html", p=p, k=k, MESES=MESES, por_mes=por_mes,
            por_comp=por_comp, recientes=recientes, num_anios=num_anios,
            fe06_resumen=fe06_resumen(fe06_rows()))

    # ------------------------- CABECERA -------------------------------
    @app.route("/cabecera", methods=["GET", "POST"])
    def cabecera():
        p = get_proyecto()
        if request.method == "POST":
            for campo in ["nombre", "cui", "meta", "distrito", "provincia", "departamento",
                          "entidad", "unidad_ejecutora", "aprobacion", "rubro", "fuente",
                          "residente", "supervisor", "asistente",
                           "responsable_almacen",
                           "n_resolucion_adicional", "cip_supervisor", "cip_residente",
                           "colegiatura_admin", "dni_responsable_almacen",
                           "asistente_tecnico", "dni_cip_asistente"]:
                setattr(p, campo, request.form.get(campo, "").strip().upper())
            try:
                setattr(p, "monto_ampliacion",
                        float(request.form.get("monto_ampliacion", 0) or 0))
            except ValueError:
                pass
            # Guardar presupuesto desglosado (et y pim) por componente/detalle
            for i, (comp, det) in enumerate(PRESUPUESTO_DETALLE):
                try:
                    et_val = float(request.form.get('pres_et_%d' % i, 0) or 0)
                    pim_val = float(request.form.get('pres_pim_%d' % i, 0) or 0)
                except ValueError:
                    et_val = pim_val = 0
                cfg = Presupuesto.query.filter_by(componente=comp, detalle=det).first()
                if cfg:
                    cfg.et = et_val
                    cfg.pim2026 = pim_val
                else:
                    clasif = detalle_clasificador(det)
                    db.session.add(Presupuesto(componente=comp, clasificador=clasif,
                                               detalle=det, et=et_val, pim2026=pim_val))
            db.session.flush()
            # Parsear clasificadores extras desde el formulario primero
            extras = []
            i = 0
            while True:
                cod = request.form.get(f"extra_codigo_{i}", "").strip().upper()
                nom = request.form.get(f"extra_nombre_{i}", "").strip().upper()
                comp = request.form.get(f"extra_componente_{i}", "Costo Directo").strip()
                if cod and nom:
                    extras.append({"codigo": cod, "nombre": nom, "componente": comp})
                elif not cod and not nom:
                    break
                i += 1
            p.clasificadores_extra = json.dumps(extras) if extras else ""
            # Guardar presupuesto de clasificadores extras (usa lista nueva del formulario)
            extras_list = extras
            for i, ex in enumerate(extras_list):
                nombre = (ex.get("nombre") or "").strip().upper()
                componente = (ex.get("componente") or "Costo Directo").strip()
                codigo = (ex.get("codigo") or "").strip().upper()
                if not nombre:
                    continue
                try:
                    et_val = float(request.form.get('pres_et_extra_%d' % i, 0) or 0)
                    pim_val = float(request.form.get('pres_pim_extra_%d' % i, 0) or 0)
                except ValueError:
                    et_val = pim_val = 0
                cfg = Presupuesto.query.filter_by(componente=componente, detalle=nombre).first()
                if cfg:
                    cfg.et = et_val
                    cfg.pim2026 = pim_val
                    cfg.clasificador = codigo or cfg.clasificador
                else:
                    db.session.add(Presupuesto(componente=componente, clasificador=codigo,
                                               detalle=nombre, et=et_val, pim2026=pim_val))
            db.session.flush()
            # Eliminar Presupuesto rows de clasificadores extra que ya no existen
            base_dets = {det for _, det in PRESUPUESTO_DETALLE}
            extra_dets = set()
            for ex in extras_list:
                nombre = (ex.get("nombre") or "").strip().upper()
                if nombre:
                    extra_dets.add(nombre)
            for cfg in Presupuesto.query.all():
                if cfg.detalle not in base_dets and cfg.detalle not in extra_dets:
                    db.session.delete(cfg)
            db.session.flush()
            # Recalcular totales del proyecto desde Presupuesto
            from collections import defaultdict
            et_comp = defaultdict(float)
            for cfg in Presupuesto.query.filter(
                    Presupuesto.componente.in_(COMPONENTES_FE06)).all():
                et_comp[cfg.componente] = round(et_comp[cfg.componente] + (cfg.et or 0), 2)
            p.costo_directo = round(et_comp.get("Costo Directo", 0), 2)
            p.gastos_generales = round(et_comp.get("Gastos Generales", 0), 2)
            p.gastos_supervision = round(et_comp.get("Gastos de Supervisión", 0), 2)
            p.elaboracion_expediente = round(et_comp.get("Elaboración de Expediente Técnico", 0), 2)
            p.liquidacion_obra = round(et_comp.get("Liquidación de Obra", 0), 2)
            p.presupuesto_total = round(sum(et_comp.values()), 2)
            for campo in ["fecha_inicio", "fecha_fin", "nuevo_final_obra",
                          "fecha_aprobacion"]:
                setattr(p, campo, parse_fecha(request.form.get(campo)))
            clas_prev = {
                "PERSONAL": p.clasificador_personal,
                "BIENES": p.clasificador_bienes,
                "SERVICIOS": p.clasificador_servicios,
                "ELABORACION DE EXPEDIENTE TECNICO": p.clasificador_expediente,
                "COSTO DE LIQUIDACION": p.clasificador_liquidacion,
            }
            for campo in ["clasificador_personal", "clasificador_bienes",
                          "clasificador_servicios", "clasificador_expediente",
                          "clasificador_liquidacion"]:
                setattr(p, campo, request.form.get(campo, "").strip().upper())
            p.clasificador_personal = p.clasificador_personal or "2.6.2.3.99.3"
            p.clasificador_bienes = p.clasificador_bienes or "2.6.2.3.99.4"
            p.clasificador_servicios = p.clasificador_servicios or "2.6.2.3.99.5"
            p.clasificador_expediente = p.clasificador_expediente or "2.6.8.1.3.1"
            p.clasificador_liquidacion = p.clasificador_liquidacion or "LIQUIDACION"
            clas_nuevo = {
                "PERSONAL": p.clasificador_personal,
                "BIENES": p.clasificador_bienes,
                "SERVICIOS": p.clasificador_servicios,
                "ELABORACION DE EXPEDIENTE TECNICO": p.clasificador_expediente,
                "COSTO DE LIQUIDACION": p.clasificador_liquidacion,
            }
            for detalle, nuevo in clas_nuevo.items():
                previo = clas_prev.get(detalle)
                if previo and nuevo and previo != nuevo:
                    for cfg in Presupuesto.query.filter_by(detalle=detalle).all():
                        cfg.clasificador = nuevo
                    for g in Gasto.query.filter(Gasto.clasificador == previo).all():
                        g.clasificador = nuevo
            for campo in ["dias_ejecucion", "dias_ampliacion"]:
                try:
                    setattr(p, campo, int(request.form.get(campo, 0) or 0))
                except ValueError:
                    pass
            raw_ad = request.form.get("adicionales_json", "") or ""
            p.adicionales = raw_ad
            adicionales_list = []
            if raw_ad:
                try:
                    adicionales_list = json.loads(raw_ad)
                except (json.JSONDecodeError, TypeError):
                    adicionales_list = []
            p.adicional_obra = len(adicionales_list) > 0
            if adicionales_list:
                ultimo = adicionales_list[-1]
                nf_str = ultimo.get("nuevo_final_obra", "")
                if nf_str:
                    try:
                        p.fecha_fin = datetime.strptime(nf_str, "%Y-%m-%d").date()
                    except (ValueError, TypeError):
                        pass
                p.n_resolucion_adicional = ultimo.get("n_resolucion_adicional", "")
                total_dias = 0
                total_monto_amp = 0
                any_amp = False
                for ad in adicionales_list:
                    try:
                        total_dias += int(ad.get("dias_ampliacion", 0) or 0)
                    except (ValueError, TypeError):
                        pass
                    if ad.get("ampliacion_presupuestal", False):
                        any_amp = True
                        try:
                            total_monto_amp += float(ad.get("monto_ampliacion", 0) or 0)
                        except (ValueError, TypeError):
                            pass
                p.dias_ampliacion = total_dias
                p.ampliacion_presupuestal = any_amp
                p.monto_ampliacion = total_monto_amp
            else:
                p.dias_ampliacion = 0
                p.nuevo_final_obra = None
                p.n_resolucion_adicional = ""
                p.ampliacion_presupuestal = False
                p.monto_ampliacion = 0
            try:
                p.anio = int(request.form.get("anio", p.anio))
            except ValueError:
                pass
            try:
                p.mes_actual = int(request.form.get("mes_actual", p.mes_actual))
            except ValueError:
                pass
            if p.fecha_inicio and not p.adicional_obra:
                if p.fecha_fin:
                    p.dias_ejecucion = max(0, (p.fecha_fin - p.fecha_inicio).days)
                elif p.dias_ejecucion:
                    p.fecha_fin = p.fecha_inicio + timedelta(days=p.dias_ejecucion)
            db.session.commit()
            flash("Datos de cabecera guardados correctamente.", "success")
            return redirect(url_for("cabecera"))
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        rubros = leer_lista(os.path.join(base, "Rubro.txt"))
        fuentes = leer_lista(os.path.join(base, "Recursos.txt"))
        extras = []
        raw_ex = getattr(p, "clasificadores_extra", "") or ""
        if raw_ex:
            try:
                extras = json.loads(raw_ex)
            except (json.JSONDecodeError, TypeError):
                extras = []
        return render_template("cabecera.html", p=p, MESES=MESES,
                               rubros=rubros, fuentes=fuentes,
                               presupuesto=presupuesto_filas(),
                               clasificadores_extra=extras,
                               COMPONENTES=COMPONENTES)

    @app.route("/cabecera/subir-logo", methods=["POST"])
    def subir_logo():
        p = get_proyecto()
        archivo = request.files.get("logo")
        if archivo and archivo.filename:
            ext = os.path.splitext(archivo.filename)[1].lower()
            if ext in (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"):
                destino = os.path.join(app.static_folder, "uploads", f"logo{ext}")
                archivo.save(destino)
                p.logo_path = f"uploads/logo{ext}"
                db.session.commit()
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return jsonify({"ok": True, "logo_url": url_for("static", filename=p.logo_path)})
                flash("Logo actualizado correctamente.", "success")
            else:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return jsonify({"ok": False, "error": "Formato no permitido"}), 400
                flash("Formato de imagen no permitido.", "error")
        else:
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"ok": False, "error": "Seleccione un archivo"}), 400
            flash("Seleccione un archivo de imagen.", "error")
        return redirect(url_for("cabecera"))

    # ------------------------- GASTOS --------------------------------
    @app.route("/gastos")
    def gastos():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        lista = gastos_mes(mes, anio, devengado=True)
        grupos = {}
        for g in lista:
            key = (g.componente, g.clasificador)
            grupos.setdefault(key, {"componente": g.componente, "clasificador": g.clasificador,
                                    "detalle": cls_nombre(p, g.clasificador),
                                    "gastos": [], "total": 0.0})
            grupos[key]["gastos"].append(g)
            grupos[key]["total"] += g.importe
        for key in grupos:
            grupos[key]["total"] = round(grupos[key]["total"], 2)
        total = total_gastos_mes(mes, anio, devengado=True)
        return render_template("gastos.html", p=p, lista=lista, grupos=grupos,
                               total=total, mes=mes, anio=anio, MESES=MESES,
                               CLASIFICADORES=CLASIFICADORES)

    @app.route("/gastos/nuevo", methods=["GET", "POST"])
    def gasto_nuevo():
        return gasto_form(None)

    @app.route("/gastos/editar/<int:gasto_id>", methods=["GET", "POST"])
    def gasto_editar(gasto_id):
        return gasto_form(gasto_id)

    @app.route("/gastos/eliminar/<int:gasto_id>", methods=["POST"])
    def gasto_eliminar(gasto_id):
        g = db.session.get(Gasto, gasto_id)
        if g:
            mes, anio = g.mes, g.anio
            db.session.delete(g)
            db.session.commit()
            flash("Gasto eliminado.", "info")
            return redirect(url_for("gastos", mes=mes, anio=anio))
        return redirect(url_for("gastos"))

    def gasto_form(gasto_id):
        p = get_proyecto()
        g = db.session.get(Gasto, gasto_id) if gasto_id else Gasto(
            fecha=fecha_defecto(p),
            mes=p.mes_actual, anio=p.anio,
            componente="Costo Directo", clasificador="2.6.2.3.99.4")
        if gasto_id and g is None:
            abort(404)
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        error = None
        if request.method == "POST":
            g.fecha = parse_fecha(request.form.get("fecha")) or g.fecha
            try:
                g.siaf = int(request.form.get("siaf", 0) or 0)
            except ValueError:
                pass
            g.tipo_doc = request.form.get("tipo_doc", "O/C").upper()
            try:
                g.num_doc = int(request.form.get("num_doc", 0) or 0)
            except ValueError:
                pass
            g.proveedor = request.form.get("proveedor", "").upper()
            g.clasificador = request.form.get("clasificador", "2.6.2.3.99.4")
            g.componente = request.form.get("componente", "Costo Directo")
            g.pecosa = request.form.get("pecosa", "").upper()
            g.mes = p.mes_actual
            g.anio = p.anio
            try:
                g.orden = int(request.form.get("orden", 1))
            except ValueError:
                pass
            if not gasto_id:
                duplicada = (Gasto.query
                             .filter(Gasto.tipo_doc == g.tipo_doc,
                                     Gasto.num_doc == g.num_doc,
                                     Gasto.siaf == g.siaf,
                                     Gasto.anio == g.anio)
                             .order_by(Gasto.id)
                             .first())
                if duplicada:
                    error = (f"La orden ya existe: {g.tipo_doc} N° {g.num_doc} con "
                             f"SIAF {g.siaf} ya está registrada para "
                             f"{MESES[duplicada.mes - 1]} {duplicada.anio}. "
                             f"Elimine o edite esa orden si desea modificarla; "
                             f"no puede volver a registrarse en otro mes.")
            if error is None:
                db.session.add(g)
                db.session.commit()
                flash("Gasto guardado correctamente.", "success")
                if gasto_id or request.form.get("modal"):
                    return redirect(url_for("gastos", mes=g.mes, anio=g.anio))
                return redirect(url_for("gasto_detalle_nuevo", gasto_id=g.id))
        tmpl = "_gasto_form.html" if es_modal else "gasto_form.html"
        return render_template(tmpl, g=g, p=p, MESES=MESES, error=error,
                               CLASIFICADORES=CLASIFICADORES, COMPONENTES=COMPONENTES,
                               clasificadores=lista_clasificadores(p))

    # ------------------------- DETALLES DE GASTO ----------------------
    def gasto_detalle_form(gasto_id, detalle_id):
        p = get_proyecto()
        g = db.session.get(Gasto, gasto_id)
        if g is None:
            return redirect(url_for("gastos"))
        d = (db.session.get(GastoDetalle, detalle_id) if detalle_id
             else GastoDetalle(und="UND", cantidad=1))
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        if request.method == "POST":
            d.gasto_id = g.id
            d.detalle = request.form.get("detalle", "").upper()
            d.und = request.form.get("und", "UND").upper()
            try:
                d.cantidad = float(request.form.get("cantidad", 1) or 1)
                d.precio_unitario = float(request.form.get("precio_unitario", 0) or 0)
            except ValueError:
                pass
            if not detalle_id:
                d.orden = len(g.detalles) + 1
                g.detalles.append(d)
            db.session.commit()
            flash("Material/servicio agregado correctamente.", "success")
            return redirect(url_for("gastos", mes=g.mes, anio=g.anio))
        tmpl = "_gasto_detalle_form.html" if es_modal else "gasto_detalle_form.html"
        return render_template(tmpl, g=g, d=d, p=p, MESES=MESES)

    @app.route("/gastos/<int:gasto_id>/detalle/nuevo", methods=["GET", "POST"])
    def gasto_detalle_nuevo(gasto_id):
        return gasto_detalle_form(gasto_id, None)

    @app.route("/gastos/detalle/editar/<int:detalle_id>", methods=["GET", "POST"])
    def gasto_detalle_editar(detalle_id):
        d = db.session.get(GastoDetalle, detalle_id)
        if d is None:
            return redirect(url_for("gastos"))
        return gasto_detalle_form(d.gasto_id, detalle_id)

    @app.route("/gastos/detalle/eliminar/<int:detalle_id>", methods=["POST"])
    def gasto_detalle_eliminar(detalle_id):
        d = db.session.get(GastoDetalle, detalle_id)
        if d:
            g = d.gasto
            db.session.delete(d)
            db.session.commit()
            flash("Detalle eliminado.", "info")
            return redirect(url_for("gastos", mes=g.mes, anio=g.anio))
        return redirect(url_for("gastos"))

    # ------------------------- INGRESO DE O/C Y O/S -------------------
    @app.route("/ordenes")
    def ordenes():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        lista = gastos_mes(mes, anio)
        total = round(sum(g.importe for g in lista), 2)
        cls_nombres = dict(clasificadores_proyecto())
        cls_nombres.update(CLASIFICADORES)
        bloqueos_orden = {}
        bloqueos_material = {}
        es_super = es_super_actual()
        for g in lista:
            if _orden_bloqueada(g, p):
                bloqueos_orden[g.id] = ("Esta orden pertenece a un mes cerrado (distinto al "
                                        "mes de trabajo del proyecto). No se puede eliminar.")
                for d in g.detalles:
                    bloqueos_material[d.id] = ("No se puede eliminar: la orden pertenece a un "
                                               "mes cerrado (distinto al mes de trabajo del "
                                               "proyecto).")
                continue
            if g.devengado and not es_super:
                bloqueos_orden[g.id] = ("Esta orden ya está devengada. "
                                        "No se puede eliminar.")
                for d in g.detalles:
                    bloqueos_material[d.id] = ("La orden ya está devengada. "
                                               "No se puede eliminar el material.")
                continue
            movs = _movs_ingreso_orden(g)
            if movs:
                bloqueos_orden[g.id] = ("Esta orden tiene materiales con ingresos en almacén. "
                                        "Elimine primero los movimientos de entrada del almacén.")
                descs_mov = {m.descripcion for m in movs}
                for d in g.detalles:
                    if (d.detalle or "").strip().upper() in descs_mov:
                        bloqueos_material[d.id] = ("Este material tiene ingresos en almacén. "
                                                   "Elimine primero los movimientos de entrada del almacén.")
        return render_template("ordenes.html", p=p, lista=lista, total=total,
                               mes=mes, anio=anio, MESES=MESES,
                               clasificadores=clasificadores_oc(p),
                               cls_nombres=cls_nombres,
                               bloqueos_orden=bloqueos_orden,
                               bloqueos_material=bloqueos_material,
                               COMPONENTES=COMPONENTES)

    @app.route("/ordenes/nuevo", methods=["POST"])
    def orden_nuevo():
        return orden_form(None)

    @app.route("/ordenes/editar/<int:orden_id>", methods=["GET", "POST"])
    def orden_editar(orden_id):
        return orden_form(orden_id)

    def orden_form(orden_id):
        p = get_proyecto()
        g = db.session.get(Gasto, orden_id) if orden_id else Gasto(
            fecha=fecha_defecto(p),
            mes=p.mes_actual, anio=p.anio, componente="Costo Directo",
            clasificador=(p.clasificador_bienes or "2.6.2.3.99.4"))
        if orden_id and g is None:
            abort(404)
        if g.id and _orden_bloqueada(g, p):
            flash("Esta orden pertenece a un mes cerrado. No se pueden modificar sus datos.", "error")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        if g.id and g.devengado and not es_super_actual():
            flash("Esta orden ya está devengada. No se pueden modificar sus datos.", "error")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        error = None
        if request.method == "POST":
            g.fecha = parse_fecha(request.form.get("fecha")) or g.fecha
            try:
                g.siaf = int(request.form.get("siaf", 0) or 0)
            except ValueError:
                pass
            g.tipo_doc = request.form.get("tipo_doc", "O/C").upper()
            try:
                g.num_doc = int(request.form.get("num_doc", 0) or 0)
            except ValueError:
                pass
            g.proveedor = request.form.get("proveedor", "").upper()
            g.clasificador = request.form.get("clasificador", g.clasificador)
            g.componente = request.form.get("componente", "Costo Directo")
            g.pecosa = request.form.get("pecosa", "").upper()
            try:
                if request.form.get("mes"):
                    g.mes = int(request.form.get("mes"))
                if request.form.get("anio"):
                    g.anio = int(request.form.get("anio"))
            except ValueError:
                pass
            if not orden_id and not es_super_actual() and (g.anio, g.mes) != (p.anio, p.mes_actual):
                error = (f"No se puede registrar una orden en un mes cerrado. "
                         f"El mes de trabajo actual es {MESES[p.mes_actual - 1]} {p.anio}.")
            if not orden_id:
                duplicada = (Gasto.query
                             .filter(Gasto.tipo_doc == g.tipo_doc,
                                     Gasto.num_doc == g.num_doc,
                                     Gasto.siaf == g.siaf,
                                     Gasto.anio == g.anio)
                             .order_by(Gasto.id)
                             .first())
                if duplicada:
                    error = (f"La orden ya existe: {g.tipo_doc} N° {g.num_doc} con "
                             f"SIAF {g.siaf} ya está registrada para "
                             f"{MESES[duplicada.mes - 1]} {duplicada.anio}. "
                             f"Elimine o edite esa orden si desea modificarla; "
                             f"no puede volver a registrarse en otro mes.")
            if error is None:
                db.session.add(g)
                db.session.commit()
                flash("Orden registrada correctamente.", "success")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        tmpl = "_orden_form.html" if es_modal else "orden_form.html"
        clf = clasificadores_oc(p)
        if g.id and g.clasificador and g.clasificador not in [c for c, _ in clf]:
            clf = [(g.clasificador, g.clasificador)] + clf
        return render_template(tmpl, g=g, p=p, MESES=MESES, error=error,
                               clasificadores=clf,
                               COMPONENTES=COMPONENTES)

    @app.route("/ordenes/devengar/<int:orden_id>", methods=["POST"])
    def orden_devengar(orden_id):
        g = db.session.get(Gasto, orden_id)
        if g:
            p = get_proyecto()
            if es_super_actual() or (g.mes, g.anio) == (p.mes_actual, p.anio):
                if "devengado" in request.form:
                    g.devengado = True
                    g.nota_pago = request.form.get("nota_pago", "").strip()
                    fdep = request.form.get("fecha_devengado", "")
                    if fdep:
                        try:
                            from datetime import datetime as _dt
                            g.fecha_devengado = _dt.strptime(fdep, "%Y-%m-%d").date()
                        except ValueError:
                            pass
                else:
                    g.devengado = False
                    g.nota_pago = ""
                    g.fecha_devengado = None
                db.session.commit()
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        return redirect(url_for("ordenes"))

    def _movs_ingreso_orden(g):
        """Movimientos de entrada (E) en almacén vinculados a una orden por
        N° de documento y descripción de sus materiales."""
        if not g or g.num_doc is None:
            return []
        descs = [d.detalle.strip().upper() for d in g.detalles
                 if (d.detalle or "").strip()]
        if not descs:
            return []
        return (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.tipo == "E",
                        AlmacenMovimiento.numero_doc == str(g.num_doc),
                        AlmacenMovimiento.descripcion.in_(descs))
                .all())

    def _orden_bloqueada(g, p):
        """True si la orden pertenece a un mes distinto al mes de trabajo actual
        (mes pasado o posterior, es decir un mes cerrado). En ese estado no se
        permite modificar la orden ni ingresar/editar/eliminar materiales.
        El Super Usuario queda exento."""
        if es_super_actual():
            return False
        return (g.anio, g.mes) != (p.anio, p.mes_actual)

    @app.route("/ordenes/eliminar/<int:orden_id>", methods=["POST"])
    def orden_eliminar(orden_id):
        g = db.session.get(Gasto, orden_id)
        if g:
            p = get_proyecto()
            if _orden_bloqueada(g, p):
                flash("Esta orden pertenece a un mes cerrado (distinto al mes de trabajo del "
                      "proyecto). No se puede eliminar.", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            if g.devengado and not es_super_actual():
                flash("Esta orden ya está devengada. No se puede eliminar.", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            if _movs_ingreso_orden(g):
                flash("La orden tiene materiales con ingresos en almacén. "
                      "Elimine primero los movimientos de entrada del almacén.", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            mes, anio = g.mes, g.anio
            db.session.delete(g)
            db.session.commit()
            flash("Orden eliminada.", "info")
            return redirect(url_for("ordenes", mes=mes, anio=anio))
        return redirect(url_for("ordenes"))

    @app.route("/ordenes/<int:orden_id>/materiales/nuevo", methods=["GET", "POST"])
    def orden_material_nuevo(orden_id):
        return orden_material_form(orden_id, None)

    @app.route("/ordenes/materiales/editar/<int:detalle_id>", methods=["GET", "POST"])
    def orden_material_editar(detalle_id):
        d = db.session.get(GastoDetalle, detalle_id)
        if d is None:
            return redirect(url_for("ordenes"))
        return orden_material_form(d.gasto_id, detalle_id)

    @app.route("/ordenes/materiales/eliminar/<int:detalle_id>", methods=["POST"])
    def orden_material_eliminar(detalle_id):
        d = db.session.get(GastoDetalle, detalle_id)
        if d:
            g = d.gasto
            p = get_proyecto()
            if _orden_bloqueada(g, p):
                flash("No se puede eliminar: la orden pertenece a un mes cerrado (distinto al "
                      "mes de trabajo del proyecto).", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            if g.devengado and not es_super_actual():
                flash("La orden ya está devengada. No se puede eliminar el material.", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            movs = _movs_ingreso_orden(g)
            if movs and (d.detalle or "").strip().upper() in {m.descripcion for m in movs}:
                flash("Este material tiene ingresos en almacén. "
                      "Elimine primero los movimientos de entrada del almacén.", "error")
                return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
            db.session.delete(d)
            db.session.commit()
            flash("Material eliminado.", "info")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        return redirect(url_for("ordenes"))

    def orden_material_form(orden_id, detalle_id):
        p = get_proyecto()
        g = db.session.get(Gasto, orden_id)
        if g is None:
            return redirect(url_for("ordenes"))
        if _orden_bloqueada(g, p):
            flash("Esta orden pertenece a un mes cerrado. No se puede ingresar ni modificar materiales.", "error")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        if g.devengado and not es_super_actual():
            flash("Esta orden ya está devengada. No se puede ingresar ni modificar materiales.", "error")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        d = (db.session.get(GastoDetalle, detalle_id) if detalle_id
             else GastoDetalle(und="UND", cantidad=1))
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        if request.method == "POST":
            d.gasto_id = g.id
            d.detalle = request.form.get("detalle", "").upper()
            d.und = request.form.get("und", "UND").upper()
            try:
                d.cantidad = float(request.form.get("cantidad", 1) or 1)
                d.precio_unitario = float(request.form.get("precio_unitario", 0) or 0)
            except ValueError:
                pass
            if not detalle_id:
                d.orden = len(g.detalles) + 1
                g.detalles.append(d)
            db.session.commit()
            flash("Material agregado correctamente.", "success")
            return redirect(url_for("ordenes", mes=g.mes, anio=g.anio))
        tmpl = "_orden_material_form.html" if es_modal else "orden_material_form.html"
        return render_template(tmpl, g=g, d=d, p=p, MESES=MESES)

    # ------------------------- ALMACEN --------------------------------
    @app.route("/almacen")
    def almacen():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        movs = (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.mes == mes, AlmacenMovimiento.anio == anio)
                .order_by(AlmacenMovimiento.fecha, AlmacenMovimiento.id).all())
        diario = almacen_diario(mes, anio)
        valorizado = almacen_valorizado(mes, anio)
        valor_total = {
            "cant_in": round(sum(x["cant_in"] for x in valorizado), 2),
            "valor_in": round(sum(x["valor_in"] for x in valorizado), 2),
            "cant_out": round(sum(x["cant_out"] for x in valorizado), 2),
            "valor_out": round(sum(x["valor_out"] for x in valorizado), 2),
            "saldo": round(sum(x["saldo"] for x in valorizado), 2),
            "valor_saldo": round(sum(x["valor_saldo"] for x in valorizado), 2),
        }
        return render_template("almacen.html", p=p, movs=movs, mes=mes, anio=anio,
                               MESES=MESES, almacen=almacen_items(mes, anio),
                               diario=diario, valorizado=valorizado,
                               valor_total=valor_total)

    @app.route("/almacen/nuevo", methods=["GET", "POST"])
    def almacen_nuevo():
        return almacen_form(None)

    @app.route("/almacen/editar/<int:mov_id>", methods=["GET", "POST"])
    def almacen_editar(mov_id):
        return almacen_form(mov_id)

    @app.route("/almacen/eliminar/<int:mov_id>", methods=["POST"])
    def almacen_eliminar(mov_id):
        m = db.session.get(AlmacenMovimiento, mov_id)
        if m:
            mes, anio = m.mes, m.anio
            db.session.delete(m)
            db.session.commit()
            flash("Movimiento eliminado.", "info")
            return redirect(url_for("almacen", mes=mes, anio=anio))
        return redirect(url_for("almacen"))

    def almacen_form(mov_id):
        p = get_proyecto()
        m = db.session.get(AlmacenMovimiento, mov_id) if mov_id else AlmacenMovimiento(
            fecha=fecha_defecto(p), tipo="E",
            mes=p.mes_actual, anio=p.anio)
        if mov_id and m is None:
            abort(404)
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        if not mov_id and not request.method == "POST":
            m.descripcion = request.args.get("descripcion", "").upper()
            m.und = request.args.get("und", "").upper()
            m.tipo = request.args.get("tipo", "E")
            if m.tipo == "E":
                m.responsable = p.responsable_almacen or ""
            elif not request.args.get("responsable"):
                m.responsable = ""
            if request.args.get("mes"):
                try:
                    m.mes = int(request.args.get("mes"))
                except ValueError:
                    pass
            if request.args.get("anio"):
                try:
                    m.anio = int(request.args.get("anio"))
                except ValueError:
                    pass
        error = None
        alerta = None
        bloquear_descripcion = (bool(mov_id) or request.form.get("desc_bloqueada") == "1"
                                or bool(request.args.get("descripcion")))
        if request.method == "POST":
            m.descripcion = request.form.get("descripcion", "").upper()
            m.und = request.form.get("und", "UND").upper()
            m.fecha = parse_fecha(request.form.get("fecha")) or m.fecha
            m.tipo = request.form.get("tipo", "E")
            try:
                m.cantidad = float(request.form.get("cantidad", 1) or 1)
                m.precio_unitario = float(request.form.get("precio_unitario", 0) or 0)
            except ValueError:
                pass
            m.numero_doc = request.form.get("numero_doc", "").upper()
            m.numero_siaf = request.form.get("numero_siaf", "").upper()
            m.pecosa_guia = request.form.get("pecosa_guia", "").upper()
            m.proveedor = request.form.get("proveedor", "").upper()
            m.responsable = request.form.get("responsable", "").upper()
            m.actividad = request.form.get("actividad", "").upper()
            try:
                m.mes = int(request.form.get("mes", p.mes_actual))
                m.anio = int(request.form.get("anio", p.anio))
            except ValueError:
                pass
            if m.tipo == "S":
                try:
                    cantidad_salida = float(request.form.get("cantidad", 1) or 1)
                except ValueError:
                    cantidad_salida = 1
                stock = saldo_insumo(m.descripcion, excluir_id=m.id if mov_id else None)
                if stock < 0:
                    alerta = ("ALERTA: el material \"{}\" tiene saldo NEGATIVO "
                              "({:,.2f} {}). No se puede registrar la salida. "
                              "Revise los movimientos de entrada de este material.".format(
                                  m.descripcion, stock, m.und or "UND"))
                elif stock == 0:
                    error = ("No cuenta con saldo en almacén para el material "
                             "\"{}\". Saldo disponible: 0.00 {}.".format(
                                 m.descripcion, m.und or "UND"))
                elif stock < cantidad_salida:
                    error = ("La cantidad de salida ({:,.2f} {}) supera el saldo "
                             "disponible ({:,.2f} {}) del material \"{}\".".format(
                                 cantidad_salida, m.und or "UND",
                                 stock, m.und or "UND", m.descripcion))
            if (m.tipo == "E" and m.descripcion and m.numero_doc
                    and error is None and alerta is None):
                # Bloqueo servidor: no registrar dos veces el mismo material con la
                # misma O/C (aplica en todo el proyecto, igual que el autocomplete).
                dup = (db.session.query(AlmacenMovimiento)
                       .filter(AlmacenMovimiento.tipo == "E",
                               db.func.upper(AlmacenMovimiento.descripcion) == m.descripcion,
                               AlmacenMovimiento.numero_doc == m.numero_doc))
                if mov_id and m.id:
                    dup = dup.filter(AlmacenMovimiento.id != m.id)
                dup = dup.first()
                if dup:
                    error = ("No se permite registrar dos veces el mismo material con la misma O/C. "
                             "El material \"{}\" ya fue ingresado con la O/C N° {} en {} {} de este "
                             "proyecto; use otro número de orden o corrija el movimiento existente.").format(
                                 m.descripcion, m.numero_doc,
                                 MESES[dup.mes - 1], dup.anio)
            if error is None and alerta is None:
                if not mov_id:
                    db.session.add(m)
                db.session.commit()
                flash("Movimiento de almacén guardado.", "success")
                return redirect(url_for("almacen", mes=m.mes, anio=m.anio))
        insumos = [(d, u) for d, u in
                   db.session.query(AlmacenMovimiento.descripcion, AlmacenMovimiento.und)
                   .distinct().order_by(AlmacenMovimiento.descripcion).all()]
        det_mats = [(d, u) for d, u in
                    db.session.query(GastoDetalle.detalle, GastoDetalle.und)
                    .filter(GastoDetalle.detalle != None, GastoDetalle.detalle != "")
                    .distinct().order_by(GastoDetalle.detalle).all()]
        vistos = set()
        insumos = [par for par in insumos + det_mats
                   if not (par[0] or "") in vistos and not vistos.add(par[0] or "")]
        precios = {}
        for desc, pu in db.session.query(AlmacenMovimiento.descripcion,
                                         AlmacenMovimiento.precio_unitario).distinct().all():
            precios.setdefault((desc or "").upper(), set()).add(float(pu or 0))
        for det, pu in db.session.query(GastoDetalle.detalle,
                                        GastoDetalle.precio_unitario).all():
            if det:
                precios.setdefault(det.upper(), set()).add(float(pu or 0))
        precios = {d: sorted(s) for d, s in precios.items()}
        # Pares (descripcion, numero_doc) ya ingresados en almacén (entradas E).
        # Sirve para que el autocomplete de "Nuevo material" no repita el mismo
        # material con la misma orden de compra (O/C).
        usados = {}
        for desc, num in (db.session.query(AlmacenMovimiento.descripcion,
                                          AlmacenMovimiento.numero_doc)
                          .filter(AlmacenMovimiento.tipo == "E")
                          .filter(AlmacenMovimiento.descripcion.isnot(None),
                                  AlmacenMovimiento.numero_doc.isnot(None),
                                  AlmacenMovimiento.descripcion != "",
                                  AlmacenMovimiento.numero_doc != "")
                          .distinct().all()):
            d = (desc or "").upper()
            n = (num or "").strip().upper()
            if d and n:
                usados.setdefault(d, set()).add(n)
        usados = {d: sorted(v) for d, v in usados.items()}
        tipo_fijo = None if mov_id else (request.args.get("tipo") or "E")
        oc_list = []
        saldo_material = 0.0
        if tipo_fijo == "E" and m.descripcion:
            oc_list = oc_para_material(m.descripcion)
            saldo_material = saldo_insumo(m.descripcion)
        tmpl = "_almacen_form.html" if es_modal else "almacen_form.html"
        status = 400 if (es_modal and (error or alerta)) else 200
        return render_template(tmpl, m=m, p=p, MESES=MESES,
                               insumos=insumos, precios=precios, usados=usados,
                               tipo_fijo=tipo_fijo, oc_list=oc_list,
                               saldo_material=saldo_material,
                               bloquear_descripcion=bloquear_descripcion,
                               error=error, alerta=alerta), status

    @app.route("/almacen/oc", methods=["GET"])
    def almacen_oc():
        p = get_proyecto()
        num = request.args.get("num", "").strip()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        if not num:
            return render_template("_almacen_oc.html", p=p, MESES=MESES,
                                   mes=mes, anio=anio, oc_list=[], error=None)
        try:
            num_i = int(num)
        except ValueError:
            return render_template("_oc_resultados.html", p=p, mes=mes, anio=anio,
                                   oc_list=[], error=f"N° de O/C inválido: {num}")
        oc_list = (Gasto.query
                   .filter(Gasto.tipo_doc == "O/C", Gasto.num_doc == num_i,
                           Gasto.proveedor != None, Gasto.proveedor != "")
                   .order_by(Gasto.fecha, Gasto.id).all())
        error = None if oc_list else f"No se encontró ninguna O/C con N° {num}."
        return render_template("_oc_resultados.html", p=p, mes=mes, anio=anio,
                               oc_list=oc_list, error=error)

    @app.route("/almacen/agregar_oc", methods=["POST"])
    def almacen_agregar_oc():
        p = get_proyecto()
        oc_id = request.form.get("oc_id", type=int)
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        g = db.session.get(Gasto, oc_id) if oc_id else None
        if g and g.detalles:
            responsable = (p.responsable_almacen or "").upper()
            creados = 0
            saltados = 0
            oc_num = str(g.num_doc or "").upper()
            for d in g.detalles:
                detalle = (d.detalle or "").strip().upper()
                if not detalle:
                    continue
                # Evitar duplicados: si el mismo material ya fue ingresado con
                # esta O/C no se vuelve a registrar en el almacén.
                if (db.session.query(AlmacenMovimiento.id)
                        .filter(AlmacenMovimiento.tipo == "E",
                                db.func.upper(AlmacenMovimiento.descripcion) == detalle,
                                AlmacenMovimiento.numero_doc == oc_num).first()):
                    saltados += 1
                    continue
                db.session.add(AlmacenMovimiento(
                    descripcion=detalle,
                    und=(d.und or "UND").strip().upper(),
                    fecha=g.fecha or datetime.today().date(),
                    tipo="E",
                    cantidad=d.cantidad or 0,
                    precio_unitario=d.precio_unitario or 0,
                    numero_doc=oc_num,
                    proveedor=(g.proveedor or "").strip().upper(),
                    responsable=responsable,
                    actividad="",
                    mes=mes, anio=anio))
                creados += 1
            db.session.commit()
            msg = f"Se agregaron {creados} materiales de la O/C N° {g.num_doc} al almacén."
            if saltados:
                msg += f" {saltados} material(es) ya estaban ingresados con esta O/C y se omitieron."
            flash(msg, "success")
        else:
            flash("No se pudo agregar: O/C no encontrada o sin materiales.", "error")
        return redirect(url_for("almacen", mes=mes, anio=anio))

    # ------------------------- TAREO / PLANILLA ---------------------------
    @app.route("/tareo")
    def tareo():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        calendario = calendario_mes(anio, mes)
        mes_ant, anio_ant = (12, anio - 1) if mes == 1 else (mes - 1, anio)
        obreros = (Trabajador.query
                   .filter(Trabajador.tipo == "OBRERO",
                           Trabajador.mes == mes, Trabajador.anio == anio)
                   .order_by(Trabajador.nombre).all())
        tecnicos = (Trabajador.query
                    .filter(Trabajador.tipo == "TECNICO",
                            Trabajador.mes == mes, Trabajador.anio == anio)
                    .order_by(Trabajador.nombre).all())
        resumen_obreros, total_dias_obreros = resumen_tareo(obreros, calendario)
        resumen_tecnicos, total_dias_tecnicos = resumen_tareo(tecnicos, calendario)
        puede_copiar = (not (obreros or tecnicos)
                        and Trabajador.query
                            .filter(Trabajador.mes == mes_ant,
                                    Trabajador.anio == anio_ant).count() > 0)
        return render_template("tareo.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, calendario=calendario,
                               obreros=obreros, tecnicos=tecnicos,
                               total_obreros=len(obreros),
                               total_tecnicos=len(tecnicos),
                               dev_obreros=(gasto_planilla("Costo Directo", mes, anio)
                                            is not None),
                               dev_tecnicos=(gasto_planilla("Gastos Generales", mes, anio)
                                             is not None),
                               resumen_obreros=resumen_obreros,
                               resumen_tecnicos=resumen_tecnicos,
                               total_dias_obreros=total_dias_obreros,
                               total_dias_tecnicos=total_dias_tecnicos,
                               puede_copiar=puede_copiar,
                               mes_ant=mes_ant, anio_ant=anio_ant)

    @app.route("/tareo/guardar", methods=["POST"])
    def tareo_guardar():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        # El formulario envia mes/anio en el cuerpo: se respeta lo que se guarda.
        try:
            fmes = int(request.form.get("mes", ""))
            if 1 <= fmes <= 12:
                mes = fmes
        except (TypeError, ValueError):
            pass
        try:
            fanio = int(request.form.get("anio", ""))
            if 1990 <= fanio <= 2100:
                anio = fanio
        except (TypeError, ValueError):
            pass
        tipo = (request.form.get("tipo", "") or "").strip().upper()
        if tipo not in ("OBRERO", "TECNICO"):
            tipo = None
        calendario = calendario_mes(anio, mes)
        n_dias = len(calendario)
        q = Trabajador.query.filter(Trabajador.mes == mes, Trabajador.anio == anio)
        if tipo:
            q = q.filter(Trabajador.tipo == tipo)
        trabajadores = q.all()
        guardados = 0
        for t in trabajadores:
            inicio_dia = 1
            if (t.fecha_inicio and t.fecha_inicio.year == anio
                    and t.fecha_inicio.month == mes):
                inicio_dia = max(1, t.fecha_inicio.day)
            valores = [int(x) for x in request.form.getlist(f"dias_{t.id}")
                       if x.strip().isdigit()]
            valores = {d for d in valores if inicio_dia <= d <= n_dias}
            # Domingo: solo se cuenta si la semana (L-S) trabajada esta completa.
            # La semana se evalua sobre todos los dias L-S del mes (aunque esten
            # antes de la fecha de inicio): si el trabajador empezo a mitad de
            # semana, el domingo de esa semana no se activa.
            for d in calendario:
                if not d["es_domingo"] or d["dia"] < inicio_dia:
                    continue
                lunes = max(1, d["dia"] - 6)
                requeridos = list(range(lunes, d["dia"]))
                completo = bool(requeridos) and all(x in valores for x in requeridos)
                if completo:
                    valores.add(d["dia"])
                else:
                    valores.discard(d["dia"])
            valores = sorted(valores)
            if valores != t.dias_lista:
                t.dias_lista = valores
                guardados += 1
        db.session.commit()
        flash(f"Tareo de {MESES[mes - 1]} {anio} guardado "
              f"({guardados} registros actualizados).", "success")
        return redirect(url_for("tareo", mes=mes, anio=anio))

    @app.route("/tareo/copiar", methods=["POST"])
    def tareo_copiar():
        """Copia el personal (datos personales) del mes anterior al mes
        seleccionado. Solo se copian los datos de personal: no dias trabajados,
        ni cantidad de dias, ni cajas marcadas (checks)."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        # El formulario envia mes/anio en el cuerpo: se respeta lo que se copia.
        try:
            fmes = int(request.form.get("mes", ""))
            if 1 <= fmes <= 12:
                mes = fmes
        except (TypeError, ValueError):
            pass
        try:
            fanio = int(request.form.get("anio", ""))
            if 1990 <= fanio <= 2100:
                anio = fanio
        except (TypeError, ValueError):
            pass
        mes_ant, anio_ant = (12, anio - 1) if mes == 1 else (mes - 1, anio)
        if Trabajador.query.filter(Trabajador.mes == mes,
                                   Trabajador.anio == anio).count() > 0:
            flash(f"El mes de {MESES[mes - 1]} {anio} ya tiene personal "
                  "registrado. No se copió.", "warning")
            return redirect(url_for("tareo", mes=mes, anio=anio))
        previos = (Trabajador.query
                   .filter(Trabajador.mes == mes_ant, Trabajador.anio == anio_ant)
                   .all())
        copiados = 0
        for t in previos:
            db.session.add(Trabajador(
                tipo=t.tipo, nombre=t.nombre, dni=t.dni,
                fecha_nacimiento=t.fecha_nacimiento, cargo=t.cargo,
                sexo=t.sexo, fecha_inicio=t.fecha_inicio, aporte=t.aporte,
                sueldo_mensual=t.sueldo_mensual,
                mes=mes, anio=anio))
            copiados += 1
        db.session.commit()
        flash(f"Personal de {MESES[mes_ant - 1]} {anio_ant} copiado a "
              f"{MESES[mes - 1]} {anio} ({copiados} registros). Solo datos "
              "personales: los días trabajados se ingresan en el tareo.",
              "success")
        return redirect(url_for("tareo", mes=mes, anio=anio))

    def gasto_planilla(comp, mes, anio):
        """Gasto PLLA (planilla de pago del tareo) del panel indicado."""
        return (Gasto.query
                .filter(Gasto.tipo_doc == "PLLA", Gasto.componente == comp,
                        Gasto.mes == mes, Gasto.anio == anio)
                .first())

    @app.route("/tareo/devengar", methods=["GET", "POST"])
    def tareo_devengar():
        """Popup de devengado de la planilla: crea/actualiza el gasto PLLA y
        activa el flag devengado de los trabajadores del panel."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        tipo = ((request.args.get("tipo") or request.form.get("tipo", ""))
                .strip().upper())
        if tipo not in ("OBRERO", "TECNICO"):
            tipo = "OBRERO"
        comp = "Costo Directo" if tipo == "OBRERO" else "Gastos Generales"
        clasif = p.clasificador_personal or "2.6.2.3.99.3"
        trabajadores = (Trabajador.query
                        .filter(Trabajador.tipo == tipo, Trabajador.mes == mes,
                                Trabajador.anio == anio)
                        .order_by(Trabajador.nombre).all())
        g = gasto_planilla(comp, mes, anio)
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        error = None
        if request.method == "POST":
            fecha = parse_fecha(request.form.get("fecha"))
            try:
                siaf = int(request.form.get("siaf", 0) or 0)
            except ValueError:
                siaf = 0
            try:
                num_doc = int(request.form.get("num_doc", 0) or 0)
            except ValueError:
                num_doc = 0
            try:
                monto = round(float(request.form.get("monto", 0) or 0), 2)
            except (TypeError, ValueError):
                monto = 0
            proveedor = (request.form.get("proveedor", "") or "").strip().upper()
            if not fecha:
                error = "Indique la fecha de devengado."
            elif not proveedor:
                error = "Indique el nombre / proveedor de la planilla."
            elif monto <= 0:
                error = "Indique el Monto de Planilla (mayor a S/ 0.00)."
            if error is None:
                if g is None:
                    n_orden = (db.session.query(db.func.coalesce(
                        db.func.max(Gasto.orden), 0)).scalar()) + 1
                    g = Gasto(tipo_doc="PLLA", num_doc=num_doc, clasificador=clasif,
                              componente=comp, mes=mes, anio=anio,
                              devengado=True, orden=n_orden)
                    db.session.add(g)
                g.fecha = fecha
                g.siaf = siaf
                g.num_doc = num_doc
                g.proveedor = proveedor
                g.devengado = True
                if not g.detalles:
                    g.detalles.append(GastoDetalle(orden=1))
                d = g.detalles[0]
                d.detalle = (f"PLANILLA DE PAGO CORRESPONDIENTE AL MES DE "
                             f"{MESES[mes - 1]} {anio}")
                d.und = "PLLA"
                d.cantidad = 1
                d.precio_unitario = monto
                for t in trabajadores:
                    if not t.devengado:
                        t.devengado = True
                db.session.commit()
                flash(f"Planilla de {tipo.capitalize()} devengada por "
                      f"S/ {monto:,.2f}.", "success")
                return redirect(url_for("tareo", mes=mes, anio=anio))
        tmpl = "_tareo_devengar_form.html" if es_modal else "tareo_devengar.html"
        if error:
            return render_template(tmpl, p=p, mes=mes, anio=anio, tipo=tipo,
                                   g=g, MESES=MESES, error=error), 400
        return render_template(tmpl, p=p, mes=mes, anio=anio, tipo=tipo,
                               g=g, MESES=MESES, error=error)

    @app.route("/tareo/desdevengar", methods=["POST"])
    def tareo_desdevengar():
        """Retira el devengado de la planilla: elimina el gasto PLLA del panel."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        tipo = (request.form.get("tipo", "") or "").strip().upper()
        if tipo not in ("OBRERO", "TECNICO"):
            return redirect(url_for("tareo", mes=mes, anio=anio))
        comp = "Costo Directo" if tipo == "OBRERO" else "Gastos Generales"
        g = gasto_planilla(comp, mes, anio)
        if g:
            db.session.delete(g)
        for t in (Trabajador.query
                  .filter(Trabajador.tipo == tipo, Trabajador.mes == mes,
                          Trabajador.anio == anio).all()):
            if t.devengado:
                t.devengado = False
        db.session.commit()
        flash(f"Devengado de la planilla {tipo.capitalize()} retirado.", "info")
        return redirect(url_for("tareo", mes=mes, anio=anio))

    @app.route("/tareo/imprimir")
    def tareo_imprimir():
        """Vista imprimible institucional del tareo/planilla mensual."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        tipo = request.args.get("tipo", "").strip().upper()
        if tipo not in ("OBRERO", "TECNICO"):
            tipo = ""
        calendario = calendario_mes(anio, mes)
        obreros = (Trabajador.query
                   .filter(Trabajador.tipo == "OBRERO",
                           Trabajador.mes == mes, Trabajador.anio == anio)
                   .order_by(Trabajador.nombre).all())
        tecnicos = (Trabajador.query
                    .filter(Trabajador.tipo == "TECNICO",
                            Trabajador.mes == mes, Trabajador.anio == anio)
                    .order_by(Trabajador.nombre).all())
        if tipo == "OBRERO":
            tecnicos = []
        elif tipo == "TECNICO":
            obreros = []
        resumen_obreros, total_dias_obreros = resumen_tareo(obreros, calendario)
        resumen_tecnicos, total_dias_tecnicos = resumen_tareo(tecnicos, calendario)
        return render_template("tareo_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, calendario=calendario,
                               obreros=obreros, tecnicos=tecnicos,
                               resumen_obreros=resumen_obreros,
                               resumen_tecnicos=resumen_tecnicos,
                               total_dias_obreros=total_dias_obreros,
                               total_dias_tecnicos=total_dias_tecnicos)

    @app.route("/tareo/planilla/opciones")
    def planilla_opciones():
        """Popup con las dos modalidades de generacion de la planilla de pagos
        para la seccion activa (obrero o tecnico)."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        seccion = request.args.get("seccion", "obrero").strip().lower()
        if seccion not in ("obrero", "tecnico"):
            seccion = "obrero"
        return render_template("_planilla_opciones.html", p=p, mes=mes, anio=anio,
                               seccion=seccion,
                               MESES=MESES,
                               TABLA_CIVIL_POR_ANIO=TABLA_CIVIL_POR_ANIO)

    @app.route("/tareo/planilla/imprimir")
    def planilla_imprimir():
        """Planilla de pagos imprimible por seccion (obrero o tecnico),
        con o sin beneficios/descuentos de ley."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        opcion = request.args.get("opcion", "sin").strip().lower()
        if opcion not in ("sin", "con"):
            opcion = "sin"
        seccion = request.args.get("seccion", "obrero").strip().lower()
        if seccion not in ("obrero", "tecnico"):
            seccion = "obrero"
        con_beneficios = opcion == "con"
        calendario = calendario_mes(anio, mes)
        filas_obreros, filas_tecnicos = [], []
        tot_obreros = {"ingresos": 0, "descuentos": 0, "neto": 0}
        tot_tecnicos = {"ingresos": 0, "descuentos": 0, "neto": 0}
        if seccion == "obrero":
            obreros = (Trabajador.query
                       .filter(Trabajador.tipo == "OBRERO",
                               Trabajador.mes == mes, Trabajador.anio == anio)
                       .order_by(Trabajador.nombre).all())
            filas_obreros = [calcular_obrero(t, calendario, anio, mes, con_beneficios)
                             for t in obreros]
            tot_obreros = {
                "ingresos": round(sum(f["ingresos"] for f in filas_obreros), 2),
                "descuentos": round(sum(f["descuentos"] for f in filas_obreros), 2),
                "neto": round(sum(f["neto"] for f in filas_obreros), 2),
            }
        else:
            tecnicos = (Trabajador.query
                        .filter(Trabajador.tipo == "TECNICO",
                                Trabajador.mes == mes, Trabajador.anio == anio)
                        .order_by(Trabajador.nombre).all())
            filas_tecnicos = [calcular_tecnico(t, calendario, anio, con_beneficios)
                              for t in tecnicos]
            tot_tecnicos = {
                "ingresos": round(sum(f["ingresos"] for f in filas_tecnicos), 2),
                "descuentos": round(sum(f["descuentos"] for f in filas_tecnicos), 2),
                "neto": round(sum(f["neto"] for f in filas_tecnicos), 2),
            }
        return render_template("planilla_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, calendario=calendario,
                               seccion=seccion, con_beneficios=con_beneficios,
                               filas_obreros=filas_obreros,
                               filas_tecnicos=filas_tecnicos,
                               tot_obreros=tot_obreros, tot_tecnicos=tot_tecnicos,
                               tabla=tabla_civil(anio),
                               TABLA_CIVIL_POR_ANIO=TABLA_CIVIL_POR_ANIO)

    @app.route("/tareo/nuevo", methods=["GET", "POST"])
    def trabajador_nuevo():
        return trabajador_form(None)

    @app.route("/tareo/editar/<int:tid>", methods=["GET", "POST"])
    def trabajador_editar(tid):
        return trabajador_form(tid)

    @app.route("/tareo/eliminar/<int:tid>", methods=["POST"])
    def trabajador_eliminar(tid):
        t = db.session.get(Trabajador, tid)
        if t:
            mes, anio, tipo = t.mes, t.anio, t.tipo
            db.session.delete(t)
            db.session.commit()
            flash(f"Registro de {t.nombre} eliminado del tareo.", "info")
            return redirect(url_for("tareo", mes=mes, anio=anio))
        return redirect(url_for("tareo"))

    def trabajador_form(tid):
        p = get_proyecto()
        t = (db.session.get(Trabajador, tid) if tid
             else Trabajador(tipo=(request.args.get("tipo") or "OBRERO"),
                             fecha_nacimiento=None, fecha_inicio=None,
                             mes=p.mes_actual, anio=p.anio))
        if tid and t is None:
            abort(404)
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        error = None
        if request.method == "POST":
            mes_previo, anio_previo = (t.mes, t.anio) if tid else (None, None)
            t.tipo = (request.form.get("tipo", "OBRERO") or "OBRERO").upper()
            cargo_previo = t.cargo if tid else None
            t.nombre = request.form.get("nombre", "").strip().upper()
            t.dni = request.form.get("dni", "").strip()
            t.cargo = request.form.get("cargo", "").strip().upper()
            t.sexo = (request.form.get("sexo", "M") or "M").upper()
            t.fecha_nacimiento = parse_fecha(request.form.get("fecha_nacimiento"))
            t.fecha_inicio = parse_fecha(request.form.get("fecha_inicio"))
            try:
                t.mes = int(request.form.get("mes", p.mes_actual))
                t.anio = int(request.form.get("anio", p.anio))
            except ValueError:
                pass
            t.aporte = (request.form.get("aporte", "AFP") or "AFP").strip().upper()
            if t.aporte not in ("AFP", "ONP"):
                t.aporte = "AFP"
            try:
                t.sueldo_mensual = round(float(request.form.get("sueldo_mensual", 0) or 0), 2)
            except (TypeError, ValueError):
                t.sueldo_mensual = 0.0
            cargos_validos = CARGOS_OBRERO if t.tipo == "OBRERO" else CARGOS_TECNICO
            if not t.nombre:
                error = "Debe indicar el nombre completo del trabajador."
            elif not t.dni or not t.dni.isdigit() or len(t.dni) != 8:
                error = "El DNI debe tener 8 dígitos numéricos."
            elif t.tipo not in ("OBRERO", "TECNICO"):
                error = "Tipo de personal no válido."
            elif t.cargo not in cargos_validos and (not tid or t.cargo != cargo_previo):
                error = (f"El cargo debe ser uno de: {', '.join(cargos_validos)}.")
            elif (t.fecha_inicio and t.anio and t.mes
                    and (t.fecha_inicio.year > t.anio
                         or (t.fecha_inicio.year == t.anio
                             and t.fecha_inicio.month > t.mes))):
                error = "La fecha de inicio no puede ser posterior al mes tareado."
            else:
                if tid and (mes_previo, anio_previo) != (t.mes, t.anio):
                    t.dias_lista = []
                if t.fecha_inicio and t.fecha_inicio.year == t.anio and t.fecha_inicio.month == t.mes:
                    n_dias = len(calendario_mes(t.anio, t.mes))
                    t.dias_lista = [d for d in t.dias_lista if d >= t.fecha_inicio.day and d <= n_dias]
                duplicado = (Trabajador.query
                             .filter(Trabajador.dni == t.dni,
                                     Trabajador.mes == t.mes,
                                     Trabajador.anio == t.anio,
                                     Trabajador.id != t.id)
                             .first())
                if duplicado:
                    error = (f"El DNI {t.dni} ({duplicado.nombre}) ya está "
                             f"registrado en {MESES[t.mes - 1]} {t.anio}.")
            if error is None:
                if not tid:
                    db.session.add(t)
                db.session.commit()
                flash(f"Registro de {t.nombre} guardado en el tareo.", "success")
                return redirect(url_for("tareo", mes=t.mes, anio=t.anio))
        status = 400 if (es_modal and error) else 200
        return render_template("_trabajador_form.html", t=t, p=p, MESES=MESES,
                               CARGOS_OBRERO=CARGOS_OBRERO,
                               CARGOS_TECNICO=CARGOS_TECNICO,
                               error=error), status

    # ------------------------- FORMATOS -------------------------------
    @app.route("/formatos")
    def formatos():
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        secciones, total, clasif_cols = f05_datos(mes, anio)
        rows = fe06_rows()
        resumen = fe06_resumen(rows)
        fe06_totales_mensual = [round(sum(r["mensual"][i] for r in resumen.values()), 2)
                                for i in range(12)]
        sintesis = fe06_sintesis(mes, anio)
        return render_template("formatos.html", p=p, mes=mes, anio=anio, MESES=MESES,
                               secciones=secciones, f05_total=total,
                               clasif_cols=clasif_cols,
                               fe06_rows=rows, fe06_resumen=resumen,
                               fe06_totales_mensual=fe06_totales_mensual,
                               fe06_sintesis=sintesis,
                               meses_activos=meses_con_ejecucion(anio),
                               meses_vis=meses_visibles(anio, mes))

    @app.route("/formatos/actividades", methods=["GET", "POST"])
    def formatos_actividades():
        """Gestiona la seccion II (Actividades Ejecutadas) del Resumen Financiero.

        GET (modal) lista las actividades del mes/anio; POST reemplaza la lista
        completa: se conservan los ids existentes, se actualizan las
        descripciones editadas y se eliminan las marcadas para quitar.
        """
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        if request.method == "POST":
            ids = request.form.getlist("actividad_id")
            descs = request.form.getlist("actividad_desc")
            ActividadEjecutada.query.filter_by(mes=mes, anio=anio).delete()
            orden = 0
            for i, desc in enumerate(descs):
                desc = (desc or "").strip()
                if not desc:
                    continue
                orden += 1
                aid = ids[i] if i < len(ids) else ""
                a = (db.session.get(ActividadEjecutada, int(aid))
                     if aid.isdigit() else None)
                if a is None:
                    a = ActividadEjecutada(mes=mes, anio=anio)
                a.descripcion = desc
                a.orden = orden
                db.session.add(a)
            db.session.commit()
            flash("Actividades ejecutadas guardadas correctamente.", "success")
            return redirect(url_for("formatos", mes=mes, anio=anio))
        es_modal = bool(request.headers.get("X-Modal")) or request.args.get("modal")
        if not es_modal:
            return redirect(url_for("formatos", mes=mes, anio=anio))
        return render_template("_actividades_form.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, actividades=actividades_mes(mes, anio))

    @app.route("/formatos/manifiesto/imprimir", methods=["GET"])
    def imprimir_manifiesto():
        """Plantilla imprimible institucional del Manifiesto de Gasto con datos reales."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio

        # Cabecera institucional y datos generales
        ubicacion = " - ".join(x for x in [p.distrito, p.provincia, p.departamento] if x)
        # Correlativo del manifiesto: inicia en el mes en que se inicio el ingreso
        # de datos (primer mes con gastos devengados del anio) y aumenta cada mes
        # (Marzo N° 001, Abril N° 002, Mayo N° 003, Junio N° 004, ...).
        n_manifiesto = max(1, mes - mes_inicio_manifiesto(anio) + 1)
        hdr = {
            "entidad": p.entidad,
            "unidad": p.unidad_ejecutora,
            "proyecto": p.nombre,
            "cui": p.cui,
            "meta": p.meta,
            "fuente": p.fuente,
            "ubicacion": ubicacion,
            "modalidad": "ADMINISTRACION DIRECTA",
            "residente": p.residente,
            "supervisor": p.supervisor,
            "asistente": p.asistente,
            "periodo": f"{MESES[mes-1]} - {anio}",
            "n_manifiesto": f"{n_manifiesto:03d}",
            "logo_path": p.logo_path or "",
            "cip_supervisor": p.cip_supervisor or "",
            "cip_residente": p.cip_residente or "",
        }

        # Filas financieras agrupadas por rubro (componente) y clasificador
        orden_comp = ["Costo Directo", "Gastos Generales", "Gastos de Supervisión",
                      "Elaboración de Expediente Técnico", "Liquidación de Obra"]
        grupos = {}
        for g in gastos_mes(mes, anio, devengado=True):
            key = (g.componente or "", g.clasificador or "")
            gr = grupos.setdefault(key, {"componente": g.componente or "",
                                         "clasificador": g.clasificador or "",
                                         "filas": [], "subtotal": 0.0})
            # Misma orden (SIAF + tipo + numero + proveedor) = una sola linea de proveedor.
            clave_orden = (g.siaf or "", g.tipo_doc or "", g.num_doc or "",
                           g.proveedor or "")
            first = gr.get("_ultima_orden") != clave_orden
            gr["_ultima_orden"] = clave_orden
            for idx_d, d in enumerate(g.detalles):
                gr["filas"].append({
                    "fecha": g.fecha.strftime("%d/%m/%Y") if g.fecha else "",
                    "siaf": g.siaf or "", "tipo_doc": g.tipo_doc or "",
                    "num_doc": g.num_doc or "", "proveedor": g.proveedor or "",
                    "detalle": d.detalle, "und": d.und or "",
                    "cantidad": d.cantidad, "pu": d.precio_unitario,
                    "importe": d.importe, "prov_first": first and idx_d == 0,
                })
                gr["subtotal"] += d.importe

        def orden_seccion(item):
            comp, clas = item
            ci = orden_comp.index(comp) if comp in orden_comp else len(orden_comp)
            return (ci, clas)

        secciones = []
        for key in sorted(grupos, key=orden_seccion):
            gr = grupos[key]
            gr.pop("_ultima_orden", None)
            gr["subtotal"] = round(gr["subtotal"], 2)
            gr["label"] = (f"{gr['componente'].upper()} - "
                           f"{cls_nombre(p, gr['clasificador'])}")
            prov_cont = 0
            for fila in gr["filas"]:
                if fila["prov_first"]:
                    prov_cont += 1
                fila["prov_num"] = prov_cont
            secciones.append(gr)

        total_gral = round(total_gastos_mes(mes, anio, devengado=True), 2)
        return render_template("manifiesto_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, hdr=hdr, secciones=secciones,
                               total_gral=total_gral)

    # ── Helper para headers de Excel (coincidentes con print.css) ──
    def _build_xl_header(ws, ncols, title, subtitle, data, logosrc,
                         info_pairs, responsables, extra_right=None,
                         title_font_size=16, logo_w=60, logo_h=60,
                         header_heights=None, info_cols=None):
        """Encabezado Excel limpio en 2 secciones (como hoja de impresión).

        header_heights: dict opcional {row_key: height} para sobreescribir alturas.
            Keys: 'entity'=row1, 'sub'=row2, 'title'=row3, 'subtitle'=row4,
                  'project'=row6 info, 'pair'=rows7-8, 'resp'=rows responsables.
        """
        import math, os as _os
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        FN = "Calibri"
        AZUL = "1E3A5F"; AZUL_OSC = "162D45"; BG = "F8FAFC"; TEXTO2 = "64748B"
        BLANCO = "FFFFFF"; BORDE = "CBD5E1"

        thin_o = Side(style="thin", color=AZUL_OSC)
        border_all = Border(left=thin_o, right=thin_o, top=thin_o, bottom=thin_o)
        thin_g = Side(style="thin", color=BORDE)
        border_grid = Border(left=thin_g, right=thin_g, top=thin_g, bottom=thin_g)

        fill_azul = PatternFill("solid", fgColor=AZUL)
        fill_bg = PatternFill("solid", fgColor=BG)
        fill_resp = PatternFill("solid", fgColor="E8EDF3")

        font_entity = Font(name=FN, size=11, bold=True, color=AZUL_OSC)
        font_sub = Font(name=FN, size=9, color=TEXTO2)
        font_title = Font(name=FN, size=title_font_size, bold=True, color=AZUL)
        font_label = Font(name=FN, size=8, bold=True, color=AZUL_OSC)
        font_val = Font(name=FN, size=9)
        font_resp_lab = Font(name=FN, size=7, bold=True, color=BLANCO)
        font_resp_val = Font(name=FN, size=9)
        font_extra = Font(name=FN, size=9, bold=True, color=AZUL)

        align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_nw = Alignment(horizontal="left", vertical="center", wrap_text=False)
        align_r = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # ── Logo ──
        if logosrc and _os.path.isfile(logosrc):
            try:
                img = XlImage(logosrc)
                img.width = logo_w; img.height = logo_h
                ws.add_image(img, "A1")
            except Exception:
                pass

        pcol = get_column_letter(ncols - 1)
        lcol = get_column_letter(ncols)

        # ── SECCIÓN 1: Encabezado limpio sin bordes alrededor del logo ──
        ws.merge_cells(f"B1:{pcol}1")
        c = ws["B1"]; c.value = data.get("entidad", "")
        c.font = font_entity
        c.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

        ws.merge_cells(f"B2:{pcol}2")
        c = ws["B2"]; c.value = data.get("unidad", "")
        c.font = font_sub; c.alignment = align_c

        ws.merge_cells(f"B3:{pcol}3")
        c = ws["B3"]; c.value = title.upper()
        c.font = font_title; c.alignment = Alignment(horizontal="center", vertical="top")

        ws.merge_cells(f"B4:{pcol}4")
        c = ws["B4"]; c.value = subtitle.upper()
        c.font = font_sub; c.alignment = align_c

        # Periodo / extra a la derecha — sin borde en encabezado
        if extra_right:
            for rn, (val, fnt) in extra_right.items():
                c = ws.cell(row=rn, column=ncols)
                c.value = val; c.font = fnt or font_extra
                c.alignment = align_r
        elif data.get("periodo"):
            ws.merge_cells(f"{lcol}2:{lcol}3")
            c = ws[f"{lcol}2"]; c.value = data["periodo"]
            c.font = Font(name=FN, size=9, bold=True, color=AZUL)
            c.alignment = align_r

        # Alturas de encabezado (configurables)
        hh = header_heights or {}
        ws.row_dimensions[1].height = hh.get('entity', 28)
        ws.row_dimensions[2].height = hh.get('sub', 18)
        ws.row_dimensions[3].height = hh.get('title', 32)
        ws.row_dimensions[4].height = hh.get('subtitle', 18)

        # ── SECCIÓN 2: Info-grid con cards bordeadas ──
        info_start = 6
        mid = ncols // 2
        mid_col = get_column_letter(mid)

        for i, row_items in enumerate(info_pairs):
            r = info_start + i
            n = len(row_items)
            if info_cols == "left":
                # Labels A:C, values D+ (no wrap) — FE-05
                if n == 1:
                    lbl, val = row_items[0]
                    ws.merge_cells(f"A{r}:C{r}")
                    c_l = ws.cell(row=r, column=1, value=lbl)
                    c_l.font = font_label; c_l.alignment = align_nw
                    c_l.fill = fill_bg; c_l.border = border_grid
                    for cc in [2, 3]:
                        ws.cell(row=r, column=cc).fill = fill_bg
                        ws.cell(row=r, column=cc).border = border_grid
                    ws.merge_cells(f"D{r}:{lcol}{r}")
                    c_v = ws.cell(row=r, column=4, value=val)
                    c_v.font = font_val; c_v.alignment = align_nw; c_v.border = border_grid
                    for cc in range(5, ncols + 1):
                        ws.cell(row=r, column=cc).border = border_grid
                else:
                    # Dos mitades: A:C label, D:mid val | mid+1:mid+3 label, mid+4:lastcol val
                    mid = (ncols + 1) // 2
                    for idx, (lbl, val) in enumerate(row_items):
                        if idx == 0:
                            lbl_s, lbl_e = 1, 3
                            val_s, val_e = 4, mid
                        else:
                            lbl_s, lbl_e = mid + 1, mid + 3
                            val_s, val_e = mid + 4, ncols
                        cl1, cl2 = get_column_letter(lbl_s), get_column_letter(lbl_e)
                        cv1, cv2 = get_column_letter(val_s), get_column_letter(val_e)
                        ws.merge_cells(f"{cl1}{r}:{cl2}{r}")
                        c_l = ws.cell(row=r, column=lbl_s, value=lbl)
                        c_l.font = font_label; c_l.alignment = align_nw
                        c_l.fill = fill_bg; c_l.border = border_grid
                        for cc in range(lbl_s + 1, lbl_e + 1):
                            ws.cell(row=r, column=cc).fill = fill_bg
                            ws.cell(row=r, column=cc).border = border_grid
                        ws.merge_cells(f"{cv1}{r}:{cv2}{r}")
                        c_v = ws.cell(row=r, column=val_s, value=val)
                        c_v.font = font_val; c_v.alignment = align_nw; c_v.border = border_grid
                        for cc in range(val_s + 1, val_e + 1):
                            ws.cell(row=r, column=cc).border = border_grid
                ws.row_dimensions[r].height = hh.get('pair', 17)
            elif n == 1:
                # Ancho completo: label A:B, valor C:lastcol
                lbl, val = row_items[0]
                ws.merge_cells(f"A{r}:B{r}")
                c_l = ws.cell(row=r, column=1, value=lbl)
                c_l.font = font_label; c_l.alignment = align_l
                c_l.fill = fill_bg; c_l.border = border_grid
                ws.cell(row=r, column=2).fill = fill_bg
                ws.cell(row=r, column=2).border = border_grid
                ws.merge_cells(f"C{r}:{lcol}{r}")
                c_v = ws.cell(row=r, column=3, value=val)
                c_v.font = font_val; c_v.alignment = align_l; c_v.border = border_grid
                for cc in range(4, ncols + 1):
                    ws.cell(row=r, column=cc).border = border_grid
                v_len = len(str(val or ""))
                nlines = max(1, math.ceil(v_len / 50))
                ws.row_dimensions[r].height = hh.get('project', max(17, nlines * 14))
            else:
                # Dos mitades: par de cards lado a lado
                # Izq: label A:B, valor C:mid   |   Der: label (mid+1):(mid+2), valor (mid+3):lastcol
                mid = (ncols + 1) // 2  # split justo (ej: 6 para 11 cols, 10 para 19)
                for idx, (lbl, val) in enumerate(row_items):
                    if idx == 0:
                        col_lbl_s, col_lbl_e = 1, 2
                        col_val_s, col_val_e = 3, mid
                    else:
                        col_lbl_s, col_lbl_e = mid + 1, mid + 2
                        col_val_s, col_val_e = mid + 3, ncols
                    cl1 = get_column_letter(col_lbl_s)
                    cl2 = get_column_letter(col_lbl_e)
                    cv1 = get_column_letter(col_val_s)
                    cv2 = get_column_letter(col_val_e)
                    ws.merge_cells(f"{cl1}{r}:{cl2}{r}")
                    c_l = ws.cell(row=r, column=col_lbl_s, value=lbl)
                    c_l.font = font_label; c_l.alignment = align_l
                    c_l.fill = fill_bg; c_l.border = border_grid
                    ws.cell(row=r, column=col_lbl_e).fill = fill_bg
                    ws.cell(row=r, column=col_lbl_e).border = border_grid
                    ws.merge_cells(f"{cv1}{r}:{cv2}{r}")
                    c_v = ws.cell(row=r, column=col_val_s, value=val)
                    c_v.font = font_val; c_v.alignment = align_l; c_v.border = border_grid
                    for cc in range(col_val_s + 1, col_val_e + 1):
                        ws.cell(row=r, column=cc).border = border_grid
                ws.row_dimensions[r].height = hh.get('pair', 17)

        # ── Responsables ──
        resp_start = info_start + len(info_pairs)
        resp_col_end = min(5, ncols)
        resp_merge_end = get_column_letter(resp_col_end)
        for i, (lab, val) in enumerate(responsables):
            r = resp_start + i
            if info_cols == "left":
                # Labels A:C, values D:lastcol (sin salto) — FE-05
                ws.merge_cells(f"A{r}:C{r}")
                c_l = ws.cell(row=r, column=1, value=lab)
                c_l.font = font_resp_lab; c_l.alignment = align_nw
                c_l.fill = fill_azul; c_l.border = border_all
                for cc in [2, 3]:
                    ws.cell(row=r, column=cc).fill = fill_azul
                    ws.cell(row=r, column=cc).border = border_all
                ws.merge_cells(f"D{r}:{lcol}{r}")
                c_v = ws.cell(row=r, column=4, value=val)
                c_v.font = font_resp_val; c_v.alignment = align_nw
                c_v.fill = fill_resp; c_v.border = border_all
                for cc in range(5, ncols + 1):
                    ws.cell(row=r, column=cc).fill = fill_resp
                    ws.cell(row=r, column=cc).border = border_all
            else:
                ws.merge_cells(f"A{r}:B{r}")
                c_l = ws.cell(row=r, column=1, value=lab)
                c_l.font = font_resp_lab; c_l.alignment = align_l
                c_l.fill = fill_azul; c_l.border = border_all
                ws.cell(row=r, column=2).fill = fill_azul
                ws.cell(row=r, column=2).border = border_all
                ws.merge_cells(f"C{r}:{resp_merge_end}{r}")
                c_v = ws.cell(row=r, column=3, value=val)
                c_v.font = font_resp_val; c_v.alignment = align_l
                c_v.fill = fill_resp; c_v.border = border_all
            for cc in range(4, resp_col_end + 1):
                ws.cell(row=r, column=cc).fill = fill_resp
                ws.cell(row=r, column=cc).border = border_all
            for cc in range(resp_col_end + 1, ncols + 1):
                ws.cell(row=r, column=cc).border = border_grid
            ws.row_dimensions[r].height = hh.get('resp', 19)

        return resp_start + len(responsables) + 1

    @app.route("/formatos/manifiesto/excel")
    def manifiesto_excel():
        """Genera archivo .xlsx del Manifiesto de Gasto fiel a la hoja de impresión."""
        import math
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                     numbers)
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter

        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio

        ubicacion = " - ".join(x for x in [p.distrito, p.provincia, p.departamento] if x)
        n_manifiesto = max(1, mes - mes_inicio_manifiesto(anio) + 1)
        hdr = {
            "entidad": p.entidad or "",
            "unidad": p.unidad_ejecutora or "",
            "proyecto": p.nombre or "",
            "cui": p.cui or "",
            "meta": p.meta or "",
            "fuente": p.fuente or "",
            "ubicacion": ubicacion,
            "modalidad": "ADMINISTRACION DIRECTA",
            "residente": p.residente or "",
            "supervisor": p.supervisor or "",
            "asistente": p.asistente or "",
            "periodo": f"{MESES[mes-1]} - {anio}",
            "n_manifiesto": f"{n_manifiesto:03d}",
            "logo_path": p.logo_path or "",
            "cip_supervisor": p.cip_supervisor or "",
            "cip_residente": p.cip_residente or "",
        }

        orden_comp = ["Costo Directo", "Gastos Generales", "Gastos de Supervisión",
                      "Elaboración de Expediente Técnico", "Liquidación de Obra"]
        grupos = {}
        for g in gastos_mes(mes, anio, devengado=True):
            key = (g.componente or "", g.clasificador or "")
            gr = grupos.setdefault(key, {"componente": g.componente or "",
                                         "clasificador": g.clasificador or "",
                                         "filas": [], "subtotal": 0.0})
            clave_orden = (g.siaf or "", g.tipo_doc or "", g.num_doc or "",
                           g.proveedor or "")
            first = gr.get("_ultima_orden") != clave_orden
            gr["_ultima_orden"] = clave_orden
            for idx_d, d in enumerate(g.detalles):
                gr["filas"].append({
                    "fecha": g.fecha, "siaf": g.siaf or "", "tipo_doc": g.tipo_doc or "",
                    "num_doc": g.num_doc or "", "proveedor": g.proveedor or "",
                    "detalle": d.detalle, "und": d.und or "",
                    "cantidad": d.cantidad, "pu": d.precio_unitario,
                    "importe": d.importe, "prov_first": first and idx_d == 0,
                })
                gr["subtotal"] += d.importe

        def orden_seccion(item):
            comp, clas = item
            ci = orden_comp.index(comp) if comp in orden_comp else len(orden_comp)
            return (ci, clas)

        secciones = []
        for key in sorted(grupos, key=orden_seccion):
            gr = grupos[key]
            gr.pop("_ultima_orden", None)
            gr["subtotal"] = round(gr["subtotal"], 2)
            gr["label"] = (f"{gr['componente'].upper()} - "
                           f"{cls_nombre(p, gr['clasificador'])}")
            prov_cont = 0
            for fila in gr["filas"]:
                if fila["prov_first"]:
                    prov_cont += 1
                fila["prov_num"] = prov_cont
            secciones.append(gr)

        total_gral = round(total_gastos_mes(mes, anio, devengado=True), 2)

        # ── Colores fieles al print.css ──
        AZUL = "1E3A5F"
        AZUL_OSC = "162D45"
        AZUL_CLARO = "2C527A"
        BG = "F8FAFC"
        TEXTO2 = "64748B"
        BLANCO = "FFFFFF"
        BORDE = "CBD5E1"

        thin = Side(style="thin", color=AZUL_OSC)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        thin_gray = Side(style="thin", color=BORDE)
        border_grid = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

        fill_azul = PatternFill("solid", fgColor=AZUL)
        fill_azul_osc = PatternFill("solid", fgColor=AZUL_OSC)
        fill_bg = PatternFill("solid", fgColor=BG)
        fill_rubro = PatternFill("solid", fgColor="E3EAF2")
        fill_white = PatternFill("solid", fgColor=BLANCO)

        FN = "Calibri"
        font_title = Font(name=FN, size=16, bold=True, color=AZUL)
        font_entity = Font(name=FN, size=11, bold=True, color=AZUL_OSC)
        font_sub = Font(name=FN, size=9, color=TEXTO2)
        font_periodo = Font(name=FN, size=10, bold=True, color=AZUL)
        font_num_doc = Font(name=FN, size=10, bold=True, color=AZUL)
        font_label = Font(name=FN, size=8, bold=True, color=AZUL_OSC)
        font_val = Font(name=FN, size=10)
        font_th = Font(name=FN, size=8, bold=True, color=BLANCO)
        font_td = Font(name=FN, size=9)
        font_rubro = Font(name=FN, size=9, bold=True, color=AZUL_OSC)
        font_subtotal = Font(name=FN, size=9, bold=True, color=AZUL_OSC)
        font_total = Font(name=FN, size=10, bold=True, color=BLANCO)
        font_firma_nom = Font(name=FN, size=10, bold=True)
        font_firma_cargo = Font(name=FN, size=8, bold=True, color=AZUL_OSC)
        font_firma_reg = Font(name=FN, size=8, color=TEXTO2)

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        align_right_num = Alignment(horizontal="right", vertical="center")
        align_center_nw = Alignment(horizontal="center", vertical="center")
        align_left_nw = Alignment(horizontal="left", vertical="center")

        wb = Workbook()
        ws = wb.active
        ws.title = "Manifiesto"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 1.0
        ws.page_margins.right = 1.0
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = False
        ws.sheet_view.view = "pageBreakPreview"
        ws.sheet_view.showGridLines = False
        ws.print_title_rows = "17:17"

        # Columnas ampliadas para ocupar todo el ancho de la hoja A4
        # (1" margen izq/der => ancho util ~9.7in) sin verse comprimidas
        col_widths = {"A": 4, "B": 9, "C": 8, "D": 7, "E": 6,
                      "F": 22, "G": 34, "H": 5, "I": 7, "J": 11, "K": 14}
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w

        # ── ENCABEZADO (fiel al print.css) ──
        logo_file = os.path.join(app.root_path, "static", hdr["logo_path"]) if hdr["logo_path"] else None
        nc_m = 11  # A-K
        info_pairs = [
            [("Proyecto", hdr["proyecto"])],
            [("Codigo Unico", hdr["cui"]), ("Correlativo Meta", hdr["meta"])],
            [("Fuente Financ.", hdr["fuente"]), ("Modalidad", hdr["modalidad"])],
            [("Ubicacion", hdr["ubicacion"]), ("Residente de Obra", hdr["residente"])],
            [("Supervisor de Obra", hdr["supervisor"]), ("Admin. de Obra", hdr["asistente"])],
        ]
        font_num_doc = Font(name=FN, size=10, bold=True, color=AZUL)
        tabla_start = _build_xl_header(
            ws, nc_m,
            "Manifiesto de Gasto",
            hdr["modalidad"],
            {"entidad": hdr["entidad"], "unidad": hdr["unidad"],
             "periodo": hdr["periodo"]},
            logo_file, info_pairs, [],
            extra_right={
                1: (f"N° {hdr['n_manifiesto']}", font_num_doc),
                2: ("PERÍODO:", Font(name=FN, size=9, bold=True, color=AZUL)),
                3: (hdr["periodo"], font_periodo),
            },
            header_heights={'project': 40, 'pair': 15})

        # ── Filas 1-12 = encabezado fijo (se repite en cada página) ──
        ws.row_dimensions[5].height = 0
        ws.row_dimensions[5].hidden = True
        ws.row_dimensions[11].height = 7
        ws.print_title_rows = "1:12"

        # ── TABLA FINANCIERA ──

        headers = ["N°", "Fecha", "SIAF", "Tipo Doc", "N°",
                    "Nombre / Proveedor", "Detalle del Gasto", "Und.",
                    "Cant.", "P. Unitario", "Total"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=tabla_start, column=j, value=h)
            c.font = font_th
            c.fill = fill_azul
            c.alignment = align_center
            c.border = border_all
        ws.row_dimensions[tabla_start].height = 22

        r = tabla_start + 1
        subtotal_rows = []
        for s in secciones:
            # Fila de rubro
            ws.merge_cells(f"A{r}:K{r}")
            c = ws.cell(row=r, column=1, value=s["label"])
            c.font = font_rubro
            c.fill = fill_rubro
            c.alignment = align_left
            c.border = border_all
            ws.row_dimensions[r].height = 20
            r += 1

            sec_data_start = r
            for f in s["filas"]:
                prov_num = f["prov_num"] if f["prov_first"] else ""
                fecha_str = f["fecha"].strftime("%d/%m/%Y") if f["prov_first"] and f["fecha"] else ""
                siaf = f["siaf"] if f["prov_first"] else ""
                tipo_doc = f["tipo_doc"] if f["prov_first"] else ""
                num_doc = f["num_doc"] if f["prov_first"] else ""
                proveedor = f["proveedor"] if f["prov_first"] else ""

                vals = [prov_num, fecha_str, siaf, tipo_doc, num_doc,
                        proveedor, f["detalle"], f["und"],
                        f["cantidad"], f["pu"], f"=ROUND(I{r}*J{r},2)"]
                aligns = [align_center_nw, align_left_nw, align_left, align_center_nw,
                          align_center_nw, align_left, align_left, align_center_nw,
                          align_right_num, align_right_num, align_right_num]

                for j, (v, al) in enumerate(zip(vals, aligns), 1):
                    c = ws.cell(row=r, column=j, value=v)
                    c.font = font_td
                    c.alignment = al
                    # Solo aplicar borde si la celda tiene contenido
                    if v is not None and v != "":
                        c.border = border_grid
                    if j in (9, 10, 11) and v:
                        c.number_format = '#,##0.00'
                # Height: líneas según ancho de col (G=34, F=22) × 14pt
                v_texto = str(f["detalle"] or "")
                v_prov = str(proveedor or "")
                # ~1.1 chars por unidad de ancho de columna
                h_texto = math.ceil(len(v_texto) / 37) * 14 if v_texto else 0
                h_prov = math.ceil(len(v_prov) / 24) * 14 if v_prov else 0
                ws.row_dimensions[r].height = max(h_texto, h_prov, 18)
                r += 1

            # Subtotal
            ws.merge_cells(f"A{r}:J{r}")
            c_lbl = ws.cell(row=r, column=1,
                            value=f"SUBTOTAL {s['componente'].upper()} S/.")
            c_lbl.font = font_subtotal
            c_lbl.fill = fill_bg
            c_lbl.alignment = Alignment(horizontal="right", vertical="center")
            c_lbl.border = border_all
            for col in range(2, 11):
                ws.cell(row=r, column=col).fill = fill_bg
                ws.cell(row=r, column=col).border = border_all
            c_val = ws.cell(row=r, column=11,
                            value=f"=SUM(K{sec_data_start}:K{r-1})")
            c_val.font = font_subtotal
            c_val.fill = fill_bg
            c_val.alignment = align_right_num
            c_val.border = border_all
            c_val.number_format = '#,##0.00'
            subtotal_rows.append(r)
            ws.row_dimensions[r].height = 20
            r += 1

        if not secciones:
            ws.merge_cells(f"A{r}:K{r}")
            c = ws.cell(row=r, column=1,
                        value="SIN GASTOS DEVENGADOS PARA EL PERÍODO")
            c.font = font_rubro
            c.fill = fill_rubro
            c.alignment = align_center
            c.border = border_all
            r += 1

        # Total general
        ws.merge_cells(f"A{r}:J{r}")
        c_lbl = ws.cell(row=r, column=1, value="TOTAL GENERAL S/.")
        c_lbl.font = font_total
        c_lbl.fill = fill_azul_osc
        c_lbl.alignment = Alignment(horizontal="right", vertical="center")
        c_lbl.border = border_all
        for col in range(2, 11):
            ws.cell(row=r, column=col).fill = fill_azul_osc
            ws.cell(row=r, column=col).border = border_all
        # Suma todas las filas de subtotales
        grand_val = (f"=SUM(" + ",".join(f"K{sr}" for sr in subtotal_rows) + ")"
                     ) if subtotal_rows else total_gral
        c_val = ws.cell(row=r, column=11, value=grand_val)
        c_val.font = font_total
        c_val.fill = fill_azul_osc
        c_val.alignment = align_right_num
        c_val.border = border_all
        c_val.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 24
        r += 1

        # ── FIRMAS ──
        firmas_row = r + 3
        # Espacio antes de las firmas (p.ej. fila 40 → 45pt)
        ws.row_dimensions[firmas_row - 3].height = 45
        firmas = [
            (hdr["residente"], "Residente de Obra",
             f"CIP: {hdr['cip_residente']}" if hdr["cip_residente"] else ""),
            (hdr["supervisor"], "Supervisor de Obra",
             f"CIP: {hdr['cip_supervisor']}" if hdr["cip_supervisor"] else ""),
            (p.administrador_obra or p.asistente or "",
             "V°B° Administración / Entidad", ""),
        ]
        for idx, (nombre, cargo, reg) in enumerate(firmas):
            col = 1 + idx * 4
            ws.merge_cells(start_row=firmas_row, start_column=col,
                           end_row=firmas_row, end_column=col + 2)

            # Línea de firma (borde superior para simular la línea)
            c_line = ws.cell(row=firmas_row, column=col)
            c_line.border = Border(
                top=Side(style="thin", color=AZUL),
                bottom=Side(style=None)
            )
            # Apply border to all cells in merged range
            for merge_col in range(col, col + 3):
                ws.cell(row=firmas_row, column=merge_col).border = Border(
                    top=Side(style="thin", color=AZUL)
                )

            # Línea de firma compacta (8pt)
            ws.row_dimensions[firmas_row].height = 8

            # Nombre
            r_nom = firmas_row + 1
            ws.merge_cells(start_row=r_nom, start_column=col,
                           end_row=r_nom, end_column=col + 2)
            c_nom = ws.cell(row=r_nom, column=col, value=nombre)
            c_nom.font = font_firma_nom
            c_nom.alignment = Alignment(horizontal="center", vertical="bottom",
                                        wrap_text=True)
            ws.row_dimensions[r_nom].height = 20

            # Cargo
            r_cargo = r_nom + 1
            ws.merge_cells(start_row=r_cargo, start_column=col,
                           end_row=r_cargo, end_column=col + 2)
            c_cargo = ws.cell(row=r_cargo, column=col, value=cargo)
            c_cargo.font = font_firma_cargo
            c_cargo.alignment = Alignment(horizontal="center", vertical="top")
            ws.row_dimensions[r_cargo].height = 16

            # Registro
            if reg:
                r_reg = r_cargo + 1
                ws.merge_cells(start_row=r_reg, start_column=col,
                               end_row=r_reg, end_column=col + 2)
                c_reg = ws.cell(row=r_reg, column=col, value=reg)
                c_reg.font = font_firma_reg
                c_reg.alignment = Alignment(horizontal="center", vertical="top")
                ws.row_dimensions[r_reg].height = 14

        # ── Ajustes finales ──
        ws.print_area = f"A1:K{firmas_row + 4}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"Manifiesto_Gasto_{hdr['periodo'].replace(' ', '_')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ------------------------- CONFIGURACION --------------------------
    @app.route("/configuracion", methods=["GET", "POST"])
    def configuracion():
        p = get_proyecto()
        if request.method == "POST":
            num = int(request.form.get("num_anios_anteriores", 3) or 3)
            num = max(0, min(3, num))
            p.num_anios_anteriores = num
            p.incluir_anios_anteriores = num > 0
            # Guardar metas por año
            for yr, attr in [("2023", "meta_ejec2023"), ("2024", "meta_ejec2024"),
                              ("2025", "meta_ejec2025")]:
                try:
                    setattr(p, attr, float(request.form.get(f"meta_{yr}", 0) or 0))
                except ValueError:
                    pass
            db.session.commit()
            # Columnas de ejecución: mapeo inverso (ejec2025 = anio-1, ejec2024 = anio-2, etc.)
            anio = p.anio
            year_cols = {}
            for i in range(1, 4):
                yr = anio - i
                col = f"ejec{yr}"
                year_cols[col] = i  # 1=ejec2025, 2=ejec2024, 3=ejec2023
            incluir_cols = {f"ejec{anio - i}" for i in range(1, num + 1)}
            for key, value in request.form.items():
                if key.startswith("pim_") or key.startswith("et_") or \
                   (key.startswith("ejec20") and key.split("_")[0] in incluir_cols):
                    try:
                        pref, cid = key.split("_", 1)
                        cfg = db.session.get(Presupuesto, int(cid))
                        if cfg:
                            setattr(cfg, "pim2026" if pref == "pim" else pref,
                                    float(value or 0))
                    except (ValueError, TypeError):
                        pass
            db.session.commit()
            from collections import defaultdict
            et_comp = defaultdict(float)
            for cfg in Presupuesto.query.filter(
                    Presupuesto.componente.in_(COMPONENTES_FE06)).all():
                et_comp[cfg.componente] = round(et_comp[cfg.componente] + (cfg.et or 0), 2)
            p.costo_directo = round(et_comp.get("Costo Directo", 0), 2)
            p.gastos_generales = round(et_comp.get("Gastos Generales", 0), 2)
            p.gastos_supervision = round(et_comp.get("Gastos de Supervisión", 0), 2)
            p.elaboracion_expediente = round(et_comp.get("Elaboración de Expediente Técnico", 0), 2)
            p.liquidacion_obra = round(et_comp.get("Liquidación de Obra", 0), 2)
            p.presupuesto_total = round(sum(et_comp.values()), 2)
            db.session.commit()
            flash("Configuración presupuestal guardada.", "success")
            return redirect(url_for("configuracion"))
        orden_detalle = {"PERSONAL": 0, "BIENES": 1, "SERVICIOS": 2,
                         "ELABORACION DE EXPEDIENTE TECNICO": 3,
                         "COSTO DE LIQUIDACION": 4}
        comp_orden = {c: i for i, c in enumerate(COMPONENTES_FE06)}
        # Construir set de detalles válidos: base + extras actuales del proyecto
        base_dets = {det for _, det in PRESUPUESTO_DETALLE}
        raw_ex = getattr(p, "clasificadores_extra", "") or ""
        extra_dets = set()
        try:
            for ex in (json.loads(raw_ex) if raw_ex else []):
                n = (ex.get("nombre") or "").strip().upper()
                if n:
                    extra_dets.add(n)
        except (json.JSONDecodeError, TypeError):
            pass
        valid_dets = base_dets | extra_dets
        # Eliminar filas huérfanas de Presupuesto que ya no corresponden a ningún detalle
        for cfg in Presupuesto.query.all():
            if cfg.detalle not in valid_dets:
                db.session.delete(cfg)
        db.session.flush()
        configs = sorted(Presupuesto.query.all(),
                         key=lambda c: (comp_orden.get(c.componente, 99),
                                        orden_detalle.get(c.detalle, 99)))
        return render_template("configuracion.html", p=p, configs=configs)

    # ------------------------- RESPALDO DE DATOS -----------------------
    @app.route("/respaldo", methods=["GET", "POST"])
    def respaldo():
        actor = usuario_actual()
        es_super = bool(actor and actor.rol == ROL_SUPER)
        p = None if es_super else get_proyecto()
        if request.method == "POST":
            accion = request.form.get("accion")
            if accion == "crear":
                try:
                    dest = crear_respaldo()
                    flash(f"Respaldo creado correctamente: {os.path.basename(dest)}",
                          "success")
                except Exception as e:
                    flash(f"No se pudo crear el respaldo: {e}", "danger")
                return redirect(url_for("respaldo"))
            if accion == "restaurar":
                nombre = request.form.get("archivo", "")
                if nombre not in [r["nombre"] for r in listar_respaldos()]:
                    flash("Archivo de respaldo no válido.", "danger")
                    return redirect(url_for("respaldo"))
                try:
                    pre = restaurar_respaldo(nombre)
                    flash(f"Base de datos restaurada desde {nombre}. Se guardó una copia "
                          f"del estado anterior: {os.path.basename(pre)}", "success")
                except Exception as e:
                    flash(f"No se pudo restaurar el respaldo: {e}", "danger")
                return redirect(url_for("respaldo"))
            if accion == "cargar":
                archivo = request.files.get("archivo_db")
                if not archivo or not archivo.filename:
                    flash("Debe seleccionar un archivo de base de datos.", "danger")
                    return redirect(url_for("respaldo"))
                fname = archivo.filename.strip()
                if not fname.lower().endswith(".db"):
                    flash("El archivo debe ser una base de datos SQLite (.db).", "danger")
                    return redirect(url_for("respaldo"))
                # Nombre seguro: evitar路径 traversal
                safe_name = os.path.basename(fname)
                os.makedirs(RESPALDO_DIR, exist_ok=True)
                dest = os.path.join(RESPALDO_DIR, safe_name)
                archivo.save(dest)
                # Validar que sea SQLite
                try:
                    con = sqlite3.connect(dest)
                    con.execute("SELECT COUNT(*) FROM sqlite_master").fetchone()
                    con.close()
                except Exception:
                    try:
                        os.remove(dest)
                    except Exception:
                        pass
                    flash("El archivo no es una base de datos SQLite válida.", "danger")
                    return redirect(url_for("respaldo"))
                # Cifrar el archivo cargado
                try:
                    with open(dest, "rb") as f:
                        raw = f.read()
                    with open(dest, "wb") as f:
                        f.write(_cifrar_datos(raw))
                except Exception:
                    pass
                # Restaurar desde el archivo cargado
                try:
                    pre = restaurar_respaldo(safe_name)
                    flash(f"Base de datos restaurada desde {safe_name}. Se guardó una copia "
                          f"del estado anterior: {os.path.basename(pre)}", "success")
                except Exception as e:
                    flash(f"No se pudo restaurar el respaldo cargado: {e}", "danger")
                return redirect(url_for("respaldo"))
        db_path = ruta_db()
        stats = None
        if os.path.exists(db_path) and not es_super:
            stats = {"ruta": db_path, "tamano": os.path.getsize(db_path),
                     "mod": datetime.fromtimestamp(os.path.getmtime(db_path)),
                     "integridad": verificar_integridad(db_path),
                     "proyecto": Proyecto.query.count(),
                     "presupuesto": Presupuesto.query.count(),
                     "gastos": Gasto.query.count(),
                     "detalles": GastoDetalle.query.count(),
                     "almacen": AlmacenMovimiento.query.count()}
        return render_template("respaldo.html", p=p, stats=stats,
                               respaldos=listar_respaldos(),
                               es_super=es_super,
                               carpeta=RESPALDO_DIR)

    # ------------------------- FORMATOS: IMPRESION --------------------------
    @app.route("/formatos/fe05/imprimir")
    def imprimir_fe05():
        """Vista imprimible del FE-05 Ejecución Presupuestal Mensual."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        secciones, total, clasif_cols = f05_datos(mes, anio)
        return render_template("fe05_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, secciones=secciones, total=total,
                               clasif_cols=clasif_cols)

    @app.route("/formatos/fe05/excel")
    def fe05_excel():
        """Genera archivo .xlsx del FE-05 Ejecucion Presupuestal Mensual fiel
        al diseño de la hoja de impresión."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        import math

        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        secciones, total, clasif_cols = f05_datos(mes, anio)

        # ── Colores fieles al print.css ──
        AZUL = "1E3A5F"
        AZUL_OSC = "162D45"
        AZUL_CLARO = "2C527A"
        BG = "F8FAFC"
        TEXTO2 = "64748B"
        BLANCO = "FFFFFF"
        BORDE = "CBD5E1"

        thin = Side(style="thin", color=AZUL_OSC)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        thin_gray = Side(style="thin", color=BORDE)
        border_grid = Border(left=thin_gray, right=thin_gray, top=thin_gray,
                             bottom=thin_gray)

        fill_azul = PatternFill("solid", fgColor=AZUL)
        fill_azul_osc = PatternFill("solid", fgColor=AZUL_OSC)
        fill_bg = PatternFill("solid", fgColor=BG)
        fill_rubro = PatternFill("solid", fgColor="E3EAF2")

        FN = "Calibri"
        font_entity = Font(name=FN, size=11, bold=True, color=AZUL_OSC)
        font_sub = Font(name=FN, size=9, color=TEXTO2)
        font_title = Font(name=FN, size=16, bold=True, color=AZUL)
        font_periodo = Font(name=FN, size=10, bold=True, color=AZUL)
        font_label = Font(name=FN, size=8, bold=True, color=AZUL_OSC)
        font_val = Font(name=FN, size=10)
        font_th = Font(name=FN, size=8, bold=True, color=BLANCO)
        font_td = Font(name=FN, size=9)
        font_td_b = Font(name=FN, size=9, bold=True)
        font_rubro = Font(name=FN, size=9, bold=True, color=AZUL_OSC)
        font_subtotal = Font(name=FN, size=9, bold=True, color=AZUL_OSC)
        font_total = Font(name=FN, size=10, bold=True, color=BLANCO)

        align_center = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
        align_nw = Alignment(horizontal="left", vertical="center",
                             wrap_text=False)
        align_center_nw = Alignment(horizontal="center", vertical="center",
                                    wrap_text=False)
        align_right_num = Alignment(horizontal="right", vertical="center",
                                    wrap_text=False)

        wb = Workbook()
        ws = wb.active
        ws.title = "FE-05"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = True
        ws.sheet_view.view = "pageBreakPreview"
        ws.sheet_view.showGridLines = False

        ncols = 15 + len(clasif_cols)

        # Anchos en px (total = 1123px para A4 horizontal)
        # Reducir texto para ampliar clasificadores y numeric cols sin ###
        px = [22, 38, 68, 30, 30, 120, 165, 28, 32, 56, 66] \
            + [70] * len(clasif_cols) + [66, 62, 60, 72]
        for j, w in enumerate(px, 1):
            ws.column_dimensions[get_column_letter(j)].width = round(w / 7, 1)

        # Anchos reales de columna (después de asignar widths)
        col_chars = {j: max(3, ws.column_dimensions[get_column_letter(j)].width or 5)
                     for j in range(1, ncols + 1)}

        # ── ENCABEZADO (fiel al print.css) ──
        logo_file = os.path.join(app.root_path, "static", p.logo_path) if p.logo_path else None
        info_pairs = [
            [("Proyecto", p.nombre or "")],
            [("Codigo Unico", p.cui or ""), ("Correlativo Meta", p.meta or "")],
            [("Fuente Financ.", p.fuente or ""), ("Modalidad", "ADMINISTRACION DIRECTA")],
        ]
        responsables = [
            ("Supervisor de Obra", p.supervisor or ""),
            ("Residente de Obra", p.residente or ""),
            ("Admin. de Obra", p.asistente or ""),
        ]
        tabla_start = _build_xl_header(
            ws, ncols,
            "Ejecucion Presupuestal Mensual",
            "Total Costo de Construccion por Adm. Directa",
            {"entidad": p.entidad or "", "unidad": p.unidad_ejecutora or "",
             "periodo": f"PERIODO: {MESES[mes-1]} - {anio}"},
            logo_file, info_pairs, responsables,
            header_heights={'resp': 22}, info_cols="left")

        # ── Ocultar fila 5 spacer + reducir fila 12 spacer ──
        ws.row_dimensions[5].hidden = True
        ws.row_dimensions[5].height = 0
        ws.row_dimensions[12].height = 7

        # ── Auto-ajustar alturas filas 1-11 (encabezado) ──
        def _xl_row_height(texto, col_width_chars, font_size, wrap=True):
            if not wrap or not texto:
                return font_size * 1.6
            chars_por_linea = max(1, int(col_width_chars * 0.85))
            nlineas = max(1, math.ceil(len(str(texto)) / chars_por_linea))
            return max(font_size * 1.6, nlineas * (font_size * 1.6))

        # Fila 1: entidad
        w1 = sum(col_chars.get(j, 9) for j in range(2, ncols))
        ws.row_dimensions[1].height = max(22, _xl_row_height(ws.cell(1, 2).value, w1, 11))
        # Fila 2: unidad
        ws.row_dimensions[2].height = max(16, _xl_row_height(ws.cell(2, 2).value, w1, 9))
        # Fila 3: titulo
        ws.row_dimensions[3].height = max(24, _xl_row_height(ws.cell(3, 2).value, w1, 13))
        # Fila 4: subtitulo
        ws.row_dimensions[4].height = max(16, _xl_row_height(ws.cell(4, 2).value, w1, 9))
        # Fila 6: proyecto
        # Fila 6: proyecto (value in D+)
        w6 = sum(col_chars.get(j, 9) for j in range(4, ncols + 1))
        ws.row_dimensions[6].height = max(17, _xl_row_height(p.nombre, w6, 10))
        # Filas 7-8: info pairs (each value in D:half)
        half = ncols // 2
        w_half = sum(col_chars.get(j, 9) for j in range(4, half + 1))
        for info_idx, row_items in enumerate(info_pairs[1:], start=7):
            txt = max((item[1] for item in row_items),
                      key=lambda x: len(str(x or "")), default="")
            ws.row_dimensions[info_idx].height = max(15, _xl_row_height(txt, w_half, 10))
        # Filas 9-11: responsables (value in D:lastcol)
        resp_w = sum(col_chars.get(j, 9) for j in range(4, ncols + 1))
        for i, (lab, val) in enumerate(responsables):
            ws.row_dimensions[9 + i].height = max(15, _xl_row_height(val, resp_w, 9))

        ws.print_title_rows = f"{tabla_start}:{tabla_start}"
        ws.freeze_panes = f"A{tabla_start + 1}"
        headers = ["N°", "SIAF", "Fecha", "Tipo", "N°", "Proveedor", "Detalle",
                   "Und", "Cant", "P.U", "Importe"] + list(clasif_cols) \
                 + ["Directos", "Grales", "Superv.", "Total"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=tabla_start, column=j, value=h)
            c.font = font_th
            c.fill = fill_azul
            c.alignment = align_center
            c.border = border_all
        ws.row_dimensions[tabla_start].height = 30

        r = tabla_start + 1
        subtotal_rows = []
        for s in secciones:
            # Fila de rubro (.fila-rubro)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            c = ws.cell(row=r, column=1, value=s["label"])
            c.font = font_rubro
            c.alignment = align_left
            for col in range(1, ncols + 1):
                cc = ws.cell(row=r, column=col)
                cc.fill = fill_rubro
                cc.border = border_all
            ws.row_dimensions[r].height = 18
            r += 1
            sec_start = r

            for f in s["filas"]:
                pf = f["prov_first"]
                g, d = f["g"], f["d"]
                vals = [
                    f["prov_num"] if pf else "",
                    g.siaf or "" if pf else "",
                    g.fecha.strftime("%d/%m/%Y") if pf and g.fecha else "",
                    g.tipo_doc or "" if pf else "",
                    g.num_doc or "" if pf else "",
                    g.proveedor or "" if pf else "",
                    d.detalle or "",
                    d.und or "",
                    d.cantidad if d.cantidad else None,
                    d.precio_unitario,
                    d.importe,
                ] + [f["clasif"][cl] if f["clasif"][cl] else "-"
                     for cl in clasif_cols] + [
                    f["directos"] if f["directos"] else "-",
                    f["generales"] if f["generales"] else "-",
                    f["supervision"] if f["supervision"] else "-",
                    d.importe,
                ]
                aligns = [align_center_nw, align_nw, align_center_nw,
                          align_center_nw, align_center_nw, align_left,
                          align_left, align_center_nw]
                for j, v in enumerate(vals, 1):
                    c = ws.cell(row=r, column=j, value=v)
                    if v is not None and v != "":
                        c.border = border_grid
                    if j <= 8:
                        c.font = font_td
                        c.alignment = aligns[j - 1]
                    elif j == ncols:
                        c.font = font_td_b
                        c.alignment = align_right_num
                    else:
                        c.font = font_td
                        c.alignment = align_right_num
                    if isinstance(v, (int, float)):
                        # Cols N°(1), SIAF(2), N°doc(5), CANT(8) = enteros
                        if j in (1, 2, 5, 8):
                            c.number_format = '0'
                        else:
                            c.number_format = '#,##0.00'
                tc = ws.cell(row=r, column=ncols, value=f"=K{r}")
                tc.number_format = '#,##0.00'
                # Altura de fila: ajustar al contenido más alto
                h = 18
                for j, v in enumerate(vals, 1):
                    if v is None or v == "":
                        continue
                    cw = col_chars.get(j, 9)
                    txt = str(v)
                    # Proveedor (6) y Detalle (7) permiten wrap; resto no
                    if j in (6, 7):
                        chars_linea = max(1, int(cw * 0.85))
                        nlineas = max(1, math.ceil(len(txt) / chars_linea))
                        h = max(h, nlineas * 13)
                    else:
                        h = max(h, 18)
                ws.row_dimensions[r].height = h
                r += 1

            # Subtotal de sección (.fila-sub)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            c_lbl = ws.cell(row=r, column=1, value=f"SUBTOTAL {s['label']}")
            c_lbl.font = font_subtotal
            c_lbl.alignment = Alignment(horizontal="right", vertical="center")
            nums = [s["totales"]["total"]] \
                + [s["totales"]["clasif"][cl] for cl in clasif_cols] \
                + [s["totales"]["directos"], s["totales"]["generales"],
                   s["totales"]["supervision"], s["totales"]["total"]]
            for j, v in enumerate(nums):
                col = 11 + j
                if r > sec_start:
                    letra = get_column_letter(col)
                    cc = ws.cell(row=r, column=col,
                                 value=f"=SUM({letra}{sec_start}:{letra}{r - 1})")
                else:
                    cc = ws.cell(row=r, column=col, value=v)
                cc.font = font_subtotal
                cc.alignment = align_right_num
                cc.number_format = '#,##0.00'
            for col in range(1, ncols + 1):
                cc = ws.cell(row=r, column=col)
                cc.fill = fill_bg
                cc.border = border_all
            ws.row_dimensions[r].height = 18
            subtotal_rows.append(r)
            r += 1

        if not secciones:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            c = ws.cell(row=r, column=1,
                        value="SIN GASTOS DEVENGADOS PARA EL PERÍODO")
            c.font = font_rubro
            c.alignment = align_left
            for col in range(1, ncols + 1):
                cc = ws.cell(row=r, column=col)
                cc.fill = fill_rubro
                cc.border = border_all
            ws.row_dimensions[r].height = 18
            r += 1

        # Total general (.fila-total)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c_lbl = ws.cell(row=r, column=1, value="TOTAL GASTO EJECUTADO")
        c_lbl.font = font_total
        c_lbl.alignment = Alignment(horizontal="right", vertical="center")
        tnums = [total["total"]] + [total["clasif"][cl] for cl in clasif_cols] \
            + [total["directos"], total["generales"], total["supervision"],
               total["total"]]
        for j, v in enumerate(tnums):
            col = 11 + j
            if subtotal_rows:
                cc = ws.cell(row=r, column=col, value="=SUM("
                             + ",".join(f"{get_column_letter(col)}{sr}"
                                        for sr in subtotal_rows) + ")")
            else:
                cc = ws.cell(row=r, column=col, value=v)
            cc.font = font_total
            cc.alignment = align_right_num
            cc.number_format = '#,##0.00'
        for col in range(1, ncols + 1):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill_azul_osc
            cc.border = border_all
        ws.row_dimensions[r].height = 22
        r += 1

        # ── Ajustes finales ──
        ws.print_area = f"A1:{get_column_letter(ncols)}{r}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"FE05_Ejecucion_{MESES[mes-1]}_{anio}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/formatos/fe06/imprimir")
    def imprimir_fe06():
        """Vista imprimible del FE-06 Presupuesto vs Ejecutado."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        rows = fe06_rows()
        resumen = fe06_resumen(rows)
        fe06_totales_mensual = [round(sum(r["mensual"][i] for r in resumen.values()), 2)
                                for i in range(12)]
        sintesis = fe06_sintesis(mes, anio)
        return render_template("fe06_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, fe06_rows=rows, fe06_resumen=resumen,
                               fe06_totales_mensual=fe06_totales_mensual,
                                fe06_sintesis=sintesis,
                                meses_activos=meses_con_ejecucion(anio),
                                meses_vis=meses_visibles(anio, mes))

    @app.route("/formatos/fe06/excel")
    def fe06_excel():
        """Genera archivo .xlsx del FE-06 Presupuesto vs Ejecutado fiel al
        diseño de la hoja de impresión."""
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                     numbers)
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
        import math

        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        rows = fe06_rows()
        resumen = fe06_resumen(rows)
        fe06_totales_mensual = [round(sum(r["mensual"][i] for r in resumen.values()), 2)
                                for i in range(12)]
        sintesis = fe06_sintesis(mes, anio)
        meses_vis = meses_visibles(anio, mes)
        incluir_anios = p.incluir_anios_anteriores

        # ── Colores fieles al print.css ──
        AZUL = "1E3A5F"
        AZUL_OSC = "162D45"
        AZUL_CLARO = "2C527A"
        BG = "F8FAFC"
        TEXTO2 = "64748B"
        BLANCO = "FFFFFF"
        BORDE = "CBD5E1"

        thin = Side(style="thin", color=AZUL_OSC)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        thin_gray = Side(style="thin", color=BORDE)
        border_grid = Border(left=thin_gray, right=thin_gray, top=thin_gray,
                             bottom=thin_gray)

        fill_azul = PatternFill("solid", fgColor=AZUL)
        fill_azul_osc = PatternFill("solid", fgColor=AZUL_OSC)
        fill_bg = PatternFill("solid", fgColor=BG)
        fill_rubro = PatternFill("solid", fgColor="E3EAF2")
        fill_resp = PatternFill("solid", fgColor="E8EDF3")

        FN = "Calibri"
        font_entity = Font(name=FN, size=11, bold=True, color=AZUL_OSC)
        font_sub = Font(name=FN, size=9, color=TEXTO2)
        font_title = Font(name=FN, size=16, bold=True, color=AZUL)
        font_periodo = Font(name=FN, size=10, bold=True, color=AZUL)
        font_label = Font(name=FN, size=8, bold=True, color=AZUL_OSC)
        font_val = Font(name=FN, size=10)
        font_th = Font(name=FN, size=8, bold=True, color=BLANCO)
        font_td = Font(name=FN, size=9)
        font_td_b = Font(name=FN, size=9, bold=True)
        font_rubro = Font(name=FN, size=9, bold=True, color=AZUL_OSC)
        font_total = Font(name=FN, size=10, bold=True, color=BLANCO)

        align_center = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center",
                               wrap_text=True)
        align_right_num = Alignment(horizontal="right", vertical="center")

        wb = Workbook()
        ws = wb.active
        ws.title = "FE-06"
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left = 1.0
        ws.page_margins.right = 1.0
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = False
        ws.sheet_view.view = "pageBreakPreview"
        ws.sheet_view.showGridLines = False

        # ── Columnas de la tabla (segun configuracion del proyecto) ──
        col_pim = 4 + (3 if incluir_anios else 0)
        col_mes_ini = col_pim + 1
        col_ejec = col_mes_ini + len(meses_vis)
        col_pct = col_ejec + 1
        col_saldo = col_ejec + 2
        ncols = max(11, col_saldo + (2 if incluir_anios else 0))

        anchos = {1: 11, 2: 26, 3: 15}
        if incluir_anios:
            anchos.update({4: 9, 5: 9, 6: 9, col_saldo + 1: 11, col_saldo + 2: 12})
        anchos[col_pim] = 11
        for k, _m in enumerate(meses_vis):
            anchos[col_mes_ini + k] = 9
        anchos[col_ejec] = 11
        anchos[col_pct] = 9
        anchos[col_saldo] = 11
        for j in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = anchos.get(j, 9)
        last_col = get_column_letter(ncols)

        # ── ENCABEZADO (fiel al print.css) ──
        logo_file = os.path.join(app.root_path, "static", p.logo_path) if p.logo_path else None
        info_pairs = [
            [("Proyecto", p.nombre or "")],
            [("Codigo Unico", p.cui or ""), ("Correlativo Meta", p.meta or "")],
            [("Fuente Financ.", p.fuente or ""), ("Modalidad", "ADMINISTRACION DIRECTA")],
        ]
        responsables = [
            ("Supervisor de Obra", p.supervisor or ""),
            ("Residente de Obra", p.residente or ""),
            ("Admin. de Obra", p.asistente or ""),
        ]
        tabla_start = _build_xl_header(
            ws, ncols,
            "Cuadro Comparativo Presupuesto Analitico Aprobado y Ejecutado",
            f"Periodo: {MESES[mes-1]} - {anio}",
            {"entidad": p.entidad or "", "unidad": p.unidad_ejecutora or "",
             "periodo": f"PERIODO: {MESES[mes-1]} - {anio}"},
            logo_file, info_pairs, responsables,
            title_font_size=13,
            header_heights={'entity': 22, 'sub': 16, 'title': 24, 'subtitle': 16,
                            'project': 36, 'pair': 15, 'resp': 16})

        # Eliminar fila 5 (spacer vacio entre encabezado e info-grid)
        # No usar delete_rows porque desajusta los merges de _build_xl_header
        ws.row_dimensions[5].hidden = True
        ws.row_dimensions[5].height = 0

        # Reducir filas spacer (12=before tabla, 31=before PIM, 38=before EXP, 43=trailing)
        for spacer_row in [12, 31, 38, 43]:
            ws.row_dimensions[spacer_row].height = 7

        # ── Ajustar fila Proyecto (row 6 after header): fusionar B6:L6 ──
        # After hidden row 5, row 6 has Proyecto data from _build_xl_header
        no_border = Border()
        # Unmerge existing merges on row 6
        for m in list(ws.merged_cells.ranges):
            if m.min_row == 6:
                ws.unmerge_cells(str(m))
        # Label en A6
        c5a = ws.cell(row=6, column=1)
        c5a.value = "Proyecto"
        c5a.font = font_label
        c5a.alignment = align_left
        c5a.fill = fill_bg
        c5a.border = border_grid
        # Valor del proyecto: set value FIRST, then merge
        c5b = ws.cell(row=6, column=2)
        c5b.value = p.nombre or ""
        c5b.font = font_val
        c5b.alignment = align_left
        c5b.border = border_grid
        ws.merge_cells("B6:L6")
        for cc in range(3, ncols + 1):
            ws.cell(row=6, column=cc).border = border_grid

        # ── Limpiar bordes y merges en responsables cols F-L (6-12) ──
        # _build_xl_header crea merges que se extienden hasta ncols,
        # pero responsables solo deben usar A-E. Unmerge + recrear.
        resp_row_start = 6 + len(info_pairs)  # row 9, 10, 11
        resp_col_end = min(5, ncols)
        def _xl_row_height(texto, col_width_chars, font_size, wrap=True):
            """Calcula altura de fila necesaria para el texto."""
            if not wrap or not texto:
                return font_size * 1.6
            chars_por_linea = max(1, int(col_width_chars * 0.85))
            nlineas = max(1, math.ceil(len(str(texto)) / chars_por_linea))
            return max(font_size * 1.6, nlineas * (font_size * 1.6))

        # Fila 1-4: header (ya con alturas de _build_xl_header)
        # Fila 6: proyecto (B6:L6)
        w6 = sum(anchos.get(j, 9) for j in range(2, ncols + 1))
        h6 = _xl_row_height(p.nombre, w6, 10)
        ws.row_dimensions[6].height = max(17, h6)
        # Filas 7-8: info pairs (half-width)
        mid_w = sum(anchos.get(j, 9) for j in range(1, (ncols + 1) // 2 + 1))
        for info_row_idx, row_items in enumerate(info_pairs[1:], start=7):
            txt = max((item[1] for item in ws.cell(info_row_idx, 1).value and [] or []),
                      key=lambda x: len(str(x or "")), default="")
            ws.row_dimensions[info_row_idx].height = max(15, _xl_row_height(txt, mid_w, 10))
        # Filas 9-11: responsables (value width = full row)
        resp_w = sum(anchos.get(j, 9) for j in range(1, ncols + 1))
        for i, (lab, val) in enumerate(responsables):
            ws.row_dimensions[resp_row_start + i].height = max(15, _xl_row_height(val, resp_w, 9))

        # ── Limpiar bordes y merges en responsables cols F-L ──
        for resp_row in range(resp_row_start, resp_row_start + len(responsables)):
            # Unmerge all ranges in this row
            for m in list(ws.merged_cells.ranges):
                if m.min_row == resp_row:
                    ws.unmerge_cells(str(m))
            # Recreate merges: A:B label, C:E value
            ws.merge_cells(f"A{resp_row}:B{resp_row}")
            ws.merge_cells(f"C{resp_row}:{get_column_letter(resp_col_end)}{resp_row}")
            # Clear all cells cols 1-ncols and re-apply correct styling
            for cc in range(1, ncols + 1):
                cell = ws.cell(row=resp_row, column=cc)
                cell.border = no_border
                cell.fill = PatternFill()
            # Apply correct borders: A-B label with azul fill, C-E value with resp fill
            for cc in [1, 2]:
                cell = ws.cell(row=resp_row, column=cc)
                cell.border = border_all
                cell.fill = fill_azul
            for cc in range(3, resp_col_end + 1):
                cell = ws.cell(row=resp_row, column=cc)
                cell.border = border_all
                cell.fill = fill_resp

        ws.print_title_rows = f"{tabla_start}:{tabla_start}"
        ws.freeze_panes = f"A{tabla_start + 1}"
        headers = ["Especif. de gasto", "Detalle", "Exp. Tecnico"]
        if incluir_anios:
            headers += ["2023", "2024", "2025"]
        headers += [f"PIM {anio}"]
        headers += [MESES[m - 1][:3] for m in meses_vis]
        headers += [f"Ejec. {anio}", "%", f"Saldo {anio}"]
        if incluir_anios:
            headers += ["Acum. total", "Saldo proyecto"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=tabla_start, column=j, value=h)
            c.font = font_th
            c.fill = fill_azul
            c.alignment = align_center
            c.border = border_all
        ws.row_dimensions[tabla_start].height = 30

        def valores_fila(reg, fraccion):
            """Valores numericos de una fila segun las columnas activas."""
            nums = [reg["et"]]
            if incluir_anios:
                nums += [reg["e2023"], reg["e2024"], reg["e2025"]]
            nums.append(reg["pim"])
            nums += [reg["mensual"][m - 1] for m in meses_vis]
            nums += [reg["total_anio"], fraccion, reg["saldo_pim"]]
            if incluir_anios:
                nums += [reg["acum_total"], reg["saldo_et"]]
            return nums

        def pintar_fila(r, etiqueta, nums, fmt_pct, estilo):
            """Fila con A:B = etiqueta merged + columnas numericas.

            estilo: 'comp' (rubro), 'total' (fila-total); ambos con borde
            azul y relleno propio."""
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            fuente, borde, relleno = ((font_rubro, border_all, fill_rubro)
                                      if estilo == "comp"
                                      else (font_total, border_all,
                                            fill_azul_osc))
            c_lab = ws.cell(row=r, column=1, value=etiqueta)
            c_lab.font = fuente
            c_lab.alignment = align_left
            for j, v in enumerate(nums, 3):
                cc = ws.cell(row=r, column=j, value=v)
                cc.font = fuente
                cc.alignment = align_right_num
                cc.number_format = fmt_pct if j == col_pct else '#,##0.00'
            for col in range(1, ncols + 1):
                cc = ws.cell(row=r, column=col)
                cc.border = borde
                cc.fill = relleno

        r = tabla_start + 1
        filas_rubro = []
        for comp in COMPONENTES_FE06:
            rc = resumen[comp]
            frac_c = (rc["total_anio"] / rc["pim"]) if rc["pim"] else 0
            fila_rubro = r
            filas_rubro.append(r)
            pintar_fila(r, comp.upper(), valores_fila(rc, frac_c), "0.0%", "comp")
            ws.row_dimensions[r].height = 18
            r += 1

            # Filas de detalle (sub-clasificadores del componente)
            det_ini = r
            det_fin = r - 1
            for sub in (x for x in rows if x["componente"] == comp):
                ca = ws.cell(row=r, column=1, value=sub["clasificador"])
                ca.font = font_td
                ca.alignment = align_left
                cb = ws.cell(row=r, column=2, value=sub["detalle"])
                cb.font = font_td
                cb.alignment = align_left
                frac_s = (sub["total_anio"] / sub["pim"]) if sub["pim"] else 0
                for j, v in enumerate(valores_fila(sub, frac_s), 3):
                    cc = ws.cell(row=r, column=j, value=v)
                    cc.font = font_td
                    cc.alignment = align_right_num
                    cc.number_format = '0.0%' if j == col_pct else '#,##0.00'
                for col in range(1, ncols + 1):
                    ws.cell(row=r, column=col).border = border_grid
                # Auto-ajustar alto segun Detalle (col 2)
                det_text = sub["detalle"] or ""
                det_w = anchos.get(2, 25)
                h_det = _xl_row_height(det_text, det_w, 9)
                ws.row_dimensions[r].height = max(16, h_det)
                det_fin = r
                r += 1

            # El subtotal del rubro se calcula con SUM sobre sus detalles
            if det_fin >= det_ini:
                letra_pim = get_column_letter(col_pim)
                letra_ejec = get_column_letter(col_ejec)
                for col in range(3, ncols + 1):
                    if col == col_pct:
                        continue
                    letra = get_column_letter(col)
                    ws.cell(row=fila_rubro, column=col,
                            value=f"=SUM({letra}{det_ini}:{letra}{det_fin})")
                cpct = ws.cell(row=fila_rubro, column=col_pct)
                cpct.value = (f"=IF({letra_pim}{fila_rubro}=0,0,"
                              f"{letra_ejec}{fila_rubro}/{letra_pim}{fila_rubro})")
                cpct.number_format = '0.0%'

        # Fila final (.fila-total): sumatoria de todos los componentes
        def suma(key):
            return round(sum(v[key] for v in resumen.values()), 2)

        tot_pim = suma("pim")
        tot_anio = suma("total_anio")
        tot_reg = {
            "et": suma("et"),
            "e2023": suma("e2023"), "e2024": suma("e2024"),
            "e2025": suma("e2025"),
            "pim": tot_pim,
            "mensual": fe06_totales_mensual,
            "total_anio": tot_anio,
            "saldo_pim": suma("saldo_pim"),
            "acum_total": suma("acum_total"),
            "saldo_et": suma("saldo_et"),
        }
        frac_t = (tot_anio / tot_pim) if tot_pim else 0
        total_row = r
        pintar_fila(r, "TOTAL RUBRO 18", valores_fila(tot_reg, frac_t),
                    "0.00%", "total")
        letra_pim = get_column_letter(col_pim)
        letra_ejec = get_column_letter(col_ejec)
        for col in range(3, ncols + 1):
            if col == col_pct:
                continue
            letra = get_column_letter(col)
            refs = ",".join(f"{letra}{fr}" for fr in filas_rubro)
            ws.cell(row=total_row, column=col).value = f"=SUM({refs})"
        cpct = ws.cell(row=total_row, column=col_pct)
        cpct.value = (f"=IF({letra_pim}{total_row}=0,0,"
                      f"{letra_ejec}{total_row}/{letra_pim}{total_row})")
        cpct.number_format = '0.00%'
        ws.row_dimensions[r].height = 20
        r += 1

        # ── RESUMENES lado a lado: PIM (A-F) | EXPEDIENTE (H-L) ──
        r += 1  # spacer

        # Columnas para cada cuadro
        col_pim_end = 6      # F
        col_exp_ini = 8      # H
        col_exp_end = ncols  # L (12)
        col_pim_lbl = get_column_letter(col_pim_end - 1)  # E
        col_pim_val = get_column_letter(col_pim_end)       # F
        col_exp_lbl = get_column_letter(col_exp_ini + 1)   # I
        col_exp_val = get_column_letter(col_exp_ini + 2)   # J

        # ── Headers ──
        # PIM header: A:E merged
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_pim_end - 1)
        c = ws.cell(row=r, column=1, value="RESUMEN A NIVEL PIM")
        c.font = font_th; c.alignment = align_left
        for col in range(1, col_pim_end):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill_azul; cc.border = border_all
        # EXPEDIENTE header: H:L merged
        ws.merge_cells(start_row=r, start_column=col_exp_ini, end_row=r, end_column=col_exp_end)
        c = ws.cell(row=r, column=col_exp_ini, value="RESUMEN A NIVEL DE EXPEDIENTE")
        c.font = font_th; c.alignment = align_left
        for col in range(col_exp_ini, col_exp_end + 1):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill_azul; cc.border = border_all
        ws.row_dimensions[r].height = 18
        r += 1

        # ── Sub-headers ──
        # PIM: DESCRIPCION A:C, % D, S/ E
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(row=r, column=1, value="DESCRIPCION")
        c.font = font_label; c.alignment = align_left
        c3 = ws.cell(row=r, column=4, value="%")
        c3.font = font_label; c3.alignment = align_center
        c2 = ws.cell(row=r, column=5, value="S/.")
        c2.font = font_label; c2.alignment = align_right_num
        for col in range(1, col_pim_end + 1):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill_bg; cc.border = border_grid

        # EXPEDIENTE: DESCRIPCION H:K, S/ L
        ws.merge_cells(start_row=r, start_column=col_exp_ini, end_row=r, end_column=col_exp_ini + 3)
        c = ws.cell(row=r, column=col_exp_ini, value="DESCRIPCION")
        c.font = font_label; c.alignment = align_left
        c2 = ws.cell(row=r, column=col_exp_ini + 4, value="S/.")
        c2.font = font_label; c2.alignment = align_right_num
        for col in range(col_exp_ini, col_exp_ini + 5):
            cc = ws.cell(row=r, column=col)
            cc.fill = fill_bg; cc.border = border_grid
        ws.row_dimensions[r].height = 18
        r += 1

        # ── Data rows ──
        pim_v = sintesis["pim"]
        def pct_de(num):
            return (num / pim_v) if pim_v else 0

        filas_pim = [
            (f"PIM ASIGNADO {anio}", pim_v, 1.0, "0%", True),
            (f"GASTO EJECUTADO {anio}",
             f"={get_column_letter(col_ejec)}{total_row}",
             f"=IF(E{r}=0,0,E{r+1}/E{r})", "0.00%", False),
            (f"GASTO FINANCIERO MES DE {MESES[mes-1]}", sintesis["mes_actual"],
             pct_de(sintesis["mes_actual"]), "0.00%", False),
            (f"SALDO {anio}", sintesis["saldo_pim"],
             pct_de(sintesis["saldo_pim"]), "0.00%", True),
        ]

        filas_exp = [("PRESUPUESTO TOTAL DE OBRA", sintesis["et"], True)]
        if incluir_anios:
            filas_exp += [("GASTO EJECUTADO 2023", sintesis["e2023"], False),
                          ("GASTO EJECUTADO 2024", sintesis["e2024"], False),
                          ("GASTO EJECUTADO 2025", sintesis["e2025"], False)]
            filas_exp.append((f"GASTO EJECUTADO AL MES DE {MESES[mes-1]} {anio}",
                              sintesis["ejec_anio"], False))
        else:
            filas_exp.append((f"GASTO EJECUTADO {anio}",
                              sintesis["ejec_anio"], False))
        if incluir_anios:
            filas_exp.append((f"SALDO {anio}", sintesis["saldo_proyecto"], True))
        else:
            filas_exp.append(("SALDO A NIVEL DE EXPEDIENTE TECNICO",
                              sintesis["saldo_proyecto"], True))

        ini_pim = r
        max_rows = max(len(filas_pim), len(filas_exp))
        for idx in range(max_rows):
            # PIM side (cols A-E): desc A:C, % D, S/ E
            if idx < len(filas_pim):
                desc, sval, fracc, fmt, strong = filas_pim[idx]
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                c = ws.cell(row=r, column=1, value=desc)
                c.font = font_td_b if strong else font_td
                c.alignment = align_left
                cp = ws.cell(row=r, column=4, value=fracc)
                cp.font = font_td; cp.alignment = align_right_num
                cp.number_format = fmt
                cs = ws.cell(row=r, column=5, value=sval)
                cs.font = font_td_b if strong else font_td
                cs.alignment = align_right_num
                cs.number_format = '#,##0.00'
                for col in range(1, col_pim_end + 1):
                    ws.cell(row=r, column=col).border = border_grid

            # EXPEDIENTE side (cols H-L): desc H:K, S/ L
            if idx < len(filas_exp):
                desc, sval, strong = filas_exp[idx]
                ws.merge_cells(start_row=r, start_column=col_exp_ini,
                               end_row=r, end_column=col_exp_ini + 3)
                c = ws.cell(row=r, column=col_exp_ini, value=desc)
                c.font = font_td_b if strong else font_td
                c.alignment = align_left
                cs = ws.cell(row=r, column=col_exp_ini + 4, value=sval)
                cs.font = font_td_b if strong else font_td
                cs.alignment = align_right_num
                cs.number_format = '#,##0.00'
                for col in range(col_exp_ini, col_exp_ini + 5):
                    ws.cell(row=r, column=col).border = border_grid

            ws.row_dimensions[r].height = 16
            r += 1

        # Fix PIM percentage formula (needs final row refs)
        # Re-read filas_pim formula with correct row references
        pim_data_start = ini_pim
        pim_data_end = ini_pim + len(filas_pim) - 1
        # GASTO EJECUTADO row is ini_pim + 1
        ejec_pim_row = ini_pim + 1
        ws.cell(row=ejec_pim_row, column=4).value = (
            f"=IF(E{ini_pim}=0,0,E{ejec_pim_row}/E{ini_pim})")

        # ── Ajustes finales ──
        ws.print_area = f"A1:{get_column_letter(ncols)}{r}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"FE06_Presupuesto_{MESES[mes-1]}_{anio}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/formatos/panel/imprimir")
    def imprimir_panel():
        """Vista imprimible del Resumen Financiero: replica de la hoja PANEL
        del informe financiero con todos los cuadros y calculos del aplicativo."""
        p = get_proyecto()
        if Presupuesto.query.count() == 0 and Gasto.query.count() == 0:
            flash("Aún no hay datos registrados para generar el Resumen "
                  "Financiero. Ingrese la configuración presupuestal y los "
                  "gastos del proyecto.", "warning")
            return redirect(url_for("formatos"))
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        datos = panel_datos(mes, anio)
        return render_template("panel_imprimir.html", p=p, MESES=MESES, d=datos)

    @app.route("/almacen/fe07/imprimir")
    def imprimir_fe07():
        """Vista imprimible del FE-07 Movimiento Diario de Almacén."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        diario = almacen_diario(mes, anio)
        material = request.args.get("material", "").strip()
        if material:
            diario = dict(diario)
            diario["items"] = [i for i in diario["items"] if i["descripcion"] == material]
        return render_template("fe07_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, diario=diario)

    @app.route("/almacen/fe08/imprimir")
    def imprimir_fe08():
        """Vista imprimible del FE-08 Almacén Valorizado."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        valorizado = almacen_valorizado(mes, anio)
        valor_total = {
            "cant_in": round(sum(x["cant_in"] for x in valorizado), 2),
            "valor_in": round(sum(x["valor_in"] for x in valorizado), 2),
            "cant_out": round(sum(x["cant_out"] for x in valorizado), 2),
            "valor_out": round(sum(x["valor_out"] for x in valorizado), 2),
            "saldo": round(sum(x["saldo"] for x in valorizado), 2),
            "valor_saldo": round(sum(x["valor_saldo"] for x in valorizado), 2),
        }
        return render_template("fe08_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, valorizado=valorizado,
                               valor_total=valor_total)

    @app.route("/almacen/kardex")
    def kardex():
        """Vista de Kardex por material con saldo corrido."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        filtro = request.args.get("material", "").strip()
        movs = (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.mes == mes, AlmacenMovimiento.anio == anio)
                .order_by(AlmacenMovimiento.fecha, AlmacenMovimiento.id).all())
        running = {}
        orden = []
        for m in movs:
            if filtro and m.descripcion != filtro:
                continue
            key = (m.descripcion, m.und)
            if key not in running:
                running[key] = {"balance": 0.0, "rows": []}
                orden.append(key)
            bal = running[key]["balance"]
            nuevo = bal + (m.cantidad or 0) if m.tipo == "E" else bal - (m.cantidad or 0)
            running[key]["rows"].append({"mov": m, "saldo_anterior": round(bal, 2),
                                         "saldo": round(nuevo, 2),
                                         "importe": round((m.cantidad or 0) * (m.precio_unitario or 0), 2)})
            running[key]["balance"] = nuevo
        items = []
        for key in orden:
            info = running[key]
            rows = info["rows"]
            ent = sum(r["mov"].cantidad or 0 for r in rows if r["mov"].tipo == "E")
            sal = sum(r["mov"].cantidad or 0 for r in rows if r["mov"].tipo == "S")
            val_ent = sum(r["importe"] for r in rows if r["mov"].tipo == "E")
            val_sal = sum((r["mov"].cantidad or 0) * (r["mov"].precio_unitario or 0) for r in rows if r["mov"].tipo == "S")
            items.append({
                "descripcion": key[0], "und": key[1], "movs": len(rows),
                "cant_in": round(ent, 2), "cant_out": round(sal, 2),
                "saldo": round(ent - sal, 2),
                "valor_in": round(val_ent, 2),
                "valor_out": round(val_sal, 2),
                "valor_saldo": round(val_ent - val_sal, 2),
                "rows": rows,
            })
        return render_template("kardex.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, items=items, filtro=filtro)

    @app.route("/almacen/inventario")
    def inventario():
        """Vista de Reporte de Inventario consolidado por material."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        movs = (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.mes == mes, AlmacenMovimiento.anio == anio)
                .order_by(AlmacenMovimiento.fecha, AlmacenMovimiento.id).all())
        inv = {}
        orden = []
        for m in movs:
            key = m.descripcion
            if not key:
                continue
            if key not in inv:
                inv[key] = {"und": m.und, "cant_in": 0.0, "cant_out": 0.0,
                            "valor_in": 0.0, "valor_out": 0.0, "movs": 0,
                            "ult_entrada": None, "ult_precio": 0.0}
                orden.append(key)
            info = inv[key]
            info["movs"] += 1
            if m.tipo == "E":
                info["cant_in"] += m.cantidad or 0
                info["valor_in"] += (m.cantidad or 0) * (m.precio_unitario or 0)
                info["ult_entrada"] = m.fecha
                info["ult_precio"] = m.precio_unitario or 0
            else:
                info["cant_out"] += m.cantidad or 0
                info["valor_out"] += (m.cantidad or 0) * (m.precio_unitario or 0)
        items = []
        for key in orden:
            info = inv[key]
            saldo = round(info["cant_in"] - info["cant_out"], 2)
            valor_saldo = round(info["valor_in"] - info["valor_out"], 2)
            items.append({
                "descripcion": key,
                "und": info["und"],
                "movs": info["movs"],
                "cant_in": round(info["cant_in"], 2),
                "cant_out": round(info["cant_out"], 2),
                "saldo": saldo,
                "pu": round(info["ult_precio"], 2) if saldo > 0 else 0.0,
                "valor_in": round(info["valor_in"], 2),
                "valor_out": round(info["valor_out"], 2),
                "valor_saldo": valor_saldo,
                "ult_entrada": info["ult_entrada"],
            })
        items.sort(key=lambda x: x["descripcion"])
        totals = {
            "cant_in": round(sum(i["cant_in"] for i in items), 2),
            "valor_in": round(sum(i["valor_in"] for i in items), 2),
            "cant_out": round(sum(i["cant_out"] for i in items), 2),
            "valor_out": round(sum(i["valor_out"] for i in items), 2),
            "saldo": round(sum(i["saldo"] for i in items), 2),
            "valor_saldo": round(sum(i["valor_saldo"] for i in items), 2),
        }
        return render_template("inventario.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, items=items, totals=totals, p_items=len(items))

    @app.route("/almacen/kardex/imprimir")
    def imprimir_kardex():
        """Vista imprimible del Kardex por material."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        filtro = request.args.get("material", "").strip()
        movs = (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.mes == mes, AlmacenMovimiento.anio == anio)
                .order_by(AlmacenMovimiento.fecha, AlmacenMovimiento.id).all())
        running = {}
        orden = []
        for m in movs:
            if filtro and m.descripcion != filtro:
                continue
            key = (m.descripcion, m.und)
            if key not in running:
                running[key] = {"balance": 0.0, "rows": []}
                orden.append(key)
            bal = running[key]["balance"]
            nuevo = bal + (m.cantidad or 0) if m.tipo == "E" else bal - (m.cantidad or 0)
            running[key]["rows"].append({"mov": m, "saldo_anterior": round(bal, 2),
                                         "saldo": round(nuevo, 2)})
            running[key]["balance"] = nuevo
        items = []
        for key in orden:
            info = running[key]
            rows = info["rows"]
            ent = sum(r["mov"].cantidad or 0 for r in rows if r["mov"].tipo == "E")
            sal = sum(r["mov"].cantidad or 0 for r in rows if r["mov"].tipo == "S")
            items.append({
                "descripcion": key[0], "und": key[1], "movs": len(rows),
                "cant_in": round(ent, 2), "cant_out": round(sal, 2),
                "saldo": round(ent - sal, 2), "rows": rows,
            })
        return render_template("kardex_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, items=items, filtro=filtro)

    @app.route("/almacen/inventario/imprimir")
    def imprimir_inventario():
        """Vista imprimible del Reporte de Inventario."""
        p = get_proyecto()
        mes = param_int("mes", p.mes_actual, lo=1, hi=12)
        anio = p.anio
        movs = (AlmacenMovimiento.query
                .filter(AlmacenMovimiento.mes == mes, AlmacenMovimiento.anio == anio)
                .order_by(AlmacenMovimiento.fecha, AlmacenMovimiento.id).all())
        inv = {}
        orden = []
        for m in movs:
            key = m.descripcion
            if not key:
                continue
            if key not in inv:
                inv[key] = {"und": m.und, "cant_in": 0.0, "cant_out": 0.0,
                            "valor_in": 0.0, "valor_out": 0.0, "movs": 0,
                            "ult_entrada": None, "ult_precio": 0.0}
                orden.append(key)
            info = inv[key]
            info["movs"] += 1
            if m.tipo == "E":
                info["cant_in"] += m.cantidad or 0
                info["valor_in"] += (m.cantidad or 0) * (m.precio_unitario or 0)
                info["ult_entrada"] = m.fecha
                info["ult_precio"] = m.precio_unitario or 0
            else:
                info["cant_out"] += m.cantidad or 0
                info["valor_out"] += (m.cantidad or 0) * (m.precio_unitario or 0)
        items = []
        for key in orden:
            info = inv[key]
            saldo = round(info["cant_in"] - info["cant_out"], 2)
            valor_saldo = round(info["valor_in"] - info["valor_out"], 2)
            items.append({
                "descripcion": key,
                "und": info["und"],
                "movs": info["movs"],
                "cant_in": round(info["cant_in"], 2),
                "cant_out": round(info["cant_out"], 2),
                "saldo": saldo,
                "pu": round(info["ult_precio"], 2) if saldo > 0 else 0.0,
                "valor_in": round(info["valor_in"], 2),
                "valor_out": round(info["valor_out"], 2),
                "valor_saldo": valor_saldo,
                "ult_entrada": info["ult_entrada"],
            })
        items.sort(key=lambda x: x["descripcion"])
        totals = {
            "cant_in": round(sum(i["cant_in"] for i in items), 2),
            "valor_in": round(sum(i["valor_in"] for i in items), 2),
            "cant_out": round(sum(i["cant_out"] for i in items), 2),
            "valor_out": round(sum(i["valor_out"] for i in items), 2),
            "saldo": round(sum(i["saldo"] for i in items), 2),
            "valor_saldo": round(sum(i["valor_saldo"] for i in items), 2),
        }
        return render_template("inventario_imprimir.html", p=p, mes=mes, anio=anio,
                               MESES=MESES, items=items, totals=totals)

    @app.route("/api/resumen")
    def api_resumen():
        p = get_proyecto()
        return jsonify({
            "por_mes": ejecucion_por_mes(p.anio),
            "por_componente": ejecucion_por_componente(p.anio),
            "meses": MESES,
            "kpis": kpis(),
        })


app = create_app()


if __name__ == "__main__":
    # Por defecto escucha en la red local (0.0.0.0). Para modo solo-local:
    #   set HOST=127.0.0.1  (o inicie con iniciar_local.bat)
    host = os.environ.get("HOST", "0.0.0.0")
    try:
        port = int(os.environ.get("PORT", "5000"))
    except ValueError:
        port = 5000
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug, host=host, port=port)

